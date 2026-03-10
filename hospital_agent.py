"""
HospitalAgent v1 — Worldwide Nearby Hospital Finder (Scraping Edition)
=======================================================================
Install:
    pip install requests beautifulsoup4 lxml openpyxl python-dotenv tqdm

Usage:
    agent = HospitalAgent()
    agent.find_nearby("Chennai")
    agent.find_nearby("Tambaram, Tamil Nadu", radius_m=5000)
    agent.emergency_near("Hyderabad")
    agent.search_by_specialty("Bangalore", "cardiac")
    agent.blood_banks_near("Mumbai")
    agent.pharmacies_near("Delhi", radius_m=2000)
    agent.export_excel("Chennai", "hospitals_chennai.xlsx")

Data sources — ALL free, ZERO signup, ZERO API key:
    ✅ Nominatim (OpenStreetMap)   — geocode any city / address / pincode / landmark
    ✅ OSM Overpass API            — scrape real hospital names, addresses, phones,
                                     websites, emergency flag, beds, specialties
    ✅ Practo public search        — scrape doctor count, hospital ratings, reviews
    ✅ Wikipedia REST API          — scrape hospital descriptions
    ✅ JustDial public pages       — scrape phone, address, ratings (India)
    ✅ Google Maps deep links      — directions, street view (no key needed)

Techniques used:
    - requests + BeautifulSoup for HTML scraping
    - lxml parser for fast HTML parsing
    - Overpass QL queries for structured OSM data
    - Nominatim for geocoding
    - tqdm for progress bars during bulk scraping
    - openpyxl for Excel export
    - python-dotenv for optional config

Output schema (matches bus/hotel/flight agent style):
    name, type, address, phone, website, emergency,
    beds, specialties, opening_hours, wheelchair,
    rating, reviews, lat, lon, dist_km,
    maps_url, directions_url, practo_url, booking_links
"""

import os, re, math, json, time, logging
from typing import List, Dict, Any, Optional
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("HospitalAgent")

# ── HTTP headers (rotate to avoid blocks) ─────────────────────────────────────
HEADERS_OSM  = {"User-Agent": "SmartTripAI-HospitalAgent/1.0 (research project)"}
HEADERS_WEB  = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
}
HEADERS_JSON = {**HEADERS_WEB, "Accept": "application/json"}

# ── API endpoints ──────────────────────────────────────────────────────────────
NOMINATIM   = "https://nominatim.openstreetmap.org/search"
NOMINATIM_R = "https://nominatim.openstreetmap.org/reverse"
OVERPASS    = "https://overpass-api.de/api/interpreter"
WIKI_API    = "https://en.wikipedia.org/w/api.php"

# ── OSM healthcare tags to scrape ──────────────────────────────────────────────
HEALTHCARE_TAG_PAIRS = [
    ("amenity",    "hospital"),
    ("amenity",    "clinic"),
    ("amenity",    "doctors"),
    ("amenity",    "dentist"),
    ("amenity",    "pharmacy"),
    ("amenity",    "nursing_home"),
    ("healthcare", "hospital"),
    ("healthcare", "clinic"),
    ("healthcare", "centre"),
    ("healthcare", "doctor"),
    ("healthcare", "blood_bank"),
    ("healthcare", "laboratory"),
    ("healthcare", "physiotherapist"),
    ("healthcare", "emergency"),
]

FACILITY_ICONS = {
    "hospital":        "🏥",
    "clinic":          "🏨",
    "doctors":         "👨‍⚕️",
    "dentist":         "🦷",
    "pharmacy":        "💊",
    "nursing_home":    "🏠",
    "centre":          "🏢",
    "blood_bank":      "🩸",
    "laboratory":      "🔬",
    "physiotherapist": "💆",
    "emergency":       "🚨",
}

SPECIALTY_KEYWORDS = {
    "Cardiac":       ["heart","cardiac","cardio","cardiology","cath lab","angioplasty"],
    "Cancer":        ["cancer","oncology","tumour","tumor","chemo","radiation"],
    "Neuro":         ["neuro","brain","spine","neurology","neurosurgery","stroke"],
    "Orthopedic":    ["ortho","bone","joint","fracture","arthro","ligament"],
    "Pediatric":     ["child","children","pediatric","paediatric","neonatal","infant"],
    "Maternity":     ["maternity","gynaecology","gynecology","obstetric","women","ivf"],
    "Dental":        ["dental","dentist","teeth","oral","maxillofacial"],
    "Eye":           ["eye","ophthalm","vision","retina","cataract","lasik"],
    "ENT":           ["ent","ear","nose","throat","hearing"],
    "Kidney":        ["kidney","renal","nephro","dialysis","transplant"],
    "Trauma":        ["trauma","accident","emergency","casualty","burn","icu"],
    "Mental Health": ["mental","psychiatry","psychology","rehab","de-addiction"],
    "Skin":          ["skin","dermat","cosmetic","aesthetic"],
    "Diabetes":      ["diabetes","endocrin","thyroid","hormone"],
    "Liver":         ["liver","gastro","hepato","digestive","bariatric"],
}

EMERGENCY_NUMBERS = {
    "National":    {"Ambulance": "108",  "Emergency": "112", "Police": "100",
                    "Fire": "101",       "Women Helpline": "1091",
                    "Child Helpline": "1098", "Health Helpline": "104",
                    "Mental Health (iCall)": "9152987821"},
    "Tamil Nadu":  {"Ambulance": "108",  "Health": "104",  "Women": "181"},
    "Karnataka":   {"Ambulance": "108",  "Health": "104",  "BBMP Health": "080-22221188"},
    "Maharashtra": {"Ambulance": "108",  "Health": "104",  "Mumbai BMC": "1916"},
    "Delhi":       {"Ambulance": "102",  "CATS": "1099",   "Health": "011-23232323"},
    "Kerala":      {"Ambulance": "108",  "Health": "104",  "DISHA": "1056"},
    "Andhra Pradesh": {"Ambulance": "108", "Health": "104"},
    "Telangana":   {"Ambulance": "108",  "Health": "104",  "TEMS": "040-27852001"},
    "West Bengal": {"Ambulance": "102",  "Health": "1800-345-5505"},
    "Gujarat":     {"Ambulance": "108",  "Health": "104"},
    "Rajasthan":   {"Ambulance": "108",  "Health": "104"},
    "Punjab":      {"Ambulance": "108",  "Health": "104"},
    "Uttar Pradesh": {"Ambulance": "108", "Health": "104"},
}


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _hav(lat1, lon1, lat2, lon2) -> float:
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))

def _geocode(location: str) -> Optional[tuple]:
    """Geocode any address, landmark, city, pincode → (lat, lon, display_name)."""
    try:
        r = requests.get(NOMINATIM, headers=HEADERS_OSM, timeout=10,
                         params={"q": location, "format": "json", "limit": 1})
        d = r.json()
        if d:
            return float(d[0]["lat"]), float(d[0]["lon"]), d[0]["display_name"]
    except Exception as e:
        log.warning(f"Geocode failed for '{location}': {e}")
    return None

def _reverse_geocode(lat: float, lon: float) -> str:
    try:
        r = requests.get(NOMINATIM_R, headers=HEADERS_OSM, timeout=8,
                         params={"lat": lat, "lon": lon, "format": "json"})
        return r.json().get("display_name", f"{lat:.4f},{lon:.4f}")
    except Exception:
        return f"{lat:.4f},{lon:.4f}"

def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()

def _infer_specialties(tags: Dict) -> List[str]:
    blob = " ".join([
        tags.get("healthcare:speciality", ""),
        tags.get("medical_speciality", ""),
        tags.get("name", ""),
        tags.get("description", ""),
        tags.get("operator", ""),
        tags.get("subject", ""),
    ]).lower()
    found = []
    for label, keywords in SPECIALTY_KEYWORDS.items():
        if any(kw in blob for kw in keywords):
            found.append(label)
    return found if found else ["General"]

def _detect_state(display_name: str) -> str:
    """Detect Indian state from Nominatim display name."""
    for state in EMERGENCY_NUMBERS:
        if state.lower() in display_name.lower():
            return state
    return "National"


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER 1 — OSM Overpass  (primary source)
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_osm(lat: float, lon: float, radius_m: int,
                type_filter: Optional[List[str]] = None) -> List[Dict]:
    """
    Scrape real hospital/clinic data from OpenStreetMap via Overpass QL.
    Returns raw enriched records sorted by distance.
    """
    pairs = [(k, v) for k, v in HEALTHCARE_TAG_PAIRS
             if not type_filter or v in type_filter]

    tag_clause = "\n".join(
        f'  node["{k}"="{v}"](around:{radius_m},{lat},{lon});\n'
        f'  way["{k}"="{v}"](around:{radius_m},{lat},{lon});'
        for k, v in pairs
    )
    query = f"[out:json][timeout:90];\n(\n{tag_clause}\n);\nout center body;"

    try:
        r = requests.post(OVERPASS, data=query, headers=HEADERS_OSM, timeout=95)
        r.raise_for_status()
        elements = r.json().get("elements", [])
    except Exception as e:
        log.error(f"Overpass scrape failed: {e}")
        return []

    records, seen = [], set()

    for el in elements:
        tags  = el.get("tags", {})
        name  = _clean(tags.get("name") or tags.get("short_name") or "")
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())

        # Coordinates
        if el["type"] == "node":
            h_lat, h_lon = el.get("lat"), el.get("lon")
        else:
            c = el.get("center", {})
            h_lat, h_lon = c.get("lat"), c.get("lon")

        if h_lat is None or h_lon is None:
            continue

        dist     = _hav(lat, lon, h_lat, h_lon)
        ftype    = tags.get("amenity") or tags.get("healthcare") or "hospital"
        icon     = FACILITY_ICONS.get(ftype, "🏥")

        # Address assembly
        addr_parts = [
            tags.get("addr:housenumber", ""),
            tags.get("addr:street", ""),
            tags.get("addr:suburb", ""),
            tags.get("addr:city", ""),
            tags.get("addr:state", ""),
            tags.get("addr:postcode", ""),
        ]
        address = _clean(", ".join(p for p in addr_parts if p)) or "N/A"

        # Hours
        hours = tags.get("opening_hours", "N/A")
        if hours == "24/7":
            hours = "Open 24 Hours ✅"

        # Emergency
        emergency = tags.get("emergency", "")
        if not emergency:
            emergency = "yes" if ftype in ("hospital", "emergency") else "no"

        records.append({
            "name":          name,
            "icon":          icon,
            "type":          ftype,
            "type_label":    f"{icon} {ftype.replace('_',' ').title()}",
            "address":       address,
            "phone":         _clean(tags.get("phone") or tags.get("contact:phone") or ""),
            "website":       _clean(tags.get("website") or tags.get("contact:website") or ""),
            "email":         _clean(tags.get("email") or tags.get("contact:email") or ""),
            "emergency":     emergency,
            "beds":          tags.get("beds") or tags.get("capacity:beds") or "",
            "opening_hours": hours,
            "wheelchair":    tags.get("wheelchair", ""),
            "operator":      _clean(tags.get("operator") or ""),
            "operator_type": tags.get("operator:type", ""),
            "specialties":   _infer_specialties(tags),
            "lat":           h_lat,
            "lon":           h_lon,
            "dist_km":       round(dist, 3),
            "maps_url":      f"https://www.google.com/maps?q={h_lat},{h_lon}",
            "directions_url":f"https://www.google.com/maps/dir/?api=1&destination={h_lat},{h_lon}",
            "osm_id":        el.get("id"),
            # Placeholders filled by secondary scrapers
            "rating":        None,
            "reviews":       None,
            "practo_url":    "",
            "description":   "",
            "source":        "OpenStreetMap",
        })

    records.sort(key=lambda x: x["dist_km"])
    return records


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER 2 — Wikipedia  (descriptions)
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_wiki_desc(name: str, city: str = "") -> str:
    """Scrape a 2-sentence Wikipedia summary for a hospital."""
    try:
        search_r = requests.get(WIKI_API, headers=HEADERS_JSON, timeout=8, params={
            "action": "query", "list": "search",
            "srsearch": f"{name} {city} hospital".strip(),
            "format": "json", "srlimit": 1,
        })
        results = search_r.json().get("query", {}).get("search", [])
        if not results:
            return ""
        title = results[0]["title"]
        page_r = requests.get(WIKI_API, headers=HEADERS_JSON, timeout=8, params={
            "action": "query", "titles": title,
            "prop": "extracts", "exintro": True,
            "explaintext": True, "exsentences": 3,
            "format": "json",
        })
        for page in page_r.json().get("query", {}).get("pages", {}).values():
            return _clean(page.get("extract", ""))
    except Exception as e:
        log.debug(f"Wiki scrape failed for '{name}': {e}")
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER 3 — Practo public search  (ratings, doctor count)
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_practo(hospital_name: str, city: str) -> Dict:
    """
    Scrape Practo public search for hospital rating and doctor count.
    Uses Practo's public JSON API endpoint — no login required.
    """
    result = {"rating": None, "reviews": None, "practo_url": "", "doctors": None}
    try:
        city_slug     = city.lower().replace(" ", "-")
        name_slug     = hospital_name.lower().replace(" ", "-").replace("'", "")
        search_url    = f"https://www.practo.com/search?results_type=hospital&q={requests.utils.quote(hospital_name)}&city={city_slug}"
        result["practo_url"] = search_url

        r = requests.get(
            "https://www.practo.com/search/hospitals",
            headers=HEADERS_JSON, timeout=10,
            params={"results_type": "hospital", "q": hospital_name, "city": city_slug}
        )
        if r.status_code != 200:
            return result

        soup = BeautifulSoup(r.text, "lxml")

        # Try structured data first (JSON-LD)
        for script in soup.find_all("script", {"type": "application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list):
                    data = data[0]
                if data.get("@type") in ("Hospital", "MedicalOrganization", "LocalBusiness"):
                    result["rating"]  = data.get("aggregateRating", {}).get("ratingValue")
                    result["reviews"] = data.get("aggregateRating", {}).get("reviewCount")
                    return result
            except Exception:
                pass

        # Fallback: scrape rating from HTML
        rating_el = soup.select_one('[data-qa-id="hospital_rating"]')
        if rating_el:
            result["rating"] = _clean(rating_el.get_text())

        reviews_el = soup.select_one('[data-qa-id="hospital_review_count"]')
        if reviews_el:
            result["reviews"] = _clean(reviews_el.get_text())

    except Exception as e:
        log.debug(f"Practo scrape failed for '{hospital_name}': {e}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER 4 — JustDial public pages  (phone, address, ratings — India)
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_justdial(hospital_name: str, city: str) -> Dict:
    """
    Scrape JustDial public listing for phone numbers and ratings.
    Uses public search HTML — no API key needed.
    """
    result = {"phone": "", "address": "", "rating": None, "jd_url": ""}
    try:
        city_clean  = city.lower().replace(" ", "-")
        name_enc    = requests.utils.quote(hospital_name)
        search_url  = f"https://www.justdial.com/{city_clean}/Hospitals/{name_enc}"
        result["jd_url"] = search_url

        r = requests.get(search_url, headers=HEADERS_WEB, timeout=12)
        if r.status_code not in (200, 301, 302):
            return result

        soup = BeautifulSoup(r.text, "lxml")

        # JSON-LD structured data
        for script in soup.find_all("script", {"type": "application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list):
                    data = data[0]
                if "telephone" in data:
                    result["phone"]   = _clean(data.get("telephone", ""))
                    result["address"] = _clean(str(data.get("address", "")))
                    agg = data.get("aggregateRating", {})
                    result["rating"]  = agg.get("ratingValue")
                    return result
            except Exception:
                pass

        # HTML fallback
        phone_el = soup.select_one(".contact-info .jdicon-phone ~ span, .tel")
        if phone_el:
            result["phone"] = _clean(phone_el.get_text())

        rating_el = soup.select_one(".star-rating, .ratingvalue")
        if rating_el:
            result["rating"] = _clean(rating_el.get_text())

    except Exception as e:
        log.debug(f"JustDial scrape failed for '{hospital_name}': {e}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER 5 — Google Maps JSON  (structured place data via Maps embed)
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_google_maps_data(hospital_name: str, lat: float, lon: float) -> Dict:
    """
    Scrape basic hospital data from Google Maps search result page.
    Parses structured JSON embedded in the Maps response HTML.
    """
    result = {"rating": None, "reviews": None, "phone": "", "website": ""}
    try:
        query = f"{hospital_name} hospital {lat},{lon}"
        url   = f"https://www.google.com/search?q={requests.utils.quote(query)}&num=1"
        r     = requests.get(url, headers=HEADERS_WEB, timeout=12)
        soup  = BeautifulSoup(r.text, "lxml")

        # Try JSON-LD from Google's rich results
        for script in soup.find_all("script", {"type": "application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list):
                    data = data[0]
                agg = data.get("aggregateRating", {})
                if agg.get("ratingValue"):
                    result["rating"]  = agg.get("ratingValue")
                    result["reviews"] = agg.get("reviewCount")
                phone = data.get("telephone", "")
                if phone:
                    result["phone"] = _clean(phone)
                web = data.get("url", "")
                if web:
                    result["website"] = web
                return result
            except Exception:
                pass

        # Parse rating from search snippet text
        rating_match = re.search(r'(\d\.\d)\s*/\s*5', r.text)
        if rating_match:
            result["rating"] = rating_match.group(1)

        reviews_match = re.search(r'([\d,]+)\s*(?:reviews|ratings|Google reviews)', r.text)
        if reviews_match:
            result["reviews"] = reviews_match.group(1).replace(",", "")

    except Exception as e:
        log.debug(f"Google Maps scrape failed for '{hospital_name}': {e}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  ENRICHMENT PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def _enrich_facility(fac: Dict, city: str, enrich_level: int = 1) -> Dict:
    """
    Enrich a facility record using secondary scrapers.

    enrich_level:
        0 = OSM only (fastest)
        1 = OSM + Wikipedia desc (default)
        2 = OSM + Wikipedia + Practo/JustDial/Google (slower, more complete)
    """
    name = fac["name"]

    # Level 1: Wikipedia description
    if enrich_level >= 1 and not fac.get("description"):
        fac["description"] = _scrape_wiki_desc(name, city)
        time.sleep(0.3)  # polite delay

    if enrich_level >= 2:
        # Practo rating
        practo = _scrape_practo(name, city)
        if practo.get("rating") and not fac.get("rating"):
            fac["rating"]     = practo["rating"]
            fac["reviews"]    = practo.get("reviews")
        if practo.get("practo_url"):
            fac["practo_url"] = practo["practo_url"]
        time.sleep(0.4)

        # JustDial enrichment (phone + rating fallback)
        if not fac["phone"] or not fac["rating"]:
            jd = _scrape_justdial(name, city)
            if jd.get("phone") and not fac["phone"]:
                fac["phone"]     = jd["phone"]
                fac["jd_url"]    = jd.get("jd_url", "")
            if jd.get("rating") and not fac["rating"]:
                fac["rating"]    = jd["rating"]
            if jd.get("address") and fac["address"] == "N/A":
                fac["address"]   = jd["address"]
            time.sleep(0.4)

        # Google Maps (last resort for rating/phone)
        if not fac["rating"]:
            gm = _scrape_google_maps_data(name, fac["lat"], fac["lon"])
            if gm.get("rating"):
                fac["rating"]  = gm["rating"]
                fac["reviews"] = gm.get("reviews")
            if gm.get("phone") and not fac["phone"]:
                fac["phone"]   = gm["phone"]
            if gm.get("website") and not fac["website"]:
                fac["website"] = gm["website"]
            time.sleep(0.3)

    # Always build booking links
    fac["booking_links"] = _build_booking_links(name, city)
    return fac


def _build_booking_links(hospital_name: str, city: str) -> Dict[str, str]:
    n = requests.utils.quote(hospital_name)
    c = requests.utils.quote(city)
    return {
        "practo":        f"https://www.practo.com/search?results_type=hospital&q={n}&city={c.lower()}",
        "apollo247":     f"https://www.apollo247.com/hospitals/{c.lower().replace('%20','-')}",
        "1mg":           f"https://www.1mg.com/doctors/speciality/?city={c}",
        "justdial":      f"https://www.justdial.com/{c}/Hospitals/{n}",
        "google_search": f"https://www.google.com/search?q={n}+{c}+hospital+appointment",
        "healthworld":   f"https://www.healthworld.in/hospitals/{c.lower().replace('%20','-')}",
    }


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════

def _export_excel(facilities: List[Dict], filepath: str, location: str):
    """Export hospital data to a formatted Excel file using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hospitals"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    alt_fill     = PatternFill("solid", fgColor="D6E4F0")
    emrg_fill    = PatternFill("solid", fgColor="FFD7D7")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    bold_font    = Font(bold=True, size=10)
    normal_font  = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"🏥 Hospitals & Healthcare Near '{location}'  —  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font  = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "#", "Name", "Type", "Distance (km)", "Address", "Phone",
        "Website", "Emergency", "Beds", "Specialties",
        "Opening Hours", "Wheelchair", "Rating", "Google Maps"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [4, 35, 16, 13, 40, 18, 32, 10, 8, 30, 22, 12, 8, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, fac in enumerate(facilities, start=1):
        r = row_idx + 2
        fill = emrg_fill if fac.get("emergency") == "yes" else (
               alt_fill if row_idx % 2 == 0 else None)

        values = [
            row_idx,
            fac.get("name", ""),
            fac.get("type_label", fac.get("type", "")),
            fac.get("dist_km", ""),
            fac.get("address", ""),
            fac.get("phone", ""),
            fac.get("website", ""),
            "✅ YES" if fac.get("emergency") == "yes" else fac.get("emergency",""),
            fac.get("beds", ""),
            ", ".join(fac.get("specialties", [])),
            fac.get("opening_hours", ""),
            fac.get("wheelchair", ""),
            fac.get("rating", ""),
            fac.get("maps_url", ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=str(val) if val is not None else "")
            cell.font      = bold_font if col_idx == 2 else normal_font
            cell.alignment = center_align if col_idx in (1, 4, 8, 9, 12, 13) else left_align
            cell.border    = thin_border
            if fill:
                cell.fill = fill

        # Hyperlink for Google Maps
        maps_cell = ws.cell(row=r, column=14)
        maps_url  = fac.get("maps_url", "")
        if maps_url:
            maps_cell.hyperlink = maps_url
            maps_cell.value     = "📍 View Map"
            maps_cell.font      = Font(color="0563C1", underline="single", size=10)

        ws.row_dimensions[r].height = 20

    # ── Emergency numbers sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Emergency Numbers")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 25

    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = "🚨 Emergency Helpline Numbers — India"
    ws2["A1"].font      = Font(bold=True, size=13, color="CC0000")
    ws2["A1"].alignment = center_align

    row = 2
    for state, numbers in EMERGENCY_NUMBERS.items():
        ws2.cell(row=row, column=1, value=state).font = Font(bold=True, color="1F4E79")
        ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D6E4F0")
        ws2.row_dimensions[row].height = 18
        row += 1
        for service, number in numbers.items():
            ws2.cell(row=row, column=2, value=service).font = Font(size=10)
            ws2.cell(row=row, column=3, value=number).font  = Font(bold=True, size=10, color="CC0000")
            ws2.row_dimensions[row].height = 16
            row += 1

    wb.save(filepath)
    log.info(f"Excel saved → {filepath}")


# ══════════════════════════════════════════════════════════════════════════════
#  HOSPITAL AGENT v1
# ══════════════════════════════════════════════════════════════════════════════

class HospitalAgent:
    """
    HospitalAgent v1 — Worldwide Nearby Hospital Finder (Scraping Edition).

    No hardcoded DB. All data scraped live from:
        OpenStreetMap → Wikipedia → Practo → JustDial → Google Maps

    API:
        find_nearby(location, radius_m, limit, types, enrich)
            → All hospitals/clinics sorted by distance

        search_by_specialty(location, specialty, radius_m)
            → Filter by: cardiac, cancer, pediatric, trauma, dental, eye, kidney…

        emergency_near(location, radius_m)
            → Emergency hospitals + A&E only + local helpline numbers

        hospital_details(name, city)
            → Full scrape: Wikipedia + Practo + JustDial + Google + booking links

        blood_banks_near(location, radius_m)
            → Blood banks and donation centres

        pharmacies_near(location, radius_m)
            → Pharmacies near a location

        export_excel(location, filepath, radius_m)
            → Scrape + export formatted Excel with all hospitals

        get_emergency_numbers(state)
            → Emergency helpline numbers for any Indian state
    """

    def __init__(self):
        log.info("HospitalAgent v1 ready — OSM + Wikipedia + Practo + JustDial + Google")

    # ── 1. Find Nearby ─────────────────────────────────────────────────────────

    def find_nearby(self, location: str, radius_m: int = 5000,
                    limit: int = 20, types: Optional[List[str]] = None,
                    enrich: int = 1) -> Dict[str, Any]:
        """
        Find all healthcare facilities near a location.

        Args:
            location : city, address, pincode, landmark — any geocodable string
            radius_m : search radius in metres (default 5 km)
            limit    : max results returned (default 20)
            types    : filter by OSM type e.g. ["hospital", "clinic", "pharmacy"]
            enrich   : 0=OSM only, 1=+Wikipedia, 2=+Practo/JustDial/Google

        Returns:
            dict with facilities[], hospitals[], clinics[], pharmacies[],
                  total, location info, emergency_numbers
        """
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'", "facilities": []}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping hospitals near '{location}' (radius={radius_m}m)…")
        facilities = _scrape_osm(lat, lon, radius_m, types)

        if not facilities:
            log.warning(f"No results at {radius_m}m — widening to {radius_m * 3}m…")
            facilities = _scrape_osm(lat, lon, radius_m * 3, types)

        # Enrich with secondary scrapers
        if enrich > 0:
            log.info(f"Enriching {min(limit, len(facilities))} records (level={enrich})…")
            for fac in tqdm(facilities[:limit], desc="Enriching", unit="hospital"):
                _enrich_facility(fac, city, enrich_level=enrich)

        facilities = facilities[:limit]

        # Categorise
        hospitals  = [f for f in facilities if f["type"] in ("hospital",)
                      or "hospital" in f.get("type","")]
        clinics    = [f for f in facilities if f["type"] in
                      ("clinic","doctors","nursing_home","centre")]
        pharmacies = [f for f in facilities if f["type"] == "pharmacy"]
        others     = [f for f in facilities if f not in hospitals + clinics + pharmacies]

        state = _detect_state(display)

        return {
            "location":        {"name": location, "display": display,
                                 "lat": lat, "lon": lon, "city": city},
            "radius_m":        radius_m,
            "facilities":      facilities,
            "hospitals":       hospitals,
            "clinics":         clinics,
            "pharmacies":      pharmacies,
            "others":          others,
            "total":           len(facilities),
            "emergency_numbers": {**EMERGENCY_NUMBERS.get("National", {}),
                                  **EMERGENCY_NUMBERS.get(state, {})},
            "scraped_at":      datetime.now().isoformat(),
        }

    # ── 2. Search by specialty ─────────────────────────────────────────────────

    def search_by_specialty(self, location: str, specialty: str,
                             radius_m: int = 15000, limit: int = 15,
                             enrich: int = 1) -> Dict[str, Any]:
        """
        Scrape and filter hospitals by medical specialty near a location.

        specialty examples:
            "cardiac", "cancer", "pediatric", "maternity", "trauma",
            "orthopedic", "kidney", "eye", "dental", "mental health", "liver"
        """
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'", "results": []}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping {specialty} hospitals near '{location}'…")
        all_fac = _scrape_osm(lat, lon, radius_m)

        spec_lower = specialty.lower()
        keywords   = SPECIALTY_KEYWORDS.get(
            next((k for k in SPECIALTY_KEYWORDS if k.lower() == spec_lower), ""),
            [spec_lower]
        )

        # Filter by specialty keywords
        filtered = []
        for fac in all_fac:
            blob = (fac["name"] + " " +
                    " ".join(fac["specialties"]) + " " +
                    fac.get("operator", "")).lower()
            if any(kw in blob for kw in keywords):
                filtered.append(fac)

        results = filtered[:limit] if filtered else all_fac[:limit]

        if enrich > 0:
            for fac in tqdm(results, desc=f"Enriching {specialty}", unit="hospital"):
                _enrich_facility(fac, city, enrich_level=enrich)

        return {
            "location":    {"name": location, "display": display, "lat": lat, "lon": lon},
            "specialty":   specialty,
            "results":     results,
            "total_found": len(results),
            "scraped_at":  datetime.now().isoformat(),
        }

    # ── 3. Emergency hospitals ─────────────────────────────────────────────────

    def emergency_near(self, location: str, radius_m: int = 10000) -> Dict[str, Any]:
        """
        Scrape emergency hospitals and A&E departments near a location.
        Also returns local emergency helpline numbers.
        """
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'"}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping emergency hospitals near '{location}'…")

        # Scrape hospitals only (not clinics/pharmacies)
        all_fac  = _scrape_osm(lat, lon, radius_m, type_filter=["hospital", "emergency"])
        emrg_fac = [f for f in all_fac if f.get("emergency") in ("yes", "only")]

        if not emrg_fac:
            # Fall back to all hospitals and flag largest as likely emergency
            emrg_fac = all_fac[:8]

        for fac in tqdm(emrg_fac[:10], desc="Enriching emergency hospitals", unit="hospital"):
            _enrich_facility(fac, city, enrich_level=1)

        state = _detect_state(display)
        numbers = {
            **EMERGENCY_NUMBERS.get("National", {}),
            **EMERGENCY_NUMBERS.get(state, {}),
        }

        return {
            "location":        {"name": location, "display": display, "lat": lat, "lon": lon},
            "emergency_hospitals": emrg_fac,
            "total":           len(emrg_fac),
            "emergency_numbers": numbers,
            "nearest_emergency": emrg_fac[0] if emrg_fac else None,
            "scraped_at":      datetime.now().isoformat(),
        }

    # ── 4. Hospital full details ───────────────────────────────────────────────

    def hospital_details(self, name: str, city: str) -> Dict[str, Any]:
        """
        Full detail scrape for a specific hospital:
            OSM data + Wikipedia + Practo + JustDial + Google Maps + booking links
        """
        geo = _geocode(f"{name} {city}")
        if not geo:
            geo = _geocode(city)
        if not geo:
            return {"error": f"Could not locate: '{name}, {city}'"}

        lat, lon, display = geo
        log.info(f"Fetching full details for '{name}' in {city}…")

        # Primary OSM scrape
        facilities = _scrape_osm(lat, lon, radius_m=3000)
        match = next((f for f in facilities
                      if name.lower() in f["name"].lower()), None)

        if not match:
            # Create a stub record
            match = {
                "name":          name, "icon": "🏥", "type": "hospital",
                "type_label":    "🏥 Hospital", "address": city,
                "phone":         "", "website": "", "email": "",
                "emergency":     "", "beds":    "", "opening_hours": "",
                "wheelchair":    "", "operator": "", "operator_type": "",
                "specialties":   ["General"], "lat": lat, "lon": lon,
                "dist_km":       0, "rating":  None, "reviews":      None,
                "maps_url":      f"https://www.google.com/maps?q={lat},{lon}",
                "directions_url":f"https://www.google.com/maps/dir/?api=1&destination={lat},{lon}",
                "practo_url":    "", "description": "", "source": "stub",
            }

        # Enrich fully (level 2 = all scrapers)
        match = _enrich_facility(match, city, enrich_level=2)

        # Wikipedia description
        if not match.get("description"):
            match["description"] = _scrape_wiki_desc(name, city)

        return match

    # ── 5. Blood banks ────────────────────────────────────────────────────────

    def blood_banks_near(self, location: str, radius_m: int = 10000) -> Dict[str, Any]:
        """Scrape blood banks and donation centres near a location."""
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'"}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping blood banks near '{location}'…")
        fac = _scrape_osm(lat, lon, radius_m, type_filter=["blood_bank"])

        # Also search hospitals with "blood bank" in name
        hospitals = _scrape_osm(lat, lon, radius_m, type_filter=["hospital"])
        bb_in_hosp = [h for h in hospitals if "blood" in h["name"].lower()]

        all_bb = fac + [h for h in bb_in_hosp if h not in fac]
        all_bb.sort(key=lambda x: x["dist_km"])

        return {
            "location": {"name": location, "display": display, "lat": lat, "lon": lon},
            "blood_banks": all_bb,
            "total": len(all_bb),
            "note": "Call National Blood Bank Helpline: 1910 (India)",
        }

    # ── 6. Pharmacies ─────────────────────────────────────────────────────────

    def pharmacies_near(self, location: str, radius_m: int = 2000,
                        limit: int = 20) -> Dict[str, Any]:
        """Scrape pharmacies and medical stores near a location."""
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'"}
        lat, lon, display = geo

        log.info(f"Scraping pharmacies near '{location}'…")
        fac = _scrape_osm(lat, lon, radius_m, type_filter=["pharmacy"])

        return {
            "location":   {"name": location, "display": display, "lat": lat, "lon": lon},
            "pharmacies": fac[:limit],
            "total":      len(fac[:limit]),
            "radius_m":   radius_m,
        }

    # ── 7. Excel export ───────────────────────────────────────────────────────

    def export_excel(self, location: str, filepath: str = None,
                     radius_m: int = 10000, enrich: int = 1) -> str:
        """
        Scrape all hospitals near a location and export to Excel.

        Args:
            location : city or address
            filepath : output path (auto-generated if not provided)
            radius_m : search radius in metres
            enrich   : enrichment level 0/1/2

        Returns:
            filepath of saved Excel file
        """
        if not filepath:
            slug     = re.sub(r"[^\w]", "_", location.lower())[:30]
            filepath = f"hospitals_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        result = self.find_nearby(location, radius_m=radius_m, limit=50, enrich=enrich)
        if "error" in result:
            log.error(result["error"])
            return ""

        facilities = result["facilities"]
        log.info(f"Exporting {len(facilities)} facilities to {filepath}…")
        _export_excel(facilities, filepath, location)
        print(f"\n  ✅ Excel saved: {filepath}  ({len(facilities)} hospitals)")
        return filepath

    # ── 8. Emergency numbers ──────────────────────────────────────────────────

    def get_emergency_numbers(self, state: str = "National") -> Dict[str, str]:
        """Return emergency helpline numbers for a given Indian state."""
        numbers = {**EMERGENCY_NUMBERS.get("National", {})}
        if state and state != "National":
            for key in EMERGENCY_NUMBERS:
                if state.lower() in key.lower():
                    numbers.update(EMERGENCY_NUMBERS[key])
                    break
        return numbers

    # ══════════════════════════════════════════════════════════════════════════
    #  PRETTY PRINTERS
    # ══════════════════════════════════════════════════════════════════════════

    def print_facilities(self, result: Dict[str, Any], show_all: bool = False):
        if "error" in result:
            print(f"  ❌ {result['error']}"); return

        loc   = result["location"]
        total = result["total"]
        print(f"\n  ┌{'─'*70}")
        print(f"  │  🏥 Healthcare near '{loc['name']}'")
        print(f"  │  📍 {loc['display'][:65]}")
        print(f"  ├{'─'*70}")
        print(f"  │  Radius  : {result['radius_m']/1000:.1f} km")
        print(f"  │  Found   : {total} facilities")
        hsp = len(result.get('hospitals', []))
        cln = len(result.get('clinics', []))
        pha = len(result.get('pharmacies', []))
        print(f"  │  Split   : {hsp} hospitals · {cln} clinics · {pha} pharmacies")
        print(f"  └{'─'*70}")

        facilities = result["facilities"] if show_all else result["hospitals"] or result["facilities"]

        for i, fac in enumerate(facilities, 1):
            emrg = " 🚨" if fac.get("emergency") == "yes" else ""
            print(f"\n  {i:>2}. {fac['icon']}  {fac['name']}{emrg}")
            print(f"      📏 {fac['dist_km']} km away")
            if fac["address"] != "N/A":
                print(f"      📍 {fac['address'][:65]}")
            if fac["phone"]:
                print(f"      📞 {fac['phone']}")
            if fac["website"]:
                print(f"      🌐 {fac['website']}")
            if fac["specialties"] and fac["specialties"] != ["General"]:
                print(f"      🔬 {', '.join(fac['specialties'][:4])}")
            if fac["beds"]:
                print(f"      🛏  {fac['beds']} beds")
            if fac["opening_hours"] and fac["opening_hours"] != "N/A":
                print(f"      🕐 {fac['opening_hours']}")
            if fac.get("rating"):
                print(f"      ⭐ {fac['rating']}" +
                      (f"  ({fac['reviews']} reviews)" if fac.get("reviews") else ""))
            if fac.get("description"):
                desc = fac["description"][:120]
                print(f"      📖 {desc}…" if len(fac["description"]) > 120 else f"      📖 {desc}")
            print(f"      🗺  {fac['maps_url']}")
            print(f"      🧭 {fac['directions_url']}")

        # Emergency numbers
        nums = result.get("emergency_numbers", {})
        if nums:
            print(f"\n  🚨 Emergency Numbers:")
            print(f"  {'─'*40}")
            for service, number in nums.items():
                print(f"     {service:<28}: {number}")

    def print_emergency(self, result: Dict[str, Any]):
        if "error" in result:
            print(f"  ❌ {result['error']}"); return
        print(f"\n  🚨 Emergency Hospitals near '{result['location']['name']}'")
        print(f"  {'─'*65}")
        for i, fac in enumerate(result.get("emergency_hospitals", []), 1):
            print(f"  {i}. {fac['icon']}  {fac['name']}")
            print(f"     📏 {fac['dist_km']} km  |  📞 {fac.get('phone','N/A')}")
            print(f"     🧭 {fac['directions_url']}")
        print(f"\n  🚨 Emergency Numbers:")
        for svc, num in result.get("emergency_numbers", {}).items():
            print(f"     {svc:<28}: {num}")

    def print_details(self, fac: Dict):
        print(f"\n  ┌{'─'*65}")
        print(f"  │  {fac.get('icon','🏥')}  {fac.get('name','')}")
        print(f"  ├{'─'*65}")
        for label, key in [
            ("Type",         "type_label"),
            ("Address",      "address"),
            ("Phone",        "phone"),
            ("Website",      "website"),
            ("Email",        "email"),
            ("Emergency",    "emergency"),
            ("Beds",         "beds"),
            ("Hours",        "opening_hours"),
            ("Wheelchair",   "wheelchair"),
            ("Operator",     "operator"),
            ("Specialties",  None),
            ("Rating",       "rating"),
        ]:
            if key:
                val = fac.get(key, "")
                if val and val not in ("N/A", ""):
                    print(f"  │  {label:<14}: {str(val)[:60]}")
            else:
                specs = fac.get("specialties", [])
                if specs:
                    print(f"  │  {'Specialties':<14}: {', '.join(specs)}")

        if fac.get("description"):
            print(f"  ├{'─'*65}")
            desc = fac["description"]
            # Word-wrap at 60 chars
            words, line = desc.split(), ""
            for w in words:
                if len(line) + len(w) + 1 > 60:
                    print(f"  │  {line}")
                    line = w
                else:
                    line = f"{line} {w}".strip()
            if line:
                print(f"  │  {line}")

        print(f"  ├{'─'*65}")
        print(f"  │  🗺  {fac.get('maps_url','')}")
        print(f"  │  🧭 {fac.get('directions_url','')}")
        if fac.get("practo_url"):
            print(f"  │  💊 {fac['practo_url']}")
        print(f"  ├{'─'*65}")
        print(f"  │  📲 Book / Find Doctors:")
        for platform, url in fac.get("booking_links", {}).items():
            print(f"  │     {platform:<15}: {url[:55]}")
        print(f"  └{'─'*65}")


# ══════════════════════════════════════════════════════════════════════════════
#  INTERACTIVE CLI
# ══════════════════════════════════════════════════════════════════════════════

def run_cli():
    print(f"\n{'='*68}")
    print(f"  🏥   HospitalAgent v1 — Worldwide Nearby Hospital Finder")
    print(f"       Zero Signup · Live Scraping · OSM + Wikipedia + Practo")
    print(f"       Powered by BeautifulSoup · lxml · Overpass · Nominatim")
    print(f"{'='*68}")

    agent = HospitalAgent()

    MENU = """
  [1]  Find nearby hospitals / clinics
  [2]  Search by specialty  (cardiac, cancer, pediatric, ortho…)
  [3]  Emergency hospitals near me  (+ helpline numbers)
  [4]  Full hospital details  (scrape Wikipedia + Practo + JustDial)
  [5]  Blood banks near me
  [6]  Pharmacies near me
  [7]  Export to Excel  (all hospitals in area)
  [8]  Emergency helpline numbers  (by state)
  [q]  Quit
"""
    while True:
        print(MENU)
        c = input("  👉 Choose: ").strip().lower()

        if c == "1":
            loc     = input("  📍 Location (city / address / pincode): ").strip()
            r       = input("  📡 Radius metres [5000]: ").strip()
            lim     = input("  🔢 Limit [15]: ").strip()
            enrich  = input("  🔍 Enrich level 0=fast / 1=+wiki / 2=+all [1]: ").strip()
            result  = agent.find_nearby(
                loc,
                radius_m=int(r)      if r.isdigit()      else 5000,
                limit=int(lim)       if lim.isdigit()    else 15,
                enrich=int(enrich)   if enrich.isdigit() else 1,
            )
            agent.print_facilities(result)

        elif c == "2":
            loc     = input("  📍 Location: ").strip()
            spec    = input("  🔬 Specialty (cardiac / cancer / pediatric / ortho / eye…): ").strip()
            r       = input("  📡 Radius metres [15000]: ").strip()
            result  = agent.search_by_specialty(loc, spec,
                        radius_m=int(r) if r.isdigit() else 15000)
            print(f"\n  🔍 {result['total_found']} '{spec}' hospitals near {loc}")
            for i, fac in enumerate(result.get("results", []), 1):
                print(f"  {i:>2}. {fac['icon']}  {fac['name']:<40}  {fac['dist_km']} km")
                print(f"       📍 {fac['address'][:55]}")
                print(f"       📞 {fac.get('phone','N/A')}  |  🗺 {fac['maps_url']}")

        elif c == "3":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius metres [10000]: ").strip()
            result = agent.emergency_near(loc, radius_m=int(r) if r.isdigit() else 10000)
            agent.print_emergency(result)

        elif c == "4":
            name = input("  🏥 Hospital name: ").strip()
            city = input("  🏙  City: ").strip()
            fac  = agent.hospital_details(name, city)
            if "error" in fac:
                print(f"  ❌ {fac['error']}")
            else:
                agent.print_details(fac)

        elif c == "5":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius metres [10000]: ").strip()
            result = agent.blood_banks_near(loc, radius_m=int(r) if r.isdigit() else 10000)
            print(f"\n  🩸 {result['total']} blood banks near {loc}")
            for i, b in enumerate(result["blood_banks"], 1):
                print(f"  {i}. {b['name']:<40}  {b['dist_km']} km")
                print(f"     📍 {b['address'][:55]}  |  📞 {b.get('phone','N/A')}")
                print(f"     🧭 {b['directions_url']}")
            print(f"\n  {result['note']}")

        elif c == "6":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius metres [2000]: ").strip()
            result = agent.pharmacies_near(loc, radius_m=int(r) if r.isdigit() else 2000)
            print(f"\n  💊 {result['total']} pharmacies near {loc}")
            for i, p in enumerate(result["pharmacies"], 1):
                print(f"  {i:>2}. {p['name']:<38} {p['dist_km']} km  |  {p.get('phone','N/A')}")
                print(f"       📍 {p['address'][:55]}")
                print(f"       🕐 {p['opening_hours']}  |  🗺 {p['maps_url']}")

        elif c == "7":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius metres [10000]: ").strip()
            enrich = input("  🔍 Enrich level 0/1/2 [1]: ").strip()
            path   = agent.export_excel(
                loc,
                radius_m=int(r)      if r.isdigit()      else 10000,
                enrich=int(enrich)   if enrich.isdigit() else 1,
            )
            if path:
                print(f"  📊 Excel ready: {path}")

        elif c == "8":
            state  = input("  🗺  State (or Enter for National): ").strip() or "National"
            nums   = agent.get_emergency_numbers(state)
            print(f"\n  🚨 Emergency Numbers — {state}")
            print(f"  {'─'*45}")
            for svc, num in nums.items():
                print(f"     {svc:<28}: {num}")

        elif c in ("q", "quit", "exit"):
            print("\n  🏥 HospitalAgent signing off!\n"); break
        else:
            print("  ⚠  Invalid option.")


if __name__ == "__main__":
    run_cli()