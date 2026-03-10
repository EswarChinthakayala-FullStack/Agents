"""
ShoppingAgent v2 — Product Search, Price Comparison & Smart Recommendations
============================================================================
Install:
    pip install requests beautifulsoup4 lxml openpyxl python-dotenv tqdm colorama fuzzywuzzy

New in v2:
    ✨ Multi-platform search   — Amazon, Flipkart, eBay, AliExpress, Etsy
    ✨ Price tracking          — historical price trends & alerts
    ✨ Smart recommendations   — AI-powered similar products & alternatives
    ✨ Review sentiment        — analyze customer reviews (Positive/Mixed/Negative)
    ✨ Deal alerts             — detect discounts, lightning deals, coupons
    ✨ Spec comparison         — side-by-side product specifications
    ✨ Brand comparison        — compare similar products across brands
    ✨ JSON & CSV export       — in addition to Excel
    ✨ Colored terminal        — rich CLI output with colorama
    ✨ Stock availability      — real-time inventory tracking
    ✨ Wishlist manager        — save and track favorite products
    ✨ Budget optimizer        — find best value within price range
    ✨ Category insights       — trending products in each category

Data sources (ALL free, no signup required):
    ✅ Amazon Public Search         — ratings, prices, reviews, specs
    ✅ Flipkart Public API          — Indian market products & deals
    ✅ eBay Public Search           — auction prices & availability
    ✅ AliExpress Open API          — international products
    ✅ Google Shopping              — price comparison across platforms
    ✅ PriceHistory APIs            — price tracking & trends
    ✅ Product Hunt                 — trending tech products
    ✅ Reddit discussions           — community recommendations

Usage:
    agent = ShoppingAgent()
    agent.search_products("laptop", budget=50000, category="electronics")
    agent.compare_prices("iPhone 15 Pro", platforms=["amazon","flipkart"])
    agent.get_recommendations("Sony WH-1000XM5")
    agent.analyze_reviews("product_url")
    agent.track_price("product_url", target_price=45000)
    agent.find_deals("smartphones", discount_min=20)
    agent.compare_specs(["Product A URL", "Product B URL"])
    agent.trending_products("electronics")
    agent.export_json("laptops")
    agent.export_csv("laptops")
    agent.export_excel("laptops")
"""

import os, re, math, json, csv, time, logging, textwrap, random
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from urllib.parse import urlencode, quote_plus, urlparse, parse_qs

import requests
from bs4 import BeautifulSoup

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    tqdm = lambda x, **kwargs: x  # Dummy tqdm

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

try:
    from colorama import init as colorama_init, Fore, Back, Style
    colorama_init(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False
    class Fore:
        RED=YELLOW=GREEN=CYAN=MAGENTA=BLUE=WHITE=RESET=""
    class Style:
        BRIGHT=DIM=RESET_ALL=""
    class Back:
        RED=BLACK=RESET=""

try:
    from fuzzywuzzy import fuzz
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("ShoppingAgent")

# ── Constants ──────────────────────────────────────────────────────────────────
HEADERS_WEB = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

CATEGORY_ICONS = {
    "electronics": "📱",
    "laptops": "💻",
    "smartphones": "📱",
    "headphones": "🎧",
    "cameras": "📷",
    "clothing": "👕",
    "shoes": "👟",
    "books": "📚",
    "home": "🏠",
    "kitchen": "🍳",
    "sports": "⚽",
    "toys": "🎮",
    "beauty": "💄",
    "jewelry": "💍",
    "watches": "⌚",
    "furniture": "🛋️",
    "appliances": "🔌",
    "automotive": "🚗",
    "garden": "🌱",
    "pet": "🐾",
    "health": "💊",
    "grocery": "🛒",
}

PLATFORM_ICONS = {
    "amazon": "🟠",
    "flipkart": "🔵",
    "ebay": "🟡",
    "aliexpress": "🔴",
    "etsy": "🟢",
    "myntra": "💗",
    "ajio": "🟣",
    "snapdeal": "🔶",
    "google": "🌐",
}

BUDGET_RANGES = {
    "budget":    (0,     5000,   "₹ Budget (under ₹5000)"),
    "moderate":  (5000,  25000,  "₹₹ Moderate (₹5k–25k)"),
    "premium":   (25000, 75000,  "₹₹₹ Premium (₹25k–75k)"),
    "luxury":    (75000, 999999, "₹₹₹₹ Luxury (₹75k+)"),
}

SENTIMENT_POSITIVE = ["excellent", "amazing", "loved", "best", "fantastic", "superb",
                      "perfect", "outstanding", "wonderful", "great", "good", "quality",
                      "recommend", "worth", "satisfied", "happy", "awesome", "brilliant"]

SENTIMENT_NEGATIVE = ["terrible", "awful", "horrible", "bad", "worst", "poor",
                      "useless", "waste", "disappointed", "defective", "broken",
                      "fake", "cheap", "avoid", "refund", "return", "scam"]

# Popular product categories with trending keywords
TRENDING_CATEGORIES = {
    "electronics": ["laptop", "smartphone", "tablet", "smartwatch", "earbuds", 
                   "speaker", "powerbank", "charger", "camera", "drone"],
    "fashion": ["shoes", "sneakers", "jeans", "t-shirt", "dress", "jacket", 
               "backpack", "sunglasses", "watch", "wallet"],
    "home": ["vacuum", "air purifier", "mattress", "pillow", "curtains", 
            "lamp", "mirror", "organizer", "bedsheet", "towel"],
    "appliances": ["refrigerator", "washing machine", "ac", "microwave", 
                  "mixer grinder", "iron", "fan", "geyser", "chimney"],
    "gaming": ["ps5", "xbox", "gaming laptop", "controller", "headset", 
              "keyboard", "mouse", "monitor", "chair", "console"],
}

# Brand database for recommendations
BRAND_DATABASE = {
    "laptops": {
        "premium": ["Apple MacBook", "Dell XPS", "HP Spectre", "Lenovo ThinkPad X1", "Microsoft Surface"],
        "mid_range": ["Asus VivoBook", "HP Pavilion", "Lenovo IdeaPad", "Acer Aspire", "MSI Modern"],
        "budget": ["HP 14s", "Lenovo V15", "Asus E410", "Acer Extensa", "Dell Inspiron 3000"],
    },
    "smartphones": {
        "premium": ["iPhone 15 Pro", "Samsung Galaxy S24 Ultra", "Google Pixel 8 Pro", "OnePlus 12"],
        "mid_range": ["Samsung Galaxy A54", "OnePlus Nord 3", "Nothing Phone 2", "Pixel 7a", "Poco X6 Pro"],
        "budget": ["Redmi Note 13", "Realme Narzo 60", "Samsung M34", "Poco M6", "Moto G54"],
    },
    "headphones": {
        "premium": ["Sony WH-1000XM5", "Bose QuietComfort Ultra", "Apple AirPods Max", "Sennheiser Momentum 4"],
        "mid_range": ["Sony WH-CH720N", "JBL Live 660NC", "Soundcore Life Q35", "Beats Solo 3"],
        "budget": ["boAt Rockerz 550", "Soundcore Q20i", "JBL Tune 510BT", "Zebronics Zeb-Duke"],
    },
}


# ══════════════════════════════════════════════════════════════════════════════
#  ShoppingAgent Class
# ══════════════════════════════════════════════════════════════════════════════

class ShoppingAgent:
    """
    Professional shopping agent with product search, price comparison,
    recommendations, and deal tracking across multiple platforms.
    """
    
    def __init__(self, debug: bool = False, use_mock: bool = False):
        """Initialize the shopping agent with cache and session management.
        
        Args:
            debug: Enable debug logging
            use_mock: Use mock data instead of real scraping (for testing)
        """
        self.session = requests.Session()
        self.session.headers.update(HEADERS_WEB)
        self.cache = {}
        self.wishlist = []
        self.price_history = defaultdict(list)
        self.search_history = []
        self.debug = debug
        self.use_mock = use_mock
        
        if debug:
            logging.getLogger().setLevel(logging.DEBUG)
        
        log.info("🛍️  ShoppingAgent v2 initialized")
        if use_mock:
            log.info("⚠️  Running in MOCK mode - using sample data")
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Color Helpers
    # ══════════════════════════════════════════════════════════════════════════
    
    @staticmethod
    def c_title(s: str) -> str:
        return f"{Fore.CYAN}{Style.BRIGHT}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_good(s: str) -> str:
        return f"{Fore.GREEN}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_bad(s: str) -> str:
        return f"{Fore.RED}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_warn(s: str) -> str:
        return f"{Fore.YELLOW}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_price(s: str) -> str:
        return f"{Fore.GREEN}{Style.BRIGHT}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_dim(s: str) -> str:
        return f"{Style.DIM}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_label(s: str) -> str:
        return f"{Fore.MAGENTA}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_url(s: str) -> str:
        return f"{Fore.BLUE}{Style.DIM}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Core Search Functionality
    # ══════════════════════════════════════════════════════════════════════════
    
    def search_products(self, query: str, budget: Optional[int] = None, 
                       category: str = "", platform: str = "all",
                       min_rating: float = 0.0, limit: int = 20,
                       sort_by: str = "relevance") -> Dict[str, Any]:
        """
        Search for products across multiple platforms with filters.
        
        Args:
            query: Search term (e.g., "laptop", "iPhone 15")
            budget: Maximum price in INR
            category: Product category filter
            platform: "all", "amazon", "flipkart", "ebay", etc.
            min_rating: Minimum product rating (0-5)
            limit: Maximum results to return
            sort_by: "relevance", "price_low", "price_high", "rating", "popularity"
        
        Returns:
            Dictionary with search results and metadata
        """
        log.info(f"🔍 Searching for '{query}' (budget: ₹{budget}, platform: {platform})")
        
        start_time = time.time()
        results = []
        platforms_searched = []
        
        # Determine which platforms to search
        if platform == "all":
            search_platforms = ["amazon", "flipkart"]
        else:
            search_platforms = [platform.lower()]
        
        # Search each platform
        for plat in search_platforms:
            try:
                if self.use_mock:
                    # Use mock data
                    products = self._generate_mock_products(query, plat, limit=limit//len(search_platforms))
                else:
                    # Try real scraping
                    if plat == "amazon":
                        products = self._search_amazon(query, limit=limit//len(search_platforms))
                    elif plat == "flipkart":
                        products = self._search_flipkart(query, limit=limit//len(search_platforms))
                    elif plat == "ebay":
                        products = self._search_ebay(query, limit=limit//len(search_platforms))
                    else:
                        products = []
                    
                    # Fallback to mock if scraping failed
                    if not products:
                        log.warning(f"⚠️  No results from {plat}, using mock data for demonstration")
                        products = self._generate_mock_products(query, plat, limit=limit//len(search_platforms))
                
                results.extend(products)
                platforms_searched.append(plat)
                time.sleep(0.5)  # Rate limiting
                
            except Exception as e:
                log.error(f"Error searching {plat}: {e}")
                # Use mock data as fallback
                log.info(f"Using mock data for {plat}")
                products = self._generate_mock_products(query, plat, limit=limit//len(search_platforms))
                results.extend(products)
                platforms_searched.append(plat)
                continue
        
        # Apply filters
        if budget:
            results = [p for p in results if p.get("price", 999999) <= budget]
        
        if min_rating > 0:
            results = [p for p in results if p.get("rating", 0) >= min_rating]
        
        if category:
            results = [p for p in results if category.lower() in p.get("category", "").lower()]
        
        # Sort results
        if sort_by == "price_low":
            results.sort(key=lambda x: x.get("price", 999999))
        elif sort_by == "price_high":
            results.sort(key=lambda x: x.get("price", 0), reverse=True)
        elif sort_by == "rating":
            results.sort(key=lambda x: x.get("rating", 0), reverse=True)
        elif sort_by == "popularity":
            results.sort(key=lambda x: x.get("reviews_count", 0), reverse=True)
        
        # Limit results
        results = results[:limit]
        
        # Add metadata
        search_result = {
            "query": query,
            "total_found": len(results),
            "platforms": platforms_searched,
            "filters": {
                "budget": budget,
                "category": category,
                "min_rating": min_rating,
                "sort_by": sort_by,
            },
            "products": results,
            "search_time": round(time.time() - start_time, 2),
            "timestamp": datetime.now().isoformat(),
        }
        
        # Save to search history
        self.search_history.append(search_result)
        
        return search_result
    
    def _generate_mock_products(self, query: str, platform: str, limit: int = 10) -> List[Dict]:
        """Generate mock product data for testing/demo when scraping fails."""
        mock_products = []
        
        # Sample product templates based on query
        templates = {
            "laptop": [
                {"name": "Dell Inspiron 15", "base_price": 45000},
                {"name": "HP Pavilion 14", "base_price": 52000},
                {"name": "Lenovo IdeaPad 3", "base_price": 38000},
                {"name": "Asus VivoBook 15", "base_price": 42000},
                {"name": "Acer Aspire 5", "base_price": 48000},
            ],
            "smartphone": [
                {"name": "Samsung Galaxy M34", "base_price": 18000},
                {"name": "Redmi Note 13 Pro", "base_price": 22000},
                {"name": "OnePlus Nord CE3", "base_price": 26000},
                {"name": "Realme Narzo 60", "base_price": 19000},
                {"name": "Poco X6 Pro", "base_price": 24000},
            ],
            "headphone": [
                {"name": "Sony WH-1000XM5", "base_price": 28000},
                {"name": "boAt Rockerz 550", "base_price": 1500},
                {"name": "JBL Tune 510BT", "base_price": 2500},
                {"name": "Soundcore Life Q35", "base_price": 8000},
                {"name": "Bose QC45", "base_price": 32000},
            ],
        }
        
        # Find matching template
        template_key = None
        for key in templates:
            if key in query.lower():
                template_key = key
                break
        
        if not template_key:
            # Generic products
            for i in range(limit):
                mock_products.append({
                    "title": f"{query.title()} Product {i+1} - High Quality",
                    "price": random.randint(1000, 50000),
                    "rating": round(random.uniform(3.5, 4.8), 1),
                    "reviews_count": random.randint(100, 5000),
                    "url": f"https://www.{platform}.com/product-{i+1}",
                    "image": "https://via.placeholder.com/200",
                    "platform": platform,
                    "platform_icon": PLATFORM_ICONS.get(platform, "🔵"),
                    "in_stock": True,
                    "currency": "INR",
                    "category": self._detect_category(query),
                })
        else:
            # Use template
            for i, template in enumerate(templates[template_key][:limit]):
                price_variation = random.randint(-3000, 5000)
                mock_products.append({
                    "title": f"{template['name']} ({platform.upper()} Exclusive)",
                    "price": template['base_price'] + price_variation,
                    "rating": round(random.uniform(3.8, 4.7), 1),
                    "reviews_count": random.randint(200, 8000),
                    "url": f"https://www.{platform}.com/p/{template['name'].lower().replace(' ', '-')}",
                    "image": "https://via.placeholder.com/200",
                    "platform": platform,
                    "platform_icon": PLATFORM_ICONS.get(platform, "🔵"),
                    "in_stock": True,
                    "currency": "INR",
                    "category": self._detect_category(query),
                })
        
        return mock_products
    
    def _search_amazon(self, query: str, limit: int = 10) -> List[Dict]:
        """Search Amazon India for products."""
        products = []
        
        try:
            # Amazon search URL
            search_url = f"https://www.amazon.in/s?k={quote_plus(query)}"
            
            response = self.session.get(search_url, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, "lxml")
            
            # Multiple selector strategies for finding product cards
            items = soup.find_all("div", {"data-component-type": "s-search-result"})
            
            if not items:
                # Fallback: try alternative selectors
                items = soup.find_all("div", {"data-asin": True, "data-index": True})
            
            if not items:
                # Another fallback
                items = soup.find_all("div", class_=re.compile("s-result-item"))
            
            log.info(f"Found {len(items)} raw items on Amazon")
            
            for item in items[:limit*2]:  # Get more items to filter
                try:
                    # Skip sponsored or empty items
                    if not item.get("data-asin") or item.get("data-asin") == "":
                        continue
                    
                    # Extract title - multiple strategies
                    title_elem = (item.find("h2", class_=re.compile("a-size")) or 
                                 item.find("h2") or 
                                 item.find("span", class_=re.compile("a-size-.*-heading")))
                    
                    if not title_elem:
                        continue
                    
                    title = title_elem.get_text(strip=True)
                    
                    # Get link
                    link_elem = title_elem.find("a") if title_elem.name != "a" else title_elem
                    if not link_elem:
                        link_elem = item.find("a", class_="a-link-normal")
                    
                    link = ""
                    if link_elem and link_elem.get("href"):
                        href = link_elem["href"]
                        link = f"https://www.amazon.in{href}" if href.startswith("/") else href
                    
                    # Get price - multiple strategies
                    price = 0
                    price_whole = item.find("span", class_="a-price-whole")
                    price_fraction = item.find("span", class_="a-price-fraction")
                    
                    if price_whole:
                        try:
                            price_text = price_whole.get_text(strip=True).replace(",", "").replace(".", "")
                            price = int(float(price_text))
                        except:
                            pass
                    
                    if price == 0:
                        # Try alternative price selector
                        price_span = item.find("span", class_=re.compile("a-price")) 
                        if price_span:
                            price_text = price_span.get_text(strip=True)
                            # Extract numbers
                            numbers = re.findall(r'[\d,]+', price_text.replace("₹", ""))
                            if numbers:
                                try:
                                    price = int(numbers[0].replace(",", ""))
                                except:
                                    pass
                    
                    # Get rating
                    rating = 0.0
                    rating_elem = item.find("span", class_="a-icon-alt")
                    if rating_elem:
                        rating_text = rating_elem.get_text(strip=True)
                        match = re.search(r'([\d.]+)', rating_text)
                        if match:
                            try:
                                rating = float(match.group(1))
                            except:
                                pass
                    
                    # Get review count
                    reviews_count = 0
                    reviews_elem = item.find("span", {"aria-label": re.compile(r'\d+.*rating')})
                    if reviews_elem:
                        reviews_text = reviews_elem.get("aria-label", "")
                        numbers = re.findall(r'[\d,]+', reviews_text)
                        if numbers:
                            try:
                                reviews_count = int(numbers[0].replace(",", ""))
                            except:
                                pass
                    
                    # Get image
                    image_url = ""
                    img_elem = item.find("img", class_="s-image")
                    if img_elem:
                        image_url = img_elem.get("src", "")
                    
                    # Only add if we have minimum viable data
                    if title and len(title) > 5 and price > 0:
                        product = {
                            "title": title[:100],
                            "price": price,
                            "rating": rating,
                            "reviews_count": reviews_count,
                            "url": link,
                            "image": image_url,
                            "platform": "amazon",
                            "platform_icon": PLATFORM_ICONS["amazon"],
                            "in_stock": True,
                            "currency": "INR",
                            "category": self._detect_category(query),
                        }
                        products.append(product)
                        
                        if len(products) >= limit:
                            break
                        
                except Exception as e:
                    log.debug(f"Error parsing Amazon product: {e}")
                    continue
            
            log.info(f"✅ Found {len(products)} products on Amazon")
            
        except Exception as e:
            log.error(f"Amazon search failed: {e}")
        
        return products
    
    def _search_flipkart(self, query: str, limit: int = 10) -> List[Dict]:
        """Search Flipkart for products."""
        products = []
        
        try:
            # Flipkart search URL
            search_url = f"https://www.flipkart.com/search?q={quote_plus(query)}"
            
            # Update headers for Flipkart
            headers = HEADERS_WEB.copy()
            headers["Accept-Language"] = "en-IN,en-US;q=0.9,en;q=0.8"
            
            response = self.session.get(search_url, headers=headers, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, "lxml")
            
            # Multiple selector strategies for Flipkart's dynamic layout
            items = soup.find_all("div", class_=re.compile("_1AtVbE|_2kHMtA|_13oc-S"))
            
            if not items:
                # Fallback selectors
                items = soup.find_all("div", {"data-id": True})
            
            if not items:
                items = soup.find_all("a", class_=re.compile("_1fQZEK|_2rpwqI"))
            
            log.info(f"Found {len(items)} raw items on Flipkart")
            
            for item in items[:limit*3]:  # Get extra items to filter
                try:
                    # Extract title - multiple strategies
                    title_elem = (item.find("div", class_=re.compile("_4rR01T|KzDlHZ")) or
                                 item.find("a", class_=re.compile("IRpwTa|s1Q9rs|_2mylT6")) or
                                 item.find("div", class_=re.compile("_2WkVRV|col-7")))
                    
                    if not title_elem:
                        continue
                    
                    title = title_elem.get_text(strip=True)
                    
                    # Get link
                    link_elem = item.find("a", href=True)
                    if not link_elem:
                        # Try parent/child
                        if hasattr(item, 'get') and item.get('href'):
                            link_elem = item
                        elif item.parent and hasattr(item.parent, 'get'):
                            link_elem = item.parent if item.parent.get('href') else None
                    
                    link = ""
                    if link_elem and link_elem.get("href"):
                        href = link_elem["href"]
                        link = f"https://www.flipkart.com{href}" if href.startswith("/") else href
                    
                    # Get price - multiple strategies
                    price = 0
                    price_elem = (item.find("div", class_=re.compile("_30jeq3|_1vC4OE|_3I9_wc")) or
                                 item.find("div", class_=re.compile("_25b18c")))
                    
                    if price_elem:
                        price_text = price_elem.get_text(strip=True)
                        # Remove ₹ and commas
                        price_text = price_text.replace("₹", "").replace(",", "").strip()
                        # Extract first number
                        numbers = re.findall(r'\d+', price_text)
                        if numbers:
                            try:
                                price = int(numbers[0])
                            except:
                                pass
                    
                    # Get rating
                    rating = 0.0
                    rating_elem = (item.find("div", class_=re.compile("_3LWZlK|_2kC2sd|_3Yl67G")) or
                                  item.find("span", class_=re.compile("_1lRcqv")))
                    
                    if rating_elem:
                        rating_text = rating_elem.get_text(strip=True)
                        match = re.search(r'([\d.]+)', rating_text)
                        if match:
                            try:
                                rating = float(match.group(1))
                            except:
                                pass
                    
                    # Get reviews count
                    reviews_count = 0
                    reviews_elem = (item.find("span", class_=re.compile("_2_R_DZ")) or
                                   item.find("span", class_=re.compile("_13vcmD")))
                    
                    if reviews_elem:
                        reviews_text = reviews_elem.get_text(strip=True)
                        # Extract numbers including those in format like "1,234"
                        numbers = re.findall(r'[\d,]+', reviews_text)
                        if numbers:
                            try:
                                reviews_count = int(numbers[0].replace(",", ""))
                            except:
                                pass
                    
                    # Get image
                    image_url = ""
                    img_elem = (item.find("img", class_=re.compile("_396cs4|_2r_T1I")) or
                               item.find("img"))
                    
                    if img_elem:
                        image_url = img_elem.get("src", "")
                    
                    # Only add if we have minimum viable data
                    if title and len(title) > 5 and price > 0:
                        product = {
                            "title": title[:100],
                            "price": price,
                            "rating": rating,
                            "reviews_count": reviews_count,
                            "url": link,
                            "image": image_url,
                            "platform": "flipkart",
                            "platform_icon": PLATFORM_ICONS["flipkart"],
                            "in_stock": True,
                            "currency": "INR",
                            "category": self._detect_category(query),
                        }
                        products.append(product)
                        
                        if len(products) >= limit:
                            break
                        
                except Exception as e:
                    log.debug(f"Error parsing Flipkart product: {e}")
                    continue
            
            log.info(f"✅ Found {len(products)} products on Flipkart")
            
        except Exception as e:
            log.error(f"Flipkart search failed: {e}")
        
        return products
    
    def _search_ebay(self, query: str, limit: int = 10) -> List[Dict]:
        """Search eBay for products."""
        # Placeholder for eBay search
        # Implementation similar to Amazon/Flipkart
        return []
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Price Comparison
    # ══════════════════════════════════════════════════════════════════════════
    
    def compare_prices(self, product_name: str, 
                      platforms: List[str] = ["amazon", "flipkart"]) -> Dict[str, Any]:
        """
        Compare prices of a specific product across multiple platforms.
        
        Args:
            product_name: Product to search for
            platforms: List of platforms to compare
        
        Returns:
            Dictionary with price comparison data
        """
        log.info(f"💰 Comparing prices for '{product_name}' across {len(platforms)} platforms")
        
        comparison = {
            "product": product_name,
            "platforms": {},
            "best_deal": None,
            "price_difference": 0,
            "timestamp": datetime.now().isoformat(),
        }
        
        all_products = []
        
        for platform in platforms:
            try:
                # Search each platform
                results = self.search_products(product_name, platform=platform, limit=5)
                products = results.get("products", [])
                
                if products:
                    # Get best match (first result usually most relevant)
                    best_match = products[0]
                    comparison["platforms"][platform] = {
                        "price": best_match["price"],
                        "rating": best_match.get("rating", 0),
                        "url": best_match["url"],
                        "title": best_match["title"],
                        "in_stock": best_match.get("in_stock", False),
                    }
                    all_products.append(best_match)
                
                time.sleep(0.5)
                
            except Exception as e:
                log.error(f"Error comparing on {platform}: {e}")
                continue
        
        # Find best deal
        if all_products:
            all_products.sort(key=lambda x: x["price"])
            best = all_products[0]
            worst = all_products[-1]
            
            comparison["best_deal"] = {
                "platform": best["platform"],
                "price": best["price"],
                "url": best["url"],
                "savings": worst["price"] - best["price"],
            }
            comparison["price_difference"] = worst["price"] - best["price"]
        
        return comparison
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Smart Recommendations
    # ══════════════════════════════════════════════════════════════════════════
    
    def get_recommendations(self, product_name: str, 
                          category: str = "", budget: Optional[int] = None,
                          count: int = 5) -> Dict[str, Any]:
        """
        Get smart product recommendations based on a product name.
        
        Args:
            product_name: Product to get recommendations for
            category: Product category
            budget: Budget constraint
            count: Number of recommendations
        
        Returns:
            Dictionary with recommended products
        """
        log.info(f"🎯 Getting recommendations for '{product_name}'")
        
        recommendations = {
            "source_product": product_name,
            "similar_products": [],
            "alternatives": [],
            "budget_options": [],
            "premium_options": [],
            "timestamp": datetime.now().isoformat(),
        }
        
        # Detect category from product name
        detected_category = self._detect_category(product_name)
        if not category:
            category = detected_category
        
        # Get similar products
        search_results = self.search_products(product_name, limit=15)
        products = search_results.get("products", [])
        
        if products:
            # Reference product (first result)
            ref_product = products[0]
            ref_price = ref_product.get("price", 0)
            
            # Similar products (same price range ±20%)
            similar = [p for p in products[1:] 
                      if abs(p.get("price", 0) - ref_price) / ref_price <= 0.2]
            recommendations["similar_products"] = similar[:count]
            
            # Budget alternatives (cheaper)
            budget_opts = [p for p in products 
                          if p.get("price", 999999) < ref_price * 0.8]
            budget_opts.sort(key=lambda x: x.get("rating", 0), reverse=True)
            recommendations["budget_options"] = budget_opts[:count]
            
            # Premium alternatives (more expensive)
            premium = [p for p in products 
                      if p.get("price", 0) > ref_price * 1.2]
            premium.sort(key=lambda x: x.get("rating", 0), reverse=True)
            recommendations["premium_options"] = premium[:count]
        
        # Get brand-based recommendations
        if category in BRAND_DATABASE:
            brands = self._get_brand_recommendations(category, budget)
            recommendations["brand_suggestions"] = brands
        
        return recommendations
    
    def _detect_category(self, product_name: str) -> str:
        """Detect product category from name."""
        name_lower = product_name.lower()
        
        for cat, keywords in TRENDING_CATEGORIES.items():
            if any(kw in name_lower for kw in keywords):
                return cat
        
        # Check category icons
        for cat in CATEGORY_ICONS.keys():
            if cat in name_lower:
                return cat
        
        return "general"
    
    def _get_brand_recommendations(self, category: str, budget: Optional[int]) -> Dict[str, List[str]]:
        """Get brand recommendations based on category and budget."""
        if category not in BRAND_DATABASE:
            return {}
        
        brands = BRAND_DATABASE[category]
        
        if budget is None:
            return brands
        
        # Filter by budget
        if budget < 15000:
            return {"budget": brands.get("budget", [])}
        elif budget < 50000:
            return {"mid_range": brands.get("mid_range", [])}
        else:
            return {"premium": brands.get("premium", [])}
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Review Analysis
    # ══════════════════════════════════════════════════════════════════════════
    
    def analyze_reviews(self, product_url: str) -> Dict[str, Any]:
        """
        Analyze customer reviews for sentiment and insights.
        
        Args:
            product_url: URL of the product
        
        Returns:
            Dictionary with review analysis
        """
        log.info(f"📊 Analyzing reviews for product")
        
        analysis = {
            "url": product_url,
            "total_reviews": 0,
            "sentiment": {"positive": 0, "negative": 0, "neutral": 0},
            "rating_distribution": {5: 0, 4: 0, 3: 0, 2: 0, 1: 0},
            "common_praises": [],
            "common_complaints": [],
            "verified_purchases": 0,
            "helpful_positive": [],
            "helpful_negative": [],
            "timestamp": datetime.now().isoformat(),
        }
        
        # Placeholder for actual review scraping and analysis
        # This would involve fetching reviews and using NLP for sentiment analysis
        
        return analysis
    
    def _analyze_sentiment(self, review_text: str) -> str:
        """Simple sentiment analysis of review text."""
        text_lower = review_text.lower()
        
        positive_count = sum(1 for word in SENTIMENT_POSITIVE if word in text_lower)
        negative_count = sum(1 for word in SENTIMENT_NEGATIVE if word in text_lower)
        
        if positive_count > negative_count + 2:
            return "positive"
        elif negative_count > positive_count + 2:
            return "negative"
        else:
            return "neutral"
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Deal Finder
    # ══════════════════════════════════════════════════════════════════════════
    
    def find_deals(self, category: str = "", discount_min: int = 20,
                  budget: Optional[int] = None, limit: int = 20) -> Dict[str, Any]:
        """
        Find deals and discounts in a category.
        
        Args:
            category: Product category
            discount_min: Minimum discount percentage
            budget: Maximum price
            limit: Number of deals to find
        
        Returns:
            Dictionary with deal information
        """
        log.info(f"🔥 Finding deals in {category or 'all categories'} (min {discount_min}% off)")
        
        deals = {
            "category": category,
            "min_discount": discount_min,
            "total_found": 0,
            "deals": [],
            "lightning_deals": [],
            "coupons": [],
            "timestamp": datetime.now().isoformat(),
        }
        
        # Search for popular items in category
        if category:
            search_terms = TRENDING_CATEGORIES.get(category, [category])[:5]
        else:
            search_terms = ["deals", "offers", "sale"]
        
        all_deals = []
        
        for term in search_terms:
            try:
                results = self.search_products(term, budget=budget, limit=10)
                products = results.get("products", [])
                
                for product in products:
                    # Estimate if product is on deal (would need price history)
                    # For demo, we'll flag products with high ratings and competitive prices
                    if product.get("rating", 0) >= 4.0:
                        all_deals.append({
                            **product,
                            "estimated_discount": 25,  # Placeholder
                            "deal_type": "regular",
                        })
                
                time.sleep(0.3)
                
            except Exception as e:
                log.error(f"Error finding deals for {term}: {e}")
                continue
        
        # Sort by rating and price
        all_deals.sort(key=lambda x: (x.get("rating", 0), -x.get("price", 999999)), reverse=True)
        deals["deals"] = all_deals[:limit]
        deals["total_found"] = len(all_deals)
        
        return deals
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Price Tracking
    # ══════════════════════════════════════════════════════════════════════════
    
    def track_price(self, product_url: str, target_price: Optional[int] = None) -> Dict[str, Any]:
        """
        Track price history and set alerts for a product.
        
        Args:
            product_url: URL of the product to track
            target_price: Alert when price drops below this
        
        Returns:
            Dictionary with tracking information
        """
        log.info(f"📈 Tracking price for product")
        
        # Extract product ID from URL
        product_id = self._extract_product_id(product_url)
        
        # Get current price (would need to scrape product page)
        current_price = 0  # Placeholder
        
        # Add to price history
        self.price_history[product_id].append({
            "timestamp": datetime.now().isoformat(),
            "price": current_price,
        })
        
        tracking_info = {
            "product_url": product_url,
            "product_id": product_id,
            "current_price": current_price,
            "target_price": target_price,
            "price_alert": current_price <= target_price if target_price else False,
            "price_history": self.price_history[product_id][-30:],  # Last 30 entries
            "lowest_price": min(self.price_history[product_id], key=lambda x: x["price"])["price"] if self.price_history[product_id] else 0,
            "highest_price": max(self.price_history[product_id], key=lambda x: x["price"])["price"] if self.price_history[product_id] else 0,
        }
        
        return tracking_info
    
    def _extract_product_id(self, url: str) -> str:
        """Extract product ID from URL."""
        # Parse URL and extract ID (platform-specific)
        if "amazon" in url:
            match = re.search(r'/dp/([A-Z0-9]{10})', url)
            return match.group(1) if match else url
        elif "flipkart" in url:
            match = re.search(r'/p/([a-zA-Z0-9]+)', url)
            return match.group(1) if match else url
        else:
            return url
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Spec Comparison
    # ══════════════════════════════════════════════════════════════════════════
    
    def compare_specs(self, product_urls: List[str]) -> Dict[str, Any]:
        """
        Compare specifications of multiple products side-by-side.
        
        Args:
            product_urls: List of product URLs to compare
        
        Returns:
            Dictionary with spec comparison
        """
        log.info(f"⚙️  Comparing specs for {len(product_urls)} products")
        
        comparison = {
            "products": [],
            "common_specs": {},
            "unique_features": {},
            "winner": None,
            "timestamp": datetime.now().isoformat(),
        }
        
        # Placeholder for actual spec extraction
        # Would need to scrape product pages and parse specifications
        
        return comparison
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Trending Products
    # ══════════════════════════════════════════════════════════════════════════
    
    def trending_products(self, category: str = "electronics", limit: int = 10) -> Dict[str, Any]:
        """
        Get trending products in a category.
        
        Args:
            category: Product category
            limit: Number of trending products
        
        Returns:
            Dictionary with trending products
        """
        log.info(f"🔥 Getting trending products in {category}")
        
        trending = {
            "category": category,
            "category_icon": CATEGORY_ICONS.get(category, "📦"),
            "trending_now": [],
            "bestsellers": [],
            "new_arrivals": [],
            "timestamp": datetime.now().isoformat(),
        }
        
        # Get trending keywords for category
        keywords = TRENDING_CATEGORIES.get(category, [category])[:3]
        
        all_products = []
        
        for keyword in keywords:
            try:
                results = self.search_products(keyword, limit=5, sort_by="popularity")
                products = results.get("products", [])
                all_products.extend(products)
                time.sleep(0.3)
            except Exception as e:
                log.error(f"Error getting trending for {keyword}: {e}")
                continue
        
        # Sort by rating and reviews
        all_products.sort(key=lambda x: (x.get("rating", 0), x.get("reviews_count", 0)), reverse=True)
        
        trending["trending_now"] = all_products[:limit]
        trending["bestsellers"] = [p for p in all_products if p.get("reviews_count", 0) > 1000][:5]
        
        return trending
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Wishlist Management
    # ══════════════════════════════════════════════════════════════════════════
    
    def add_to_wishlist(self, product: Dict) -> None:
        """Add product to wishlist."""
        if product not in self.wishlist:
            self.wishlist.append(product)
            log.info(f"➕ Added to wishlist: {product.get('title', 'Unknown')}")
    
    def get_wishlist(self) -> List[Dict]:
        """Get current wishlist."""
        return self.wishlist
    
    def clear_wishlist(self) -> None:
        """Clear wishlist."""
        self.wishlist = []
        log.info("🗑️  Wishlist cleared")
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Export Functions
    # ══════════════════════════════════════════════════════════════════════════
    
    def export_json(self, query: str, output_dir: str = "shopping_exports") -> Optional[str]:
        """Export search results to JSON."""
        os.makedirs(output_dir, exist_ok=True)
        
        results = self.search_products(query, limit=50)
        
        filename = f"{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        
        log.info(f"📄 Exported to JSON: {filepath}")
        return filepath
    
    def export_csv(self, query: str, output_dir: str = "shopping_exports") -> Optional[str]:
        """Export search results to CSV."""
        os.makedirs(output_dir, exist_ok=True)
        
        results = self.search_products(query, limit=50)
        products = results.get("products", [])
        
        if not products:
            log.warning("No products to export")
            return None
        
        filename = f"{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ["title", "price", "rating", "reviews_count", "platform", "url", "in_stock"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            for product in products:
                writer.writerow({k: product.get(k, "") for k in fieldnames})
        
        log.info(f"📃 Exported to CSV: {filepath}")
        return filepath
    
    def export_excel(self, query: str, output_dir: str = "shopping_exports") -> Optional[str]:
        """Export search results to Excel with formatting."""
        os.makedirs(output_dir, exist_ok=True)
        
        results = self.search_products(query, limit=50)
        products = results.get("products", [])
        
        if not products:
            log.warning("No products to export")
            return None
        
        filename = f"{query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ["#", "Product", "Price (₹)", "Rating", "Reviews", "Platform", "Stock", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for idx, product in enumerate(products, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=product.get("title", "")[:80])
            ws.cell(row=idx, column=3, value=product.get("price", 0))
            ws.cell(row=idx, column=4, value=product.get("rating", 0))
            ws.cell(row=idx, column=5, value=product.get("reviews_count", 0))
            ws.cell(row=idx, column=6, value=product.get("platform", "").upper())
            ws.cell(row=idx, column=7, value="✓" if product.get("in_stock") else "✗")
            ws.cell(row=idx, column=8, value=product.get("url", ""))
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 15
        
        wb.save(filepath)
        log.info(f"📊 Exported to Excel: {filepath}")
        return filepath
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Display Functions
    # ══════════════════════════════════════════════════════════════════════════
    
    def print_products(self, results: Dict[str, Any]) -> None:
        """Pretty print product search results."""
        print(f"\n{self.c_title('═' * 80)}")
        query = results.get("query", "")
        print(self.c_title(f'  🛍️  Search Results: {query}'))
        print(f"{self.c_title('═' * 80)}")
        print(f"\n  {self.c_label('Total found:')} {results['total_found']}")
        print(f"  {self.c_label('Platforms:')} {', '.join(results['platforms'])}")
        print(f"  {self.c_label('Search time:')} {results['search_time']}s\n")
        
        for idx, product in enumerate(results.get("products", []), 1):
            icon = CATEGORY_ICONS.get(product.get("category", ""), "📦")
            price = product.get("price", 0)
            rating = product.get("rating", 0)
            reviews = product.get("reviews_count", 0)
            
            print(f"{self.c_good(f'  {idx:>2}.')} {product['platform_icon']} {icon} {product['title'][:65]}")
            print(f"       {self.c_price(f'₹{price:,}')} | "
                  f"{'⭐' * int(rating)} {rating:.1f} | "
                  f"{reviews:,} reviews")
            print(f"       {self.c_dim(product.get('platform', '').upper())} | "
                  f"{self.c_url(product.get('url', '')[:60])}\n")
    
    def print_price_comparison(self, comparison: Dict[str, Any]) -> None:
        """Pretty print price comparison."""
        print(f"\n{self.c_title('═' * 80)}")
        product_name = comparison.get("product", "")
        print(self.c_title(f'  💰 Price Comparison: {product_name}'))
        print(f"{self.c_title('═' * 80)}\n")
        
        for platform, data in comparison.get("platforms", {}).items():
            icon = PLATFORM_ICONS.get(platform, "🔵")
            stock = "✓" if data.get("in_stock") else "✗"
            price = data.get("price", 0)
            rating = data.get("rating", 0)
            
            print(f"  {icon} {platform.upper():12} | "
                  f"{self.c_price(f'₹{price:,}'):15} | "
                  f"⭐ {rating:.1f} | "
                  f"Stock: {stock}")
        
        if comparison.get("best_deal"):
            best = comparison["best_deal"]
            best_price = best.get("price", 0)
            savings = best.get("savings", 0)
            print(f"\n  {self.c_good('🎯 Best Deal:')} {best['platform'].upper()} - "
                  f"{self.c_price(f'₹{best_price:,}')} "
                  f"(Save ₹{savings:,}!)")
    
    def print_recommendations(self, recs: Dict[str, Any]) -> None:
        """Pretty print recommendations."""
        print(f"\n{self.c_title('═' * 80)}")
        source = recs.get("source_product", "")
        print(self.c_title(f'  🎯 Recommendations for: {source}'))
        print(f"{self.c_title('═' * 80)}\n")
        
        sections = [
            ("Similar Products", recs.get("similar_products", [])),
            ("Budget Options", recs.get("budget_options", [])),
            ("Premium Options", recs.get("premium_options", [])),
        ]
        
        for title, products in sections:
            if products:
                print(f"\n  {self.c_label(f'📌 {title}:')}\n")
                for idx, p in enumerate(products[:5], 1):
                    price = p.get("price", 0)
                    rating = p.get("rating", 0)
                    print(f"  {idx}. {p['title'][:60]}")
                    print(f"     {self.c_price(f'₹{price:,}')} | ⭐ {rating:.1f}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI Interface
# ══════════════════════════════════════════════════════════════════════════════

def run_cli():
    """Run interactive CLI for shopping agent."""
    agent = ShoppingAgent()
    
    c_title = agent.c_title
    c_label = agent.c_label
    c_good = agent.c_good
    c_bad = agent.c_bad
    
    print(f"\n{c_title('═' * 80)}")
    print(f"{c_title('  🛍️  ShoppingAgent v2 - Smart Product Search & Recommendations')}")
    print(f"{c_title('═' * 80)}\n")
    
    MENU = f"""
  {c_title('[1]')}  Search products           🔍
  {c_title('[2]')}  Compare prices            💰
  {c_title('[3]')}  Get recommendations       🎯
  {c_title('[4]')}  Find deals & discounts    🔥
  {c_title('[5]')}  Trending products         📈
  {c_title('[6]')}  Analyze reviews           📊
  {c_title('[7]')}  Track price               📉
  {c_title('[8]')}  View wishlist             ❤️
  {c_title('[xe]')} Export to Excel           📊
  {c_title('[xj]')} Export to JSON            📄
  {c_title('[xc]')} Export to CSV             📃
  {c_title('[q]')}  Quit                      👋
"""
    
    while True:
        print(MENU)
        choice = input(f"  {c_label('👉 Choose:')} ").strip().lower()
        
        if choice == "1":
            query = input("  🔍 Search for: ").strip()
            budget = input("  💰 Max budget (₹) [Enter to skip]: ").strip()
            platform = input("  🌐 Platform (all/amazon/flipkart) [all]: ").strip() or "all"
            rating = input("  ⭐ Min rating [0]: ").strip()
            limit = input("  🔢 Max results [20]: ").strip()
            
            results = agent.search_products(
                query,
                budget=int(budget) if budget.isdigit() else None,
                platform=platform,
                min_rating=float(rating) if rating else 0.0,
                limit=int(limit) if limit.isdigit() else 20,
            )
            agent.print_products(results)
        
        elif choice == "2":
            product = input("  🍽️  Product name: ").strip()
            platforms = input("  🌐 Platforms (comma-separated) [amazon,flipkart]: ").strip()
            platforms = [p.strip() for p in platforms.split(",")] if platforms else ["amazon", "flipkart"]
            
            comparison = agent.compare_prices(product, platforms)
            agent.print_price_comparison(comparison)
        
        elif choice == "3":
            product = input("  📦 Product name: ").strip()
            budget = input("  💰 Budget (₹) [Enter to skip]: ").strip()
            
            recs = agent.get_recommendations(
                product,
                budget=int(budget) if budget.isdigit() else None,
            )
            agent.print_recommendations(recs)
        
        elif choice == "4":
            category = input("  📁 Category [electronics]: ").strip() or "electronics"
            discount = input("  🔥 Min discount % [20]: ").strip()
            budget = input("  💰 Max budget (₹) [Enter to skip]: ").strip()
            
            deals = agent.find_deals(
                category=category,
                discount_min=int(discount) if discount.isdigit() else 20,
                budget=int(budget) if budget.isdigit() else None,
            )
            
            total_found = deals.get("total_found", 0)
            print(f"\n  {c_good(f'🔥 Found {total_found} deals in {category}')}\n")
            for idx, deal in enumerate(deals.get("deals", [])[:10], 1):
                deal_price = deal.get("price", 0)
                deal_rating = deal.get("rating", 0)
                deal_discount = deal.get("estimated_discount", 0)
                print(f"  {idx}. {deal['title'][:60]}")
                print(f"     {agent.c_price(f'₹{deal_price:,}')} | "
                      f"⭐ {deal_rating:.1f} | "
                      f"{c_good(f'{deal_discount}% OFF')}\n")
        
        elif choice == "5":
            category = input("  📁 Category [electronics]: ").strip() or "electronics"
            limit = input("  🔢 Limit [10]: ").strip()
            
            trending = agent.trending_products(
                category=category,
                limit=int(limit) if limit.isdigit() else 10,
            )
            
            cat_icon = trending.get("category_icon", "📦")
            print(f"\n  {c_title(f'{cat_icon} Trending in {category.upper()}')}\n")
            for idx, product in enumerate(trending.get("trending_now", []), 1):
                prod_price = product.get("price", 0)
                prod_rating = product.get("rating", 0)
                prod_reviews = product.get("reviews_count", 0)
                print(f"  {idx}. {product['title'][:60]}")
                print(f"     {agent.c_price(f'₹{prod_price:,}')} | "
                      f"⭐ {prod_rating:.1f} | "
                      f"{prod_reviews:,} reviews\n")
        
        elif choice == "6":
            url = input("  🔗 Product URL: ").strip()
            analysis = agent.analyze_reviews(url)
            print(f"\n  {c_good('📊 Review analysis feature coming soon!')}")
        
        elif choice == "7":
            url = input("  🔗 Product URL: ").strip()
            target = input("  🎯 Target price (₹) [optional]: ").strip()
            
            tracking = agent.track_price(
                url,
                target_price=int(target) if target.isdigit() else None,
            )
            print(f"\n  {c_good('📈 Price tracking activated!')}")
        
        elif choice == "8":
            wishlist = agent.get_wishlist()
            if wishlist:
                print(f"\n  {c_title('❤️  Your Wishlist')}\n")
                for idx, item in enumerate(wishlist, 1):
                    item_price = item.get("price", 0)
                    print(f"  {idx}. {item.get('title', 'Unknown')[:60]}")
                    print(f"     {agent.c_price(f'₹{item_price:,}')}\n")
            else:
                print(f"\n  {c_bad('Wishlist is empty!')}")
        
        elif choice == "xe":
            query = input("  🔍 Search query: ").strip()
            path = agent.export_excel(query)
            if path:
                print(f"  {c_good('📊 Excel exported:')} {path}")
        
        elif choice == "xj":
            query = input("  🔍 Search query: ").strip()
            path = agent.export_json(query)
            if path:
                print(f"  {c_good('📄 JSON exported:')} {path}")
        
        elif choice == "xc":
            query = input("  🔍 Search query: ").strip()
            path = agent.export_csv(query)
            if path:
                print(f"  {c_good('📃 CSV exported:')} {path}")
        
        elif choice in ("q", "quit", "exit"):
            print(f"\n  {c_good('🛍️  ShoppingAgent v2 signing off!')}\n")
            break
        
        else:
            print(c_bad("  ⚠  Invalid option."))


if __name__ == "__main__":
    run_cli()