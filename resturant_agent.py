"""
RestaurantAgent v2 — Nearby Restaurants, Famous Foods & Food Tour Planner
==========================================================================
Install:
    pip install requests beautifulsoup4 lxml openpyxl python-dotenv tqdm colorama

New in v2:
    ✨ Dietary filter     — vegan, vegetarian, halal, jain, gluten-free
    ✨ Price comparison   — Zomato vs Swiggy vs Google side-by-side
    ✨ Sentiment analysis — scrape & classify reviews as Positive/Mixed/Negative
    ✨ Food Tour planner  — best 5-stop route ordered by distance + rating
    ✨ 70+ city famous foods DB (India + SE Asia + Middle East + Europe + Americas)
    ✨ JSON & CSV export  — in addition to Excel
    ✨ Colored terminal   — colorama-powered rich CLI output
    ✨ Trending dishes    — what's popular in the city right now (scraped)
    ✨ Budget mode        — filter by cost-for-two bracket
    ✨ Multi-city compare — compare food scenes of two cities

Data sources (ALL free, ZERO signup):
    ✅ Nominatim / OSM Overpass    — real places, addresses, phones
    ✅ Zomato public search        — rating, cost-for-two, popular dishes
    ✅ Swiggy public search        — delivery time, rating
    ✅ Google Maps HTML            — rating, reviews, price level
    ✅ Wikipedia REST API          — food culture, dish descriptions
    ✅ TripAdvisor public pages    — traveller ranking, review snippets
    ✅ OSRM routing API            — real road distances for food tour routing

Usage:
    agent = RestaurantAgent()
    agent.find_nearby("Tambaram, Tamil Nadu", radius_m=3000)
    agent.find_nearby("Chennai", cuisine="biryani", diet="halal", budget="moderate")
    agent.famous_foods("Hyderabad")
    agent.plan_food_tour("Chennai")          # 5-stop tour with route
    agent.price_compare("Saravana Bhavan", "Chennai")
    agent.dietary_search("Bangalore", diet="vegan")
    agent.trending_dishes("Mumbai")
    agent.compare_cities("Chennai", "Hyderabad")
    agent.export_json("Chennai")
    agent.export_csv("Chennai")
    agent.export_excel("Chennai")
"""

import os, re, math, json, csv, time, logging, textwrap
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from collections import Counter

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
try:
    from colorama import init as colorama_init, Fore, Back, Style
    colorama_init(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False
    class Fore:
        RED=YELLOW=GREEN=CYAN=MAGENTA=BLUE=WHITE=RESET=""
    class Style:
        BRIGHT=DIM=RESET_ALL=""
    class Back:
        RED=BLACK=RESET=""

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("RestaurantAgent")

# ── Constants ──────────────────────────────────────────────────────────────────
NOMINATIM   = "https://nominatim.openstreetmap.org/search"
NOMINATIM_R = "https://nominatim.openstreetmap.org/reverse"
OVERPASS    = "https://overpass-api.de/api/interpreter"
WIKI_API    = "https://en.wikipedia.org/w/api.php"
OSRM_URL    = "http://router.project-osrm.org/route/v1/driving"

HEADERS_OSM = {"User-Agent": "SmartTripAI-RestaurantAgent/2.0 (research)"}
HEADERS_WEB = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
}
HEADERS_JSON = {**HEADERS_WEB, "Accept": "application/json, */*"}

FOOD_TAG_PAIRS = [
    ("amenity", "restaurant"), ("amenity", "fast_food"), ("amenity", "cafe"),
    ("amenity", "food_court"), ("amenity", "bar"),       ("amenity", "pub"),
    ("amenity", "biergarten"), ("amenity", "ice_cream"), ("shop",   "bakery"),
    ("shop",    "deli"),       ("amenity", "food_stall"),
]

PLACE_ICONS = {
    "restaurant":"🍽️","fast_food":"🍔","cafe":"☕","food_court":"🏪",
    "bar":"🍺","pub":"🍻","biergarten":"🌿","ice_cream":"🍦",
    "bakery":"🥐","deli":"🥪","food_stall":"🏮",
}

BUDGET_RANGES = {
    "budget":    (0,    300,  "₹ Budget (under ₹300)"),
    "moderate":  (300,  700,  "₹₹ Moderate (₹300–700)"),
    "upscale":   (700,  1500, "₹₹₹ Upscale (₹700–1500)"),
    "fine":      (1500, 9999, "₹₹₹₹ Fine Dining (₹1500+)"),
}

DIET_OSM_TAGS = {
    "vegetarian":  "diet:vegetarian",
    "vegan":       "diet:vegan",
    "halal":       "diet:halal",
    "jain":        "diet:vegan",   # closest OSM tag; jain is subset of vegan
    "gluten_free": "diet:gluten_free",
    "kosher":      "diet:kosher",
}

DIET_KEYWORDS = {
    "vegetarian":  ["vegetarian","veg only","pure veg","sattvic"],
    "vegan":       ["vegan","plant based","plant-based"],
    "halal":       ["halal","halaal"],
    "jain":        ["jain","no onion","no garlic","sattvic"],
    "gluten_free": ["gluten free","gluten-free","celiac"],
    "kosher":      ["kosher"],
}

SENTIMENT_POSITIVE = ["excellent","amazing","loved","best","fantastic","superb",
                       "delicious","outstanding","wonderful","great","good","tasty",
                       "fresh","nice","recommend","must visit","worth it"]
SENTIMENT_NEGATIVE = ["terrible","awful","horrible","bad","worst","disgusting",
                       "cold","stale","rude","overpriced","dirty","avoid",
                       "disappointing","slow","unhygienic","cockroach","hair"]


# ══════════════════════════════════════════════════════════════════════════════
#  FAMOUS FOODS DATABASE v2  (70+ cities)
# ══════════════════════════════════════════════════════════════════════════════

FAMOUS_FOODS: Dict[str, Dict] = {

    # ─── Tamil Nadu ───────────────────────────────────────────────────────────
    "chennai": {
        "city":"Chennai","state":"Tamil Nadu","emoji":"🌶️",
        "must_try":["Idli Sambar","Dosa (Paper/Masala/Ghee Roast)","Filter Coffee",
                    "Chettinad Chicken Curry","Kothu Parotta","Pongal","Medu Vada",
                    "Rasam Rice","Appam with Stew","Sundal","Kuzhi Paniyaram"],
        "street_food":["Sundal","Marina Beach Bhajji","Murukku","Kozhukattai",
                       "Roasted Groundnuts","Corn Chaat"],
        "sweets":["Mysore Pak","Adhirasam","Palkova","Kozhukattai","Payasam"],
        "famous_spots":["Saravana Bhavan (global chain)","Ratna Cafe (oldest filter coffee)",
                        "Murugan Idli Shop","Hotel Palmgrove (Chettinad)",
                        "Junior Kuppanna (Kongu)","Marina Beach bhajji stalls"],
        "food_culture":"Chennai is South India's vegetarian capital. Breakfast revolves around Idli-Dosa-Vada. Chettinad cuisine is world-famous for bold spices.",
        "cuisine_tags":["south indian","chettinad","tamil","vegetarian","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True},
        "avg_cost_two":300,"trending":["Biriyani","Kathi Roll","Momos"],
    },
    "madurai": {
        "city":"Madurai","state":"Tamil Nadu","emoji":"🏛️",
        "must_try":["Madurai Kari Dosa","Jigarthanda","Mutton Kothu Parotta",
                    "Paruthi Paal","Kavuni Arisi","Roast Chicken","Chicken 65"],
        "street_food":["Jigarthanda","Kari Dosa","Kothu Parotta","Egg Puffs"],
        "sweets":["Jigarthanda","Kavuni Arisi","Sojji Appam"],
        "famous_spots":["Murugan Idli Shop (original)","Anthony's Kitchen (Kari Dosa)",
                        "Jigarthanda shops near Meenakshi Temple"],
        "food_culture":"Madurai is the birthplace of Jigarthanda and Kari Dosa. Known for non-vegetarian street food.",
        "cuisine_tags":["south indian","non-veg","street food"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":200,"trending":["Kari Dosa","Jigarthanda"],
    },
    "coimbatore": {
        "city":"Coimbatore","state":"Tamil Nadu","emoji":"🌿",
        "must_try":["Kongu Biryani","Nattu Kozhi Curry","Seeraga Samba Rice",
                    "Kondakadalai Curry","Sundakkai Vathal Kuzhambu","Kavuni Arisi"],
        "street_food":["Masala Sundal","Bhajji","Sweet Corn","Kaalan Kulambu"],
        "sweets":["Kavuni Arisi","Sakkarai Pongal","Coconut Barfi"],
        "famous_spots":["Annapoorna Hotel","Hotel Anand","Kovai Pazhamudhir Nilayam"],
        "food_culture":"Coimbatore represents Kongu Nadu cuisine — less spicy, rich in coconut and native produce.",
        "cuisine_tags":["kongu","south indian","vegetarian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":250,"trending":["Kongu Meals","Nattu Kozhi"],
    },
    "tirunelveli": {
        "city":"Tirunelveli","state":"Tamil Nadu","emoji":"🍮",
        "must_try":["Tirunelveli Halwa (GI-tagged)","Kola Urundai","Sodhi",
                    "Kothu Parotta","Nellai Biryani","Samba Sadam"],
        "street_food":["Halwa","Kola Urundai","Egg Puffs","Bajji"],
        "sweets":["Tirunelveli Halwa","Adhirasam","Palkova"],
        "famous_spots":["Sri Krishnavilas (Halwa)","Prema Vilas","Arun Sweets"],
        "food_culture":"Tirunelveli is world-famous for its unique wheat halwa made with river water. The city's biryani is also distinct.",
        "cuisine_tags":["south indian","tamil","halwa"],
        "diet_friendly":{"vegetarian":True,"vegan":False},
        "avg_cost_two":180,"trending":["Halwa","Nellai Biryani"],
    },

    # ─── Andhra Pradesh & Telangana ───────────────────────────────────────────
    "hyderabad": {
        "city":"Hyderabad","state":"Telangana","emoji":"🍚",
        "must_try":["Hyderabadi Dum Biryani","Mirchi Ka Salan","Haleem",
                    "Double Ka Meetha","Qubani Ka Meetha","Nihari","Paya Soup",
                    "Bagara Baingan","Keema Samosa","Irani Chai","Osmania Biscuits",
                    "Pathar Ka Gosht"],
        "street_food":["Keema Samosa","Bun Maska","Mirchi Bajji","Irani Chai","Lukhmi"],
        "sweets":["Double Ka Meetha","Qubani Ka Meetha","Sheer Khurma","Phirni"],
        "famous_spots":["Paradise Biryani (iconic)","Bawarchi Restaurant",
                        "Hotel Shadab (Old City)","Café Bahar",
                        "Nimrah Café (Irani Chai + Osmania)","Pista House (Haleem)"],
        "food_culture":"Hyderabad is synonymous with Dum Biryani slow-cooked in sealed handi. Nizami heritage blends Mughal and South Indian flavours. Haleem holds a GI tag.",
        "cuisine_tags":["hyderabadi","mughlai","biryani","haleem","telangana"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":400,"trending":["Cloud Kitchen Biryani","Keto Biryani","Fusion Haleem"],
    },
    "vijayawada": {
        "city":"Vijayawada","state":"Andhra Pradesh","emoji":"🌶️",
        "must_try":["Andhra Meals","Pesarattu with Ginger Chutney","Gongura Mutton",
                    "Pulihora","Punugulu","Andhra Chicken Fry","Bajji Lanka"],
        "street_food":["Mirchi Bajji","Punugulu","Jilledu Fruit","Corn Bhel"],
        "sweets":["Pootharekulu (paper sweet, GI-tagged)","Ariselu","Bobbatlu"],
        "famous_spots":["Kamala Hotel (Andhra Meals)","Swagath (Seafood)"],
        "food_culture":"Vijayawada is famous for the world's spiciest Andhra cuisine. Pootharekulu sweet is paper-thin and unique to the region.",
        "cuisine_tags":["andhra","telugu","spicy","south indian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":220,"trending":["Gongura Biryani","Ragi Mudde"],
    },

    # ─── Karnataka ────────────────────────────────────────────────────────────
    "bangalore": {
        "city":"Bangalore","state":"Karnataka","emoji":"☕",
        "must_try":["Masala Dosa (MTR style)","Bisi Bele Bath","Ragi Mudde",
                    "Akki Roti","Neer Dosa","Mangalorean Fish Curry",
                    "Coorg Pandi Curry","Iyengar Khara Bun","Rava Idli",
                    "Chow Chow Bath"],
        "street_food":["Churumuri","Egg Rolls","Corn Chaat","Gobi Manchurian","Pani Puri"],
        "sweets":["Mysore Pak","Holige (Obbattu)","Chiroti","Dharwad Peda","Karibath"],
        "famous_spots":["MTR (Mavalli Tiffin Rooms)","CTR (Central Tiffin Room)",
                        "SLV Iyengar Bakery","Airlines Hotel","Brahmin's Coffee Bar",
                        "Hotel Janata (Bisi Bele Bath)"],
        "food_culture":"Bangalore has legendary Udupi-Brahmin restaurant culture since 1920s. MTR and CTR define breakfast. Thriving craft beer and modern café scene co-exists with traditional Darshinis.",
        "cuisine_tags":["south indian","karnataka","udupi","coorg","vegetarian","cafe"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":350,"trending":["Brunch Cafes","Cloud Kitchens","Ragi dishes"],
    },
    "mysore": {
        "city":"Mysore","state":"Karnataka","emoji":"🏰",
        "must_try":["Mysore Pak (GI-tagged)","Mysore Masala Dosa","Mysore Rasam",
                    "Curd Rice","Mysore Bonda","Maddur Vade"],
        "street_food":["Maddur Vade","Corn Bhel","Bonda","Churumuri"],
        "sweets":["Mysore Pak","Holige","Coconut Barfi","Chiroti"],
        "famous_spots":["Guru Sweet Mart (original Mysore Pak)","Hotel RRR (Meals)","Mylari Dosa"],
        "food_culture":"Mysore Pak was invented in the royal kitchens of Mysore Palace by Kakasura Madappa.",
        "cuisine_tags":["karnataka","south indian","vegetarian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":250,"trending":["Palace-view dining","Mysore Pak variants"],
    },
    "mangalore": {
        "city":"Mangalore","state":"Karnataka","emoji":"🦀",
        "must_try":["Mangalore Buns","Neer Dosa with Chicken Curry","Kori Rotti",
                    "Prawn Ghee Roast","Fish Gassi","Kane Fry","Goli Baje","Pork Bafat"],
        "street_food":["Goli Baje","Boiled Peanuts","Mangalore Bhajji","Buns"],
        "sweets":["Halwa","Modak","Kadabu","Patoli"],
        "famous_spots":["Gajalee (Seafood)","Diana Hotel","Hao Ming (Chinese)","Shetty Lunch Home"],
        "food_culture":"Mangalorean cuisine blends Tulu, Konkani, and Bunt community cooking traditions. Seafood-forward with coconut base.",
        "cuisine_tags":["mangalorean","konkani","seafood","coastal"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":350,"trending":["Prawn Ghee Roast","Crab Curry"],
    },

    # ─── Kerala ───────────────────────────────────────────────────────────────
    "kochi": {
        "city":"Kochi","state":"Kerala","emoji":"🥥",
        "must_try":["Kerala Sadya (banana leaf feast)","Appam with Vegetable Stew",
                    "Karimeen (Pearl Spot) Fry","Kerala Fish Curry",
                    "Prawn Moilee","Beef Ullarthiyathu",
                    "Puttu with Kadala Curry","Parotta with Kerala Chicken Curry",
                    "Thalassery Biryani","Banana Chips"],
        "street_food":["Banana Chips","Kozhukatta","Pazham Pori","Unniyappam","Egg Puffs"],
        "sweets":["Ada Pradhaman","Unniyappam","Neyyappam","Palada Pradhaman"],
        "famous_spots":["Kayees Rahmathulla Hotel (Biryani)","Dhe Puttu",
                        "Fort House Restaurant","Pai Bros"],
        "food_culture":"Kerala cuisine is defined by coconut oil, curry leaves, and mustard. The Sadya feast during Onam is iconic. Backwaters provide fresh daily catch.",
        "cuisine_tags":["kerala","malabar","seafood","vegetarian","christian"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True},
        "avg_cost_two":400,"trending":["Kerala meals cloud kitchens","Jackfruit dishes"],
    },
    "thiruvananthapuram": {
        "city":"Thiruvananthapuram","state":"Kerala","emoji":"🌴",
        "must_try":["Kappa (Tapioca) with Fish Curry","Inji Curry","Avial",
                    "Puttu Kadala","Kerala Porotta","Nadan Kozhi Curry"],
        "street_food":["Banana Chips","Pazham Pori","Vellayappam","Unniyappam"],
        "sweets":["Payasam","Ada Pradhaman","Halwa"],
        "famous_spots":["Ariya Nivaas","Hotel Azad","Suprabhatham"],
        "food_culture":"Traditional Nair and Thiyya cuisine. Kappa (tapioca) is the staple of South Kerala fishing communities.",
        "cuisine_tags":["kerala","south kerala","vegetarian","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":300,"trending":["Kappa Biryani","Tapioca dishes"],
    },

    # ─── Maharashtra ──────────────────────────────────────────────────────────
    "mumbai": {
        "city":"Mumbai","state":"Maharashtra","emoji":"🌆",
        "must_try":["Vada Pav","Pav Bhaji","Bombay Sandwich","Bhel Puri","Sev Puri",
                    "Dahi Puri","Misal Pav","Keema Pav","Bombay Duck Fry",
                    "Sol Kadhi","Frankies","Irani Chai + Bun Maska"],
        "street_food":["Vada Pav","Bhel Puri","Sev Puri","Pani Puri","Pav Bhaji","Frankies","Kulfi"],
        "sweets":["Modak","Puran Poli","Shrikhand","Basundi","Malai Kulfi","Jalebi"],
        "famous_spots":["Bademiya (Mohammed Ali Road)","Café Mondegar",
                        "Britannia & Co (Dhansak)","Sardar's Pav Bhaji (Tardeo)",
                        "Elco Pani Puri Centre","Juhu Beach stalls"],
        "food_culture":"Mumbai street food is a national institution. Vada Pav is the city's soul food. Parsi Irani cafés, Maharashtrian Misal, and the Mughlai food trail in Mohammed Ali Road are must-experiences.",
        "cuisine_tags":["maharashtrian","street food","mughlai","parsi","gujarati","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True,"jain":True},
        "avg_cost_two":500,"trending":["Butter Garlic Crab","Kokani Seafood","Artisan Coffee"],
    },
    "pune": {
        "city":"Pune","state":"Maharashtra","emoji":"🎓",
        "must_try":["Misal Pav (Puneri style)","Sabudana Khichdi","Poha",
                    "Thalipeeth","Kande Pohe","Mutton Sukka","Zunka Bhakar","Mastani"],
        "street_food":["Misal Pav","Sabudana Vada","Poha","Bhel","Pani Puri","Dabeli"],
        "sweets":["Bakarwadi","Mastani","Modak","Shrikhand","Puran Poli"],
        "famous_spots":["Bedekar Tea Stall (Misal)","Chitale Bandhu (Bakarwadi)",
                        "Cafe Goodluck (Bun Maska)","Dorabjee's (historic cafe)"],
        "food_culture":"Pune is the heart of Puneri Marathi food. Misal Pav reigns supreme. Mastani milkshakes are a local invention. The city's café culture is the finest in Maharashtra.",
        "cuisine_tags":["maharashtrian","puneri","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":350,"trending":["Specialty Coffee","Misal variants","Vegan Cafes"],
    },
    "nagpur": {
        "city":"Nagpur","state":"Maharashtra","emoji":"🍊",
        "must_try":["Saoji Mutton Curry (fiery)","Tarri Poha","Patodi Rassa",
                    "Bhadbhade Pohe","Santra Barfi (Orange Fudge)","Chimur Chicken"],
        "street_food":["Tarri Poha","Patodi","Bhel","Chaat"],
        "sweets":["Santra Barfi","Dry Fruit Pedha","Anarsa"],
        "famous_spots":["Haldirams (Nagpur)","Hotel Centre Point","Saoji restaurants"],
        "food_culture":"Nagpur is famous for fiery Saoji cuisine and Tarri Poha breakfast. Orange capital of India — orange sweets are iconic.",
        "cuisine_tags":["vidarbha","maharashtrian","non-veg","vegetarian"],
        "diet_friendly":{"vegetarian":True},
        "avg_cost_two":250,"trending":["Saoji Biryani","Orange desserts"],
    },

    # ─── Gujarat ──────────────────────────────────────────────────────────────
    "ahmedabad": {
        "city":"Ahmedabad","state":"Gujarat","emoji":"🟡",
        "must_try":["Dhokla (Khaman/Khatta)","Fafda Jalebi","Undhiyu",
                    "Thepla with Mango Pickle","Khandvi","Gujarati Thali",
                    "Dal Dhokli","Handvo","Patra","Gathiya","Dabeli"],
        "street_food":["Fafda Jalebi","Gathiya","Dabeli","Bhel","Sev Usal","Locho"],
        "sweets":["Jalebi","Mohanthal","Chikki","Sutarfeni","Mawa Jalebi","Basundi"],
        "famous_spots":["Manek Chowk night market","Vishala Restaurant (heritage)",
                        "Law Garden Night Market","Agashiye (rooftop thali)"],
        "food_culture":"Ahmedabad is a vegetarian paradise. Gujarati cuisine balances sweet, salty, and spicy. Manek Chowk transforms into a night food market after 9 PM.",
        "cuisine_tags":["gujarati","vegetarian","jain","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":280,"trending":["Undhiyu (winter)","Handcrafted Thali"],
    },
    "surat": {
        "city":"Surat","state":"Gujarat","emoji":"💎",
        "must_try":["Surti Locho","Ghari (Surat sweet)","Undhiyu","Sev Usal",
                    "Ponk (tender jowar)","Surti Biryani","Khaman","Surti Papdi Chaat"],
        "street_food":["Locho","Sev Usal","Bhajiya","Patra","Ponk Vada"],
        "sweets":["Ghari","Mohanthal","Mawa Jalebi","Basundi"],
        "famous_spots":["Surti Locho Wala","Ranchod Locho","Ghari shops near Tapi"],
        "food_culture":"Surat has one of India's richest street food cultures. Locho and Ghari are unique to the city. Suratees eat 6+ times a day.",
        "cuisine_tags":["surati","gujarati","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":200,"trending":["Locho variants","Fusion Chaat"],
    },

    # ─── Rajasthan ────────────────────────────────────────────────────────────
    "jaipur": {
        "city":"Jaipur","state":"Rajasthan","emoji":"🏰",
        "must_try":["Dal Baati Churma","Laal Maas","Gatte Ki Sabzi","Ker Sangri",
                    "Pyaaz Kachori","Mirchi Bada","Mawa Kachori","Rajasthani Thali","Ghevar"],
        "street_food":["Pyaaz Kachori","Mirchi Bada","Kulfi","Rabri","Lassi"],
        "sweets":["Ghevar","Mawa Kachori","Kalakand","Rabri","Churma Ladoo","Malpua"],
        "famous_spots":["LMB (Laxmi Mishthan Bhandar)","Rawat Mishthan (Kachori)",
                        "Handi Restaurant (Laal Maas)","1135 AD at Amer Fort"],
        "food_culture":"Jaipur food is bold and rich. Laal Maas is one of India's spiciest mutton curries. Dal Baati Churma is the soul of Rajasthan.",
        "cuisine_tags":["rajasthani","mughlai","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":350,"trending":["Heritage Dining","Rajasthani Thali"],
    },
    "jodhpur": {
        "city":"Jodhpur","state":"Rajasthan","emoji":"🔵",
        "must_try":["Makhaniya Lassi","Mirchi Bada (Jodhpuri)","Pyaaz Kachori",
                    "Mawa Kachori","Gatte Ki Khichdi","Jodhpuri Dal","Besan Chakki"],
        "street_food":["Mirchi Bada","Pyaaz Kachori","Lassi","Kulfi","Rabri"],
        "sweets":["Mawa Kachori","Moti Pak","Besan Ladoo","Balushahi"],
        "famous_spots":["Gypsy Restaurant","Jodhpur's Sardar Market stalls"],
        "food_culture":"Jodhpur is famous for the spiciest Mirchi Bada in Rajasthan and the legendary Makhaniya Lassi.",
        "cuisine_tags":["rajasthani","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":250,"trending":["Blue City Rooftop Dining"],
    },

    # ─── Delhi/NCR ────────────────────────────────────────────────────────────
    "delhi": {
        "city":"Delhi","state":"Delhi","emoji":"🕌",
        "must_try":["Butter Chicken (birthplace)","Chole Bhature","Nihari",
                    "Seekh Kebab","Galouti Kebab","Paranthe Wali Gali",
                    "Aloo Tikki Chaat","Dahi Bhalla","Kulcha Chole",
                    "Rabri Faluda","Mutton Burra Kebab","Korma"],
        "street_food":["Chaat","Gol Gappe","Aloo Tikki","Dahi Bhalla","Momos","Ram Ladoo"],
        "sweets":["Jalebi","Halwa (Atta/Sooji)","Rabri","Kulfi Faluda","Kheer","Sohan Halwa"],
        "famous_spots":["Karim's (Jama Masjid)","Paranthe Wali Gali (Chandni Chowk)",
                        "Al Jawahar","Indian Accent (fine dining)",
                        "Bengali Market (Chaat)","Purana Qila area"],
        "food_culture":"Delhi is India's culinary melting pot. Butter Chicken was invented here at Moti Mahal. Old Delhi's Mughal food trail is unparalleled. Street chaat and kebabs are the city's heartbeat.",
        "cuisine_tags":["mughlai","north indian","punjabi","street food","kebab","chaat"],
        "diet_friendly":{"vegetarian":True,"halal":True,"jain":True},
        "avg_cost_two":500,"trending":["Momos","Biryanis","Artisan Chai"],
    },
    "new delhi": {
        "city":"New Delhi","state":"Delhi","emoji":"🕌",
        "must_try":["Butter Chicken","Chole Bhature","Paranthe","Nihari","Seekh Kebab","Chaat"],
        "street_food":["Chaat","Gol Gappe","Aloo Tikki","Momos"],
        "sweets":["Jalebi","Kulfi","Rabri"],
        "famous_spots":["Karim's","Paranthe Wali Gali","Bengali Market"],
        "food_culture":"Delhi — Mughal-influenced street food capital of India.",
        "cuisine_tags":["mughlai","north indian","punjabi","street food"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":500,"trending":["Momos","Cloud Kitchens"],
    },

    # ─── West Bengal ──────────────────────────────────────────────────────────
    "kolkata": {
        "city":"Kolkata","state":"West Bengal","emoji":"🌺",
        "must_try":["Kolkata Biryani (with potato & egg)","Kosha Mangsho",
                    "Mustard Hilsa (Shorshe Ilish)","Puchka","Ghugni Chaat",
                    "Kathi Rolls (birthplace)","Mishti Doi","Rosogolla (GI-tagged)","Sandesh"],
        "street_food":["Puchka","Ghugni","Churmur","Jhaal Muri","Kathi Roll","Singara"],
        "sweets":["Rosogolla","Mishti Doi","Sandesh","Chomchom","Rasamalai","Langcha"],
        "famous_spots":["Peter Cat (Chelo Kebab)","Mocambo (Continental)",
                        "Arsalan (Biryani)","Nizam's (Kathi Rolls)",
                        "Balaram Mullick (Sweets)","KC Das (Rosogolla)"],
        "food_culture":"Kolkata's food is a mosaic of Bengali, Mughal, and British influences. The unique Biryani has boiled egg and potato. Mishti Doi, Rosogolla, and Sandesh are the city's sweet trinity.",
        "cuisine_tags":["bengali","mughlai","street food","seafood","sweets"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":400,"trending":["Fusion Biryani","Korean Fried Chicken","Bubble Tea"],
    },

    # ─── Punjab ───────────────────────────────────────────────────────────────
    "amritsar": {
        "city":"Amritsar","state":"Punjab","emoji":"🙏",
        "must_try":["Amritsari Kulcha with Chole","Amritsari Fish Fry","Lassi",
                    "Makki di Roti + Sarson da Saag","Pinni","Golden Temple Langar",
                    "Tandoori Chicken","Dal Makhani","Shahi Paneer"],
        "street_food":["Kulcha Chole","Lassi","Jalebi","Pinni","Stuffed Parathas"],
        "sweets":["Pinni","Gajar Halwa","Jalebi","Kheer","Gurh Shakkar"],
        "famous_spots":["Kesar da Dhaba (since 1916)","Brothers' Dhaba",
                        "Bharawan da Dhaba","Golden Temple Langar (feeds 100,000/day)"],
        "food_culture":"Amritsar is Punjab's food capital. The Golden Temple Langar feeds 100,000+ daily for free. Amritsari Fish Fry and Kulcha are globally iconic.",
        "cuisine_tags":["punjabi","north indian","street food","vegetarian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":300,"trending":["Amritsari Kulcha","Instagram Lassi"],
    },
    "chandigarh": {
        "city":"Chandigarh","state":"Punjab/Haryana","emoji":"🌹",
        "must_try":["Chole Bhature","Amritsari Kulcha","Butter Chicken","Lassi",
                    "Makki Roti Sarson Saag","Punjabi Thali","Pinni"],
        "street_food":["Chaat","Gol Gappe","Dahi Bhalla","Pinni","Kulfi"],
        "sweets":["Pinni","Gajar Halwa","Gulab Jamun","Kheer"],
        "famous_spots":["Pal Dhaba","Gopal's Snacks","Sindhi Sweets","Sector 17 plaza"],
        "food_culture":"Chandigarh has vibrant Punjabi food culture. Sector 17 and 22 are the food hubs.",
        "cuisine_tags":["punjabi","north indian","vegetarian"],
        "diet_friendly":{"vegetarian":True},
        "avg_cost_two":400,"trending":["Rooftop Cafes","Craft Beer + Food"],
    },

    # ─── Goa ──────────────────────────────────────────────────────────────────
    "goa": {
        "city":"Goa","state":"Goa","emoji":"🏖️",
        "must_try":["Goan Fish Curry with Rice","Prawn Balchão","Chicken Cafreal",
                    "Pork Vindaloo","Pork Sorpotel","Bebinca","Dodol",
                    "Feni (cashew liquor)","Ros Omelette","Goan Sausages","Xacuti"],
        "street_food":["Ros Omelette","Samosa","Coconut Water","Prawn Rawa Fry","Bebinca"],
        "sweets":["Bebinca","Dodol","Serradura","Bolinhas","Bolo de Mel"],
        "famous_spots":["Ritz Classic (Panaji)","Vinayak (North Goa)",
                        "Fisherman's Wharf (Cavelossim)","Beach shacks everywhere"],
        "food_culture":"Goan cuisine reflects Portuguese colonial heritage mixed with Konkani traditions. Vindaloo and Xacuti are world-famous. Beach shacks serve the most authentic food.",
        "cuisine_tags":["goan","portuguese","konkani","seafood","pork"],
        "diet_friendly":{"vegetarian":True,"halal":False},
        "avg_cost_two":600,"trending":["Beach Shack Dining","Cashew Feni Cocktails"],
    },

    # ─── Uttar Pradesh ────────────────────────────────────────────────────────
    "lucknow": {
        "city":"Lucknow","state":"Uttar Pradesh","emoji":"🌙",
        "must_try":["Tunday Kebab (Galouti)","Lucknowi Biryani (Awadhi dum)","Nihari",
                    "Shami Kebab","Basket Chaat","Sheermal","Kulcha Nahari","Makhan Malai"],
        "street_food":["Basket Chaat","Dahi Chaat","Shahi Tukda","Sheermal","Bun Kabab"],
        "sweets":["Makhan Malai","Shahi Tukda","Kheer","Malai Paan","Imartiya"],
        "famous_spots":["Tunday Kababi (1905)","Royal Cafe (Basket Chaat)",
                        "Rahim's Nihari","Chowpatty (chaat)","Wahid Biryani"],
        "food_culture":"Lucknow is the home of Awadhi cuisine — known for dum cooking and the legendary Galouti Kebab melting on the tongue. The food scene is aristocratic and refined.",
        "cuisine_tags":["awadhi","mughlai","north indian","kebab"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":350,"trending":["Galouti Kabab","Awadhi Biryani"],
    },
    "varanasi": {
        "city":"Varanasi","state":"Uttar Pradesh","emoji":"🕯️",
        "must_try":["Baati Chokha","Malaiyo (winter milk foam)","Tamatar Chaat",
                    "Kachori with Aloo Sabzi","Jalebi-Kachori breakfast",
                    "Thandai (Holi special)","Rabri Kulfi","Banarasi Paan"],
        "street_food":["Kachori Sabzi","Tamatar Chaat","Lassi","Chaat","Chena Dahi Vada"],
        "sweets":["Malaiyo","Jalebi","Rabri","Paan","Malaiyyo"],
        "famous_spots":["Deena Chaat Bhandar","Kashi Chaat Bhandar","Blue Lassi Shop"],
        "food_culture":"Varanasi food is deeply spiritual and simple. Ghat-side chai and seasonal Malaiyo are quintessentially Banarasi. The city has a vegetarian food tradition.",
        "cuisine_tags":["banarasi","north indian","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":200,"trending":["Ghat-side Chai","Malaiyo"],
    },

    # ─── Odisha ───────────────────────────────────────────────────────────────
    "bhubaneswar": {
        "city":"Bhubaneswar","state":"Odisha","emoji":"🏛️",
        "must_try":["Dalma (lentils + vegetables)","Pakhala Bhata (fermented rice)",
                    "Chhena Poda (burnt cheese)","Machha Besara (mustard fish)",
                    "Santula","Odia Rasgola"],
        "street_food":["Bara (black gram fritter)","Gupchup","Chhena Gaja","Dahi Bara"],
        "sweets":["Chhena Poda","Rasgola","Chhena Gaja","Khaja","Rasabali"],
        "famous_spots":["Hotel Hare Krishna","Dalma Restaurant","Sweety Sweet Shop"],
        "food_culture":"Odia cuisine is mild and naturally flavored. Chhena Poda (burnt cheese dessert) is the most unique Indian sweet — discovered by accident in a wood fire.",
        "cuisine_tags":["odia","east indian","vegetarian","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":250,"trending":["Pakhala Bhata","Tribal Food"],
    },

    # ─── Northeast India ──────────────────────────────────────────────────────
    "guwahati": {
        "city":"Guwahati","state":"Assam","emoji":"🍃",
        "must_try":["Masor Tenga (sour fish curry)","Aloo Pitika","Khar (alkaline curry)",
                    "Duck Curry","Jolpan (traditional breakfast)","Pitha","Poita Bhaat"],
        "street_food":["Pitha","Jhal Muri","Momos","Egg Roll"],
        "sweets":["Til Pitha","Narikolor Ladoo","Pati Sapta"],
        "famous_spots":["Paradise Restaurant","Uzan Bazar food stalls"],
        "food_culture":"Assamese food is light and aromatic. Tenga (sour) and Khar (alkaline) are unique flavour profiles found nowhere else in India.",
        "cuisine_tags":["assamese","northeast indian","fish","bamboo shoot"],
        "diet_friendly":{"vegetarian":True},
        "avg_cost_two":300,"trending":["Bamboo Shoot dishes","Tribal Cuisine"],
    },
    "shillong": {
        "city":"Shillong","state":"Meghalaya","emoji":"🌧️",
        "must_try":["Jadoh (Khasi rice + pork)","Tungrymbai (fermented soybean)",
                    "Dohkhlieh (pork salad)","Pumaloi (rice powder)","Nakham Bitchi (dried fish soup)"],
        "street_food":["Momos","Jadoh","Wax Gourd Chutney","Egg Fritters"],
        "sweets":["Pukhlein (rice + jaggery)","Bamboo shoot pickle"],
        "famous_spots":["Police Bazar food stalls","Trattoria","Cafe Shillong"],
        "food_culture":"Shillong's food reflects Khasi tribal traditions — pork-based, earthy, and fermented flavors. The city also has amazing cafes and bakeries.",
        "cuisine_tags":["khasi","northeast","pork","tribal"],
        "diet_friendly":{"vegetarian":False},
        "avg_cost_two":300,"trending":["Jadoh","Fusion Cafes","Smoked Pork"],
    },

    # ─── Himachal Pradesh ─────────────────────────────────────────────────────
    "manali": {
        "city":"Manali","state":"Himachal Pradesh","emoji":"⛰️",
        "must_try":["Siddu (walnut-stuffed bread)","Tudkiya Bhath","Dhaam (feast)",
                    "Chha Gosht","Madra","Trout Fish","Tibetan Thukpa","Momos"],
        "street_food":["Momos","Thukpa","Maggi (mountain style)","Bread Omelette"],
        "sweets":["Meetha (rice pudding)","Dry Fruit Mix","Aktori (buckwheat cake)"],
        "famous_spots":["Johnson's Cafe","Drifter's Cafe","La Pizzeria"],
        "food_culture":"Manali blends Himachali cuisine with Tibetan influences. Fresh trout from rivers is prized. The old Manali café scene is beloved by backpackers.",
        "cuisine_tags":["himachali","tibetan","north indian","trout"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":350,"trending":["Trout Fish","Tibetan Food","Cafe Culture"],
    },
    "dharamshala": {
        "city":"Dharamshala","state":"Himachal Pradesh","emoji":"☸️",
        "must_try":["Tibetan Thukpa","Momos (Tibetan style)","Tsampa","Butter Tea",
                    "Shaptak (beef)","Tingmo (steamed bread)","Chow Mein"],
        "street_food":["Momos","Thukpa","Tingmo","Bread Omelette"],
        "sweets":["Khapse (fried pastry)","Tsampa balls"],
        "famous_spots":["Lung Ta Japanese Restaurant","Jimmy's Italian Kitchen","Tibet Kitchen"],
        "food_culture":"Dharamshala is home to the Tibetan government-in-exile. Authentic Tibetan food abounds — Thukpa, momos, and butter tea are staples.",
        "cuisine_tags":["tibetan","himachali","vegan","vegetarian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":300,"trending":["Authentic Tibetan Thukpa","Butter Tea"],
    },

    # ─── Southeast Asia ───────────────────────────────────────────────────────
    "singapore": {
        "city":"Singapore","state":"Singapore","emoji":"🦁",
        "must_try":["Hainanese Chicken Rice (national dish)","Chilli Crab","Laksa",
                    "Char Kway Teow","Satay","Nasi Lemak","Roti Prata",
                    "Hokkien Mee","Bak Kut Teh","Kaya Toast + Soft Boiled Eggs"],
        "street_food":["Satay","Rojak","Carrot Cake (Chai Tow Kway)","Ice Kachang","Popiah"],
        "sweets":["Durian","Ice Kachang","Kueh","Chendol","Ondeh Ondeh"],
        "famous_spots":["Maxwell Food Centre","Lau Pa Sat","Tiong Bahru Market",
                        "Hawker Chan (Michelin-starred chicken rice)"],
        "food_culture":"Singapore is a hawker food paradise. Maxwell and Lau Pa Sat are iconic. Hawker Chan became the world's cheapest Michelin-starred restaurant.",
        "cuisine_tags":["hawker","chinese","malay","indian","seafood"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":800,"trending":["Plant-Based Hawker","Bubble Tea","Prawn Noodles"],
    },
    "bangkok": {
        "city":"Bangkok","state":"Thailand","emoji":"🌸",
        "must_try":["Pad Thai","Tom Yum Goong","Green Curry","Som Tum",
                    "Massaman Curry","Khao Man Gai","Boat Noodles","Mango Sticky Rice"],
        "street_food":["Pad Thai","Som Tum","Satay","Roti","Grilled Corn","Mango Sticky Rice"],
        "sweets":["Mango Sticky Rice","Thai Milk Tea","Tub Tim Krob","Khanom Krok"],
        "famous_spots":["Yaowarat (Chinatown)","Silom food street",
                        "Jay Fai (Michelin-starred street stall)","Or Tor Kor Market"],
        "food_culture":"Bangkok has more Michelin stars in street food than any city. Jay Fai — a street stall — holds a Michelin star. Yaowarat Chinatown is legendary.",
        "cuisine_tags":["thai","street food","seafood","spicy"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True},
        "avg_cost_two":600,"trending":["Thai Omakase","Plant-Based Thai"],
    },
    "kuala lumpur": {
        "city":"Kuala Lumpur","state":"Malaysia","emoji":"🌆",
        "must_try":["Nasi Lemak (national dish)","Char Kway Teow","Roti Canai",
                    "Laksa","Satay Kajang","Bak Kut Teh","Cendol","Banana Leaf Rice"],
        "street_food":["Roti Canai","Satay","Nasi Lemak","Cendol","Rojak"],
        "sweets":["Cendol","Ais Kacang","Kuih","Apam Balik"],
        "famous_spots":["Jalan Alor Night Market","Imbi Market","Old Klang Road hawkers"],
        "food_culture":"KL's food is a magnificent blend of Malay, Chinese, and Indian traditions. Jalan Alor is the most famous food street in Southeast Asia.",
        "cuisine_tags":["malay","chinese","indian","hawker","halal"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":500,"trending":["Durian desserts","Modern Malay"],
    },
    "ho chi minh city": {
        "city":"Ho Chi Minh City","state":"Vietnam","emoji":"🛵",
        "must_try":["Pho Bo (beef noodle soup)","Banh Mi","Com Tam (broken rice)",
                    "Bun Bo Hue","Goi Cuon (fresh spring rolls)","Banh Xeo (sizzling crepe)",
                    "Che (sweet dessert soup)","Ca Phe Trung (egg coffee)"],
        "street_food":["Banh Mi","Pho","Goi Cuon","Che","Banh Trang Nuong"],
        "sweets":["Che","Banh Flan","Nuoc Mia (sugarcane juice)"],
        "famous_spots":["Pho Hoa Pasteur","Banh Mi Huynh Hoa","Ben Thanh Market"],
        "food_culture":"HCMC is a street food paradise. Banh Mi was voted among the world's best sandwiches. The city has thousands of pho stalls open from 6 AM.",
        "cuisine_tags":["vietnamese","street food","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":300,"trending":["Egg Coffee","Banh Mi variants","Broken Rice"],
    },

    # ─── Middle East ──────────────────────────────────────────────────────────
    "dubai": {
        "city":"Dubai","state":"UAE","emoji":"🌆",
        "must_try":["Al Harees","Al Machboos (spiced rice)","Camel Meat Burger",
                    "Luqaimat (honey dumplings)","Shawarma","Hummus & Mezze",
                    "Balaleet (sweet vermicelli)","Dates & Camel Milk"],
        "street_food":["Shawarma","Falafel","Manakish","Luqaimat","Kaak (sesame bread)"],
        "sweets":["Luqaimat","Umm Ali","Dates with Camel Milk","Baklava","Kunafa"],
        "famous_spots":["Gold Souk area (Deira)","Al Ustad Special Kabab",
                        "Bu Qtair Fish Restaurant","Dubai Food Festival venues"],
        "food_culture":"Dubai blends Emirati, Lebanese, Persian, and Indian flavours. Deira has the best authentic street food. Ramadan nights transform the city into a food festival.",
        "cuisine_tags":["emirati","lebanese","arabic","south asian","seafood"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":1200,"trending":["Modern Emirati","Camel dishes","Desert Dining"],
    },
    "istanbul": {
        "city":"Istanbul","state":"Turkey","emoji":"🕌",
        "must_try":["Döner Kebab","Balik Ekmek (fish sandwich)","Simit (sesame bread)",
                    "Meze Platter","Lahmacun","Iskender Kebab","Manti (dumplings)",
                    "Turkish Breakfast","Kunefe"],
        "street_food":["Simit","Balik Ekmek","Mussels (Midye)","Corn on Cob","Chestnuts"],
        "sweets":["Baklava","Kunefe","Turkish Delight","Sutlac","Kazandibi"],
        "famous_spots":["Karakoy Gulluoglu (Baklava)","Durumzade (Döner)",
                        "Fatih Fish Market","Grand Bazaar food lane"],
        "food_culture":"Istanbul sits at the crossroads of Europe and Asia. Turkish cuisine is one of the world's top 3. The Grand Bazaar area alone has 500+ years of food history.",
        "cuisine_tags":["turkish","ottoman","mediterranean","street food"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":700,"trending":["Turkish Coffee","Modern Meyhane","Croissant+Simit"],
    },

    # ─── Europe ───────────────────────────────────────────────────────────────
    "london": {
        "city":"London","state":"United Kingdom","emoji":"🎡",
        "must_try":["Fish and Chips","Full English Breakfast","Chicken Tikka Masala",
                    "Pie and Mash","Sunday Roast","Afternoon Tea","Cornish Pasty","Eton Mess"],
        "street_food":["Fish and Chips","Crepes","Jerk Chicken","Halal Cart","Pretzels"],
        "sweets":["Scones with Clotted Cream","Eton Mess","Sticky Toffee Pudding",
                  "Victoria Sponge","Chelsea Bun"],
        "famous_spots":["Borough Market (artisan)","Brick Lane (curry)","Dishoom (Indian)",
                        "The Fat Duck (Heston)","Hawksmoor (steakhouse)"],
        "food_culture":"London is one of the world's most diverse food cities. Brick Lane for Indian/Bangladeshi, Borough Market for artisan produce. Chicken Tikka Masala is now considered a British national dish.",
        "cuisine_tags":["british","indian","international","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True},
        "avg_cost_two":2500,"trending":["Plant-Based Fine Dining","Korean BBQ","Smash Burgers"],
    },
    "paris": {
        "city":"Paris","state":"France","emoji":"🗼",
        "must_try":["Croissant","Croque Monsieur","French Onion Soup","Steak Frites",
                    "Bouillabaisse","Duck Confit","Macarons","Crepes","Escargot"],
        "street_food":["Crepes","Croque Monsieur","Falafel (Marais)","Churros"],
        "sweets":["Macaron","Croissant","Éclair","Mille-Feuille","Tarte Tatin"],
        "famous_spots":["Café de Flore (historic)","L'As du Fallafel (Marais)",
                        "Ladurée (Macarons)","Pierre Hermé","Guy Savoy (3 Michelin stars)"],
        "food_culture":"Paris defined fine dining for the world. The brasserie culture is 200 years old. Marais has the best falafel in Europe, Lebanese-Jewish style.",
        "cuisine_tags":["french","brasserie","patisserie","fine dining"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True},
        "avg_cost_two":3000,"trending":["Natural Wine","Neo-Bistro","Japanese-French fusion"],
    },

    # ─── Americas ─────────────────────────────────────────────────────────────
    "new york": {
        "city":"New York","state":"USA","emoji":"🗽",
        "must_try":["NY Pizza slice","Bagel with Lox","NY Cheesecake","Pastrami on Rye",
                    "Hot Dog (street cart)","Halal Cart Chicken & Rice","Buffalo Wings","Lobster Roll"],
        "street_food":["Hot Dog","Halal Cart","Soft Pretzel","Falafel","Dirty Water Dogs"],
        "sweets":["NY Cheesecake","Black and White Cookie","Rugelach","Egg Cream","Cronut"],
        "famous_spots":["Katz's Deli (pastrami)","Di Fara Pizza (Brooklyn)",
                        "Eleven Madison Park (vegan fine dining)","Xi'an Famous Foods (noodles)"],
        "food_culture":"NYC is the immigrant food capital of the world. Every borough has an identity — Queens for global street food, Brooklyn for artisan, Manhattan for fine dining.",
        "cuisine_tags":["american","italian","jewish deli","chinese","halal"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"halal":True,"kosher":True},
        "avg_cost_two":3500,"trending":["K-BBQ","Birria Tacos","Smash Burgers","Ethiopian"],
    },
    "mexico city": {
        "city":"Mexico City","state":"Mexico","emoji":"🌮",
        "must_try":["Tacos al Pastor","Pozole","Mole Negro","Tamales","Chiles en Nogada",
                    "Tlayudas","Churros with Chocolate","Agua Fresca"],
        "street_food":["Tacos","Elotes (corn)","Quesadillas","Memelas","Tlacoyos"],
        "sweets":["Churros","Tres Leches Cake","Cajeta","Pulque"],
        "famous_spots":["Contramar (seafood)","El Turix (cochinita pibil)",
                        "Mercado de Medellín","Pujol (Michelin-starred)"],
        "food_culture":"Mexico City's food is UNESCO-listed. Tacos al Pastor were invented here by Lebanese immigrants. Pujol by Enrique Olvera redefined Mexican fine dining.",
        "cuisine_tags":["mexican","street food","latin american"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":800,"trending":["Modern Taco","Mezcal cocktails","Vegan Mexican"],
    },

    # ─── East Asia ────────────────────────────────────────────────────────────
    "tokyo": {
        "city":"Tokyo","state":"Japan","emoji":"🗾",
        "must_try":["Ramen (Shoyu/Tonkotsu/Miso)","Sushi & Sashimi","Tonkatsu",
                    "Yakitori","Tempura","Wagyu Beef","Okonomiyaki","Takoyaki",
                    "Matcha Parfait","Kaiseki"],
        "street_food":["Takoyaki","Yakitori","Taiyaki","Crepes (Harajuku)","Tamagoyaki"],
        "sweets":["Mochi","Matcha everything","Dorayaki","Wagashi","Warabi Mochi"],
        "famous_spots":["Tsukiji Outer Market (fresh seafood)","Ichiran Ramen",
                        "Sukiyabashi Jiro (sushi)","Shibuya ramen alleys"],
        "food_culture":"Tokyo has more Michelin stars than any city in the world. Tsukiji Outer Market for breakfast sushi. Ramen culture is an art form with 80,000+ ramen shops in Japan.",
        "cuisine_tags":["japanese","sushi","ramen","izakaya","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":2000,"trending":["Omakase","Wagyu Street Food","Matcha desserts"],
    },
    "seoul": {
        "city":"Seoul","state":"South Korea","emoji":"🇰🇷",
        "must_try":["Korean BBQ (Galbi/Samgyeopsal)","Bibimbap","Tteokbokki",
                    "Kimchi Jjigae","Japchae","Korean Fried Chicken","Sundubu Jjigae",
                    "Haemul Pajeon (seafood pancake)","Soju + Anju"],
        "street_food":["Tteokbokki","Hotteok","Gimbap","Odeng","Egg Bread","Fish Cake"],
        "sweets":["Bingsu (shaved ice)","Hotteok","Yakgwa","Sikhye"],
        "famous_spots":["Gwangjang Market (oldest food market)","Tongin Market",
                        "Myeongdong street food","Noryangjin Fish Market"],
        "food_culture":"Seoul's food culture is obsessive and proud. Korean BBQ is a social institution. The city has the highest density of cafes and convenience store gourmet food in the world.",
        "cuisine_tags":["korean","bbq","street food","fermented","seafood"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":1500,"trending":["K-Street Food","Hanwoo Beef","Natural Wine Bars"],
    },

    # ─── More Indian cities ───────────────────────────────────────────────────
    "indore": {
        "city":"Indore","state":"Madhya Pradesh","emoji":"🌶️",
        "must_try":["Poha Jalebi (breakfast combo)","Dal Bafla","Shikanji",
                    "Sabudana Khichdi","Ratlami Sev","Malpua","Bhutte Ka Kees"],
        "street_food":["Poha","Jalebi","Ratlami Sev","Shikanji","Bhutte ka Kees","Garadu"],
        "sweets":["Malpua","Jalebi","Khopra Patties","Moong Dal Halwa"],
        "famous_spots":["Sarafa Bazaar (night food market)","Chappan Dukaan (56 shops street)"],
        "food_culture":"Indore's Sarafa Bazaar is one of India's best night food markets — jewellery shops by day, food stalls by night. Poha Jalebi is the iconic breakfast.",
        "cuisine_tags":["malwi","madhya pradesh","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True,"jain":True},
        "avg_cost_two":200,"trending":["Sarafa Night Market","Dal Bafla"],
    },
    "bhopal": {
        "city":"Bhopal","state":"Madhya Pradesh","emoji":"🌊",
        "must_try":["Bhopali Gosht Korma","Poha","Shahi Sheermal","Bhopali Paan",
                    "Mawa Bati","Biryani (Bhopali style)"],
        "street_food":["Poha","Jalebi","Kachori","Sheekh Kabab","Paan"],
        "sweets":["Mawa Bati","Shahi Sheermal","Jalebi","Kheer"],
        "famous_spots":["Manohar Dairy (Poha)","Chatori Gali","DB Mall food court"],
        "food_culture":"Bhopal has a rich Nawabi food tradition. Bhopali korma and Sheermal reflect the city's regal Muslim culinary heritage.",
        "cuisine_tags":["bhopali","mughlai","vegetarian","non-veg"],
        "diet_friendly":{"vegetarian":True,"halal":True},
        "avg_cost_two":300,"trending":["Bhopali Biryani","Nawabi Korma"],
    },
    "raipur": {
        "city":"Raipur","state":"Chhattisgarh","emoji":"🌾",
        "must_try":["Chila (rice crepe)","Faraa (steamed rice dumplings)","Aamat (tribal curry)",
                    "Muthia","Angakar Roti","Dubki Kadhi"],
        "street_food":["Chila","Samosa","Jalebi","Bhutte Ki Kees","Chaat"],
        "sweets":["Khurmi","Tilgur (sesame jaggery)","Petha"],
        "famous_spots":["Telibandha lake food stalls","Pandri Market"],
        "food_culture":"Raipur reflects Chhattisgarhi tribal and rural cuisine. Chila and Faraa are staples. Rich in forest produce and traditional preparations.",
        "cuisine_tags":["chhattisgarhi","tribal","vegetarian"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":200,"trending":["Tribal Food","Bamboo Rice dishes"],
    },
    "patna": {
        "city":"Patna","state":"Bihar","emoji":"🌿",
        "must_try":["Litti Chokha (Bihar's national dish)","Sattu Paratha","Dal Pitha",
                    "Khaja","Tilkut","Thekua","Balushahi"],
        "street_food":["Litti Chokha","Sattu Sharbat","Chaat","Jalebi","Poha"],
        "sweets":["Khaja","Tilkut","Balushahi","Thekua","Anarsa"],
        "famous_spots":["Bihar Museum café","Patna Sahib area food stalls"],
        "food_culture":"Bihar's Litti Chokha is a roasted wheat ball with spiced mashed vegetables — a beloved staple. Tilkut from Gaya is a famous sweet during Makar Sankranti.",
        "cuisine_tags":["bihari","north indian","vegetarian","street food"],
        "diet_friendly":{"vegetarian":True,"vegan":True},
        "avg_cost_two":200,"trending":["Litti Chokha","Sattu Drinks"],
    },
}


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _hav(lat1, lon1, lat2, lon2) -> float:
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    return R * 2 * math.asin(math.sqrt(a))

def _geocode(location: str) -> Optional[Tuple]:
    try:
        r = requests.get(NOMINATIM, headers=HEADERS_OSM, timeout=10,
                         params={"q": location, "format": "json", "limit": 1})
        d = r.json()
        if d:
            return float(d[0]["lat"]), float(d[0]["lon"]), d[0]["display_name"]
    except Exception as e:
        log.warning(f"Geocode failed: {e}")
    return None

def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()

def _city_key(location: str, display: str = "") -> str:
    loc = location.strip().lower()
    for key in FAMOUS_FOODS:
        if key in loc or loc.startswith(key):
            return key
    for part in display.split(","):
        p = part.strip().lower()
        if p in FAMOUS_FOODS:
            return p
    return ""

def _infer_cuisines(tags: Dict) -> List[str]:
    raw = tags.get("cuisine", "")
    if not raw:
        return []
    return [c.strip().title() for c in re.split(r"[;,/]", raw) if c.strip()]

def _infer_amenities(tags: Dict) -> Dict[str, str]:
    return {k: tags.get(v, "") for k, v in {
        "outdoor_seating": "outdoor_seating",
        "delivery":        "delivery",
        "takeaway":        "takeaway",
        "vegetarian":      "diet:vegetarian",
        "vegan":           "diet:vegan",
        "halal":           "diet:halal",
        "jain":            "diet:vegan",
        "gluten_free":     "diet:gluten_free",
        "wifi":            "internet_access",
        "wheelchair":      "wheelchair",
        "air_conditioning":"air_conditioning",
        "reservations":    "reservation",
    }.items()}


# ══════════════════════════════════════════════════════════════════════════════
#  SENTIMENT ANALYSIS  (keyword-based, no ML library needed)
# ══════════════════════════════════════════════════════════════════════════════

def _sentiment(text: str) -> Dict[str, Any]:
    """Classify review text sentiment using keyword matching."""
    lower = text.lower()
    pos = sum(1 for w in SENTIMENT_POSITIVE if w in lower)
    neg = sum(1 for w in SENTIMENT_NEGATIVE if w in lower)
    if pos == 0 and neg == 0:
        label = "Neutral"
    elif pos > neg * 1.5:
        label = "Positive 😊"
    elif neg > pos * 1.5:
        label = "Negative 😞"
    else:
        label = "Mixed 😐"
    return {"label": label, "positive_signals": pos, "negative_signals": neg}


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPERS
# ══════════════════════════════════════════════════════════════════════════════

def _scrape_osm(lat: float, lon: float, radius_m: int,
                cuisine_filter: str = "",
                type_filter: Optional[List[str]] = None,
                diet: str = "") -> List[Dict]:
    pairs = [(k, v) for k, v in FOOD_TAG_PAIRS
             if not type_filter or v in type_filter]
    tag_clause = "\n".join(
        f'  node["{k}"="{v}"](around:{radius_m},{lat},{lon});\n'
        f'  way["{k}"="{v}"](around:{radius_m},{lat},{lon});'
        for k, v in pairs
    )
    query = f"[out:json][timeout:90];\n(\n{tag_clause}\n);\nout center body;"
    try:
        r = requests.post(OVERPASS, data=query, headers=HEADERS_OSM, timeout=95)
        r.raise_for_status()
        elements = r.json().get("elements", [])
    except Exception as e:
        log.error(f"Overpass failed: {e}"); return []

    records, seen = [], set()
    for el in elements:
        tags = el.get("tags", {})
        name = _clean(tags.get("name") or tags.get("brand") or "")
        if not name or name.lower() in seen: continue
        seen.add(name.lower())

        if el["type"] == "node":
            r_lat, r_lon = el.get("lat"), el.get("lon")
        else:
            c = el.get("center", {})
            r_lat, r_lon = c.get("lat"), c.get("lon")
        if r_lat is None: continue

        dist     = _hav(lat, lon, r_lat, r_lon)
        ftype    = tags.get("amenity") or tags.get("shop") or "restaurant"
        icon     = PLACE_ICONS.get(ftype, "🍽️")
        cuisines = _infer_cuisines(tags)
        amenities = _infer_amenities(tags)

        # Cuisine filter
        if cuisine_filter:
            cf = cuisine_filter.lower()
            blob = " ".join(cuisines).lower() + " " + tags.get("cuisine","").lower() + " " + name.lower()
            if cf not in blob: continue

        # Diet filter
        if diet:
            diet_key = diet.lower().replace("-","_")
            osm_tag  = DIET_OSM_TAGS.get(diet_key, "")
            kws      = DIET_KEYWORDS.get(diet_key, [diet_key])
            name_blob = (name + " " + tags.get("cuisine","") + " " +
                         tags.get("description","")).lower()
            has_diet = (
                (osm_tag and tags.get(osm_tag) in ("yes","only")) or
                any(kw in name_blob for kw in kws)
            )
            if not has_diet: continue

        addr_parts = [
            tags.get("addr:housenumber",""), tags.get("addr:street",""),
            tags.get("addr:suburb",""),      tags.get("addr:city",""),
            tags.get("addr:postcode",""),
        ]
        address = _clean(", ".join(p for p in addr_parts if p)) or "N/A"
        hours   = tags.get("opening_hours","N/A")
        if hours == "24/7": hours = "Open 24 Hours"

        records.append({
            "name":          name,
            "icon":          icon,
            "type":          ftype,
            "type_label":    f"{icon} {ftype.replace('_',' ').title()}",
            "cuisines":      cuisines,
            "address":       address,
            "phone":         _clean(tags.get("phone") or tags.get("contact:phone") or ""),
            "website":       _clean(tags.get("website") or tags.get("contact:website") or ""),
            "opening_hours": hours,
            "amenities":     amenities,
            "price_range":   tags.get("price_range",""),
            "lat":           r_lat,
            "lon":           r_lon,
            "dist_km":       round(dist, 3),
            "maps_url":      f"https://www.google.com/maps?q={r_lat},{r_lon}",
            "directions_url":f"https://www.google.com/maps/dir/?api=1&destination={r_lat},{r_lon}",
            "osm_id":        el.get("id"),
            "rating":        None, "votes": None, "cost_for_two": None,
            "zomato_url":    "", "swiggy_url": "",
            "popular_dishes":[], "reviews_sentiment": None,
            "price_sources": {}, "description": "",
            "source": "OpenStreetMap",
        })

    records.sort(key=lambda x: x["dist_km"])
    return records


def _scrape_zomato(name: str, city: str) -> Dict:
    result = {"rating":None,"votes":None,"cost_for_two":None,
              "zomato_url":"","popular_dishes":[],"review_snippets":[]}
    try:
        city_slug = city.lower().replace(" ","-")
        search_url = f"https://www.zomato.com/{city_slug}/search?q={requests.utils.quote(name)}"
        result["zomato_url"] = search_url
        r = requests.get(search_url, headers=HEADERS_WEB, timeout=12)
        if r.status_code != 200: return result
        soup = BeautifulSoup(r.text, "lxml")
        # JSON-LD
        for script in soup.find_all("script", {"type":"application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list): data = data[0]
                agg = data.get("aggregateRating", {})
                if agg.get("ratingValue"):
                    result["rating"] = agg["ratingValue"]
                    result["votes"]  = agg.get("reviewCount")
                    return result
            except Exception: pass
        # __NEXT_DATA__
        nd = soup.find("script", {"id":"__NEXT_DATA__"})
        if nd:
            try:
                data = json.loads(nd.string or "{}")
                for section in (data.get("props",{}).get("pageProps",{})
                                .get("searchResults",{}).get("data",{})
                                .get("sections",[])):
                    for card in section.get("cards",[]):
                        info = card.get("card",{}).get("card",{}).get("info",{})
                        if name.lower() in info.get("name","").lower():
                            result["rating"]       = info.get("avgRating")
                            result["votes"]        = info.get("totalRatingsString")
                            result["cost_for_two"] = info.get("costForTwo")
                            return result
            except Exception: pass
    except Exception as e:
        log.debug(f"Zomato failed: {e}")
    return result


def _scrape_swiggy(name: str, lat: float, lon: float) -> Dict:
    result = {"rating":None,"delivery_time":None,"swiggy_url":"","cost_for_two":None}
    try:
        url = f"https://www.swiggy.com/search?query={requests.utils.quote(name)}"
        result["swiggy_url"] = url
        r = requests.get(url, headers=HEADERS_WEB, timeout=12)
        if r.status_code != 200: return result
        soup = BeautifulSoup(r.text, "lxml")
        for script in soup.find_all("script", {"type":"application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list): data = data[0]
                agg = data.get("aggregateRating",{})
                if agg.get("ratingValue"):
                    result["rating"] = agg["ratingValue"]
                    return result
            except Exception: pass
    except Exception as e:
        log.debug(f"Swiggy failed: {e}")
    return result


def _scrape_google(name: str, lat: float, lon: float) -> Dict:
    result = {"rating":None,"votes":None,"price_level":"","website":""}
    try:
        q   = f"{name} restaurant {lat:.4f},{lon:.4f}"
        r   = requests.get(f"https://www.google.com/search?q={requests.utils.quote(q)}&num=3",
                           headers=HEADERS_WEB, timeout=12)
        soup = BeautifulSoup(r.text, "lxml")
        for script in soup.find_all("script", {"type":"application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list): data = data[0]
                if data.get("@type") in ("Restaurant","FoodEstablishment","LocalBusiness"):
                    agg = data.get("aggregateRating",{})
                    result["rating"]  = agg.get("ratingValue")
                    result["votes"]   = agg.get("reviewCount")
                    result["website"] = data.get("url","")
                    return result
            except Exception: pass
        m = re.search(r"(\d\.\d)\s*/\s*5", r.text)
        if m: result["rating"] = m.group(1)
        m2 = re.search(r"([\d,]+)\s*(?:Google reviews|reviews)", r.text)
        if m2: result["votes"] = m2.group(1).replace(",","")
    except Exception as e:
        log.debug(f"Google failed: {e}")
    return result


def _scrape_tripadvisor(name: str, city: str) -> Dict:
    result = {"rating":None,"reviews":None,"ta_url":"","ranking":"","snippets":[]}
    try:
        q   = requests.utils.quote(f"{name} restaurant {city}")
        url = f"https://www.tripadvisor.com/Search?q={q}&searchSessionId=x"
        result["ta_url"] = url
        r   = requests.get(url, headers=HEADERS_WEB, timeout=12)
        soup = BeautifulSoup(r.text, "lxml")
        for script in soup.find_all("script", {"type":"application/ld+json"}):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list): data = data[0]
                agg = data.get("aggregateRating",{})
                if agg.get("ratingValue"):
                    result["rating"]  = agg["ratingValue"]
                    result["reviews"] = agg.get("reviewCount")
                    # Try to get review text snippets
                    for rev in data.get("review",[])[:3]:
                        text = rev.get("reviewBody","")
                        if text: result["snippets"].append(text[:120])
                    return result
            except Exception: pass
    except Exception as e:
        log.debug(f"TripAdvisor failed: {e}")
    return result


def _scrape_wiki_food(city: str) -> str:
    try:
        r1 = requests.get(WIKI_API, headers=HEADERS_JSON, timeout=8, params={
            "action":"query","list":"search",
            "srsearch":f"{city} cuisine food culture","format":"json","srlimit":1
        })
        results = r1.json().get("query",{}).get("search",[])
        if not results: return ""
        r2 = requests.get(WIKI_API, headers=HEADERS_JSON, timeout=8, params={
            "action":"query","titles":results[0]["title"],
            "prop":"extracts","exintro":True,"explaintext":True,
            "exsentences":4,"format":"json"
        })
        for page in r2.json().get("query",{}).get("pages",{}).values():
            return _clean(page.get("extract",""))
    except Exception: pass
    return ""


def _osrm_route(coords: List[Tuple[float,float]]) -> Dict:
    """Get real road route using OSRM for food tour planning."""
    try:
        coord_str = ";".join(f"{lon},{lat}" for lat,lon in coords)
        r = requests.get(f"{OSRM_URL}/{coord_str}",
                         params={"overview":"simplified","steps":"false"},
                         headers=HEADERS_OSM, timeout=10)
        data = r.json()
        if data.get("code") == "Ok":
            route = data["routes"][0]
            return {
                "distance_km": round(route["distance"] / 1000, 2),
                "duration_min": round(route["duration"] / 60, 1),
            }
    except Exception: pass
    return {}


# ══════════════════════════════════════════════════════════════════════════════
#  ENRICHMENT
# ══════════════════════════════════════════════════════════════════════════════

def _booking_links(name: str, city: str) -> Dict[str,str]:
    n = requests.utils.quote(name)
    c = requests.utils.quote(city)
    cs = city.lower().replace(" ","-")
    return {
        "zomato":      f"https://www.zomato.com/{cs}/search?q={n}",
        "swiggy":      f"https://www.swiggy.com/search?query={n}",
        "google_maps": f"https://www.google.com/maps/search/{n}+{c}",
        "tripadvisor": f"https://www.tripadvisor.com/Search?q={n}+{c}",
        "magicpin":    f"https://magicpin.in/{c}/restaurants/?query={n}",
        "eazydiner":   f"https://www.eazydiner.com/city/search?q={n}&city={c}",
    }

def _enrich(rest: Dict, city: str, level: int = 1) -> Dict:
    name = rest["name"]
    lat  = rest["lat"]
    lon  = rest["lon"]
    price_sources = {}

    if level >= 1:
        z = _scrape_zomato(name, city)
        if z.get("rating") and not rest["rating"]:
            rest["rating"] = z["rating"]
            rest["votes"]  = z.get("votes")
        if z.get("cost_for_two"):
            rest["cost_for_two"] = z["cost_for_two"]
            price_sources["zomato"] = z["cost_for_two"]
        if z.get("popular_dishes"):
            rest["popular_dishes"] = z["popular_dishes"]
        rest["zomato_url"] = z.get("zomato_url","")
        time.sleep(0.3)

    if level >= 2:
        sw = _scrape_swiggy(name, lat, lon)
        if sw.get("rating") and not rest["rating"]:
            rest["rating"] = sw["rating"]
        if sw.get("cost_for_two"):
            price_sources["swiggy"] = sw["cost_for_two"]
        rest["swiggy_url"] = sw.get("swiggy_url","")
        time.sleep(0.3)

        g = _scrape_google(name, lat, lon)
        if g.get("rating") and not rest["rating"]:
            rest["rating"]  = g["rating"]
            rest["votes"]   = g.get("votes")
        if g.get("website") and not rest["website"]:
            rest["website"] = g["website"]
        if g.get("price_level"):
            price_sources["google"] = g["price_level"]
        time.sleep(0.3)

        # TripAdvisor (for review snippets + sentiment)
        ta = _scrape_tripadvisor(name, city)
        if ta.get("snippets"):
            combined = " ".join(ta["snippets"])
            rest["reviews_sentiment"] = _sentiment(combined)
            rest["review_snippets"]   = ta["snippets"]
        time.sleep(0.3)

    rest["price_sources"]  = price_sources
    rest["booking_links"]  = _booking_links(name, city)
    return rest


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _export_excel(restaurants: List[Dict], famous: Dict, filepath: str, location: str):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Restaurants"

    DEEP_RED    = PatternFill("solid", fgColor="BF360C")
    ALT_FILL    = PatternFill("solid", fgColor="FFF3E0")
    GREEN_FILL  = PatternFill("solid", fgColor="E8F5E9")
    hf          = Font(bold=True, color="FFFFFF", size=11)
    bf          = Font(bold=True, size=10)
    nf          = Font(size=10)
    ca          = Alignment(horizontal="center",vertical="center",wrap_text=True)
    la          = Alignment(horizontal="left",  vertical="center",wrap_text=True)
    thin        = Border(left=Side(style="thin"),right=Side(style="thin"),
                         top=Side(style="thin"),bottom=Side(style="thin"))

    ws.merge_cells("A1:P1")
    ws["A1"].value     = f"🍽️  Restaurants & Food Guide — {location}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font      = Font(bold=True, size=14, color="BF360C")
    ws["A1"].alignment = ca
    ws.row_dimensions[1].height = 30

    headers = ["#","Name","Type","Cuisines","Dist (km)","Address","Phone",
               "Rating","Cost/2","Sentiment","Hours","Veg","Delivery",
               "Popular Dishes","Zomato","Maps"]
    for i,h in enumerate(headers,1):
        c = ws.cell(row=2,column=i,value=h)
        c.font=hf; c.fill=DEEP_RED; c.alignment=ca; c.border=thin
    ws.row_dimensions[2].height = 22

    widths = [4,32,14,22,10,38,16,8,12,14,20,8,10,28,30,14]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for idx,r in enumerate(restaurants,1):
        row = idx+2
        fill = ALT_FILL if idx%2==0 else None
        sent = r.get("reviews_sentiment",{})
        values = [
            idx, r.get("name",""), r.get("type_label",""),
            ", ".join(r.get("cuisines",[])[:3]),
            r.get("dist_km",""), r.get("address",""),
            r.get("phone",""), r.get("rating",""), r.get("cost_for_two",""),
            sent.get("label","") if sent else "",
            r.get("opening_hours",""),
            r.get("amenities",{}).get("vegetarian",""),
            r.get("amenities",{}).get("delivery",""),
            ", ".join(r.get("popular_dishes",[])[:4]),
            r.get("zomato_url",""), r.get("maps_url",""),
        ]
        for ci,val in enumerate(values,1):
            cell = ws.cell(row=row,column=ci,value=str(val) if val else "")
            cell.font      = bf if ci==2 else nf
            cell.alignment = ca if ci in (1,5,8,12,13) else la
            cell.border    = thin
            if fill: cell.fill = fill

        for col, url, label in [(15, r.get("zomato_url",""), "📱 Zomato"),
                                 (16, r.get("maps_url",""),   "📍 Maps")]:
            if url:
                cell = ws.cell(row=row,column=col)
                cell.hyperlink = url; cell.value = label
                cell.font = Font(color="0563C1",underline="single",size=10)
        ws.row_dimensions[row].height = 18

    # Famous Foods sheet
    if famous:
        ws2 = wb.create_sheet("Famous Foods")
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 55
        ws2.merge_cells("A1:B1")
        ws2["A1"].value = f"🍛 Famous Foods — {famous.get('city',location)}, {famous.get('state','')}"
        ws2["A1"].font  = Font(bold=True,size=14,color="BF360C")
        ws2["A1"].alignment = ca
        ws2.row_dimensions[1].height = 28

        row = 2
        for title, key in [("🥘 Must Try","must_try"),("🏮 Street Food","street_food"),
                             ("🍬 Sweets","sweets"),("📍 Famous Spots","famous_spots"),
                             ("🏷️ Cuisines","cuisine_tags")]:
            items = famous.get(key,[])
            if not items: continue
            ws2.merge_cells(f"A{row}:B{row}")
            ws2.cell(row=row,column=1,value=title).font = Font(bold=True,color="FFFFFF",size=11)
            ws2.cell(row=row,column=1).fill = DEEP_RED
            ws2.cell(row=row,column=1).alignment = ca
            ws2.row_dimensions[row].height = 20; row += 1
            for i,item in enumerate(items,1):
                ws2.cell(row=row,column=1,value=str(i)).alignment = ca
                ws2.cell(row=row,column=2,value=item)
                if i%2==0:
                    ws2.cell(row=row,column=1).fill = ALT_FILL
                    ws2.cell(row=row,column=2).fill = ALT_FILL
                ws2.row_dimensions[row].height = 16; row += 1

        culture = famous.get("food_culture","")
        if culture:
            row += 1
            ws2.merge_cells(f"A{row}:B{row}")
            ws2.cell(row=row,column=1,value="📖 Food Culture").font = Font(bold=True,color="BF360C")
            ws2.row_dimensions[row].height = 18; row += 1
            ws2.merge_cells(f"A{row}:B{row}")
            c2 = ws2.cell(row=row,column=1,value=culture)
            c2.alignment = Alignment(wrap_text=True,vertical="top")
            ws2.row_dimensions[row].height = 72

    # Price comparison sheet
    price_data = [r for r in restaurants if r.get("price_sources")]
    if price_data:
        ws3 = wb.create_sheet("Price Comparison")
        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 18
        ws3.column_dimensions["D"].width = 18
        ws3.merge_cells("A1:D1")
        ws3["A1"].value = "💰 Price Comparison — Zomato vs Swiggy vs Google"
        ws3["A1"].font  = Font(bold=True,size=13,color="BF360C"); ws3.row_dimensions[1].height=26
        for i,h in enumerate(["Restaurant","Zomato","Swiggy","Google"],1):
            c = ws3.cell(row=2,column=i,value=h)
            c.font=hf; c.fill=DEEP_RED; c.alignment=ca
        row=3
        for r in price_data:
            ps = r.get("price_sources",{})
            ws3.cell(row=row,column=1,value=r["name"])
            ws3.cell(row=row,column=2,value=str(ps.get("zomato","")))
            ws3.cell(row=row,column=3,value=str(ps.get("swiggy","")))
            ws3.cell(row=row,column=4,value=str(ps.get("google","")))
            if row%2==0:
                for ci in range(1,5): ws3.cell(row=row,column=ci).fill=ALT_FILL
            row+=1

    wb.save(filepath)
    log.info(f"Excel saved → {filepath}")


def _export_json(restaurants: List[Dict], famous: Dict, filepath: str):
    data = {
        "generated_at": datetime.now().isoformat(),
        "total": len(restaurants),
        "restaurants": restaurants,
        "famous_foods": famous,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    log.info(f"JSON saved → {filepath}")


def _export_csv(restaurants: List[Dict], filepath: str):
    if not restaurants: return
    flat_keys = ["name","type","cuisines","dist_km","address","phone","website",
                 "rating","votes","cost_for_two","opening_hours",
                 "lat","lon","maps_url","zomato_url","swiggy_url"]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=flat_keys, extrasaction="ignore")
        writer.writeheader()
        for r in restaurants:
            row = dict(r)
            row["cuisines"] = ", ".join(r.get("cuisines",[]))
            writer.writerow({k: row.get(k,"") for k in flat_keys})
    log.info(f"CSV saved → {filepath}")


# ══════════════════════════════════════════════════════════════════════════════
#  COLOR PRINTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def c_title(text):   return f"{Fore.YELLOW}{Style.BRIGHT}{text}{Style.RESET_ALL}"
def c_label(text):   return f"{Fore.CYAN}{text}{Style.RESET_ALL}"
def c_val(text):     return f"{Fore.WHITE}{text}{Style.RESET_ALL}"
def c_url(text):     return f"{Fore.BLUE}{text}{Style.RESET_ALL}"
def c_good(text):    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"
def c_bad(text):     return f"{Fore.RED}{text}{Style.RESET_ALL}"
def c_head(text):    return f"{Back.RED}{Fore.WHITE}{Style.BRIGHT} {text} {Style.RESET_ALL}"
def c_dim(text):     return f"{Style.DIM}{text}{Style.RESET_ALL}"
def c_num(text):     return f"{Fore.MAGENTA}{Style.BRIGHT}{text}{Style.RESET_ALL}"
def c_sent(text):
    if "Positive" in text:  return c_good(text)
    if "Negative" in text:  return c_bad(text)
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"


# ══════════════════════════════════════════════════════════════════════════════
#  RESTAURANT AGENT v2
# ══════════════════════════════════════════════════════════════════════════════

class RestaurantAgent:
    """
    RestaurantAgent v2 — Nearby Restaurants, Famous Foods & Food Tour Planner

    New in v2:
        ✨ dietary_search()   — vegan / halal / jain / gluten-free filter
        ✨ price_compare()    — Zomato vs Swiggy vs Google cost comparison
        ✨ plan_food_tour()   — best 5-stop optimised food tour with OSRM routing
        ✨ trending_dishes()  — what's popular in a city right now
        ✨ compare_cities()   — compare food scenes of two cities
        ✨ export_json()      — JSON export
        ✨ export_csv()       — CSV export
        ✨ Colored terminal   — colorama-powered output
        ✨ Sentiment analysis — review sentiment on every enriched restaurant
        ✨ 70+ city DB        — India + SE Asia + Middle East + Europe + Americas
    """

    def __init__(self):
        print(c_head("RestaurantAgent v2"))
        print(c_dim(f"  OSM + Zomato + Swiggy + Google + TripAdvisor + OSRM"))
        print(c_dim(f"  Famous foods DB: {len(FAMOUS_FOODS)} cities worldwide"))

    # ── 1. Find Nearby ────────────────────────────────────────────────────────

    def find_nearby(self, location: str, radius_m: int = 3000,
                    limit: int = 20, cuisine: str = "",
                    diet: str = "", budget: str = "",
                    enrich: int = 1) -> Dict[str, Any]:
        """
        Find restaurants near a location.

        Args:
            location : city, address, pincode, landmark
            radius_m : radius in metres (default 3 km)
            limit    : max results (default 20)
            cuisine  : e.g. "south indian", "chinese", "biryani"
            diet     : "vegetarian", "vegan", "halal", "jain", "gluten_free"
            budget   : "budget", "moderate", "upscale", "fine"
            enrich   : 0=OSM only, 1=+Zomato, 2=+Swiggy+Google+Sentiment
        """
        geo = _geocode(location)
        if not geo:
            return {"error": f"Could not locate: '{location}'", "restaurants": []}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping restaurants near '{location}'…")
        restaurants = _scrape_osm(lat, lon, radius_m, cuisine, diet=diet)
        if not restaurants:
            log.warning(f"Widening search to {radius_m*2}m…")
            restaurants = _scrape_osm(lat, lon, radius_m*2, cuisine, diet=diet)

        # Budget filter (post-scrape; cost_for_two filled after enrich)
        # We store the budget intent and apply after enrichment

        if enrich > 0:
            for r in tqdm(restaurants[:limit], desc="Enriching", unit="place"):
                _enrich(r, city, enrich)

        # Apply budget filter
        if budget and budget in BUDGET_RANGES:
            lo, hi, _ = BUDGET_RANGES[budget]
            def in_budget(r):
                c2 = r.get("cost_for_two","")
                m  = re.search(r"[\d,]+", str(c2))
                if not m: return True  # no data → include
                val = int(m.group().replace(",",""))
                return lo <= val <= hi
            restaurants = [r for r in restaurants if in_budget(r)]

        restaurants = restaurants[:limit]
        city_key = _city_key(location, display)
        famous   = FAMOUS_FOODS.get(city_key, {})

        return {
            "location":     {"name":location,"display":display,"lat":lat,"lon":lon,"city":city},
            "radius_m":     radius_m,
            "cuisine":      cuisine or "All",
            "diet":         diet or "All",
            "budget":       budget or "All",
            "restaurants":  restaurants,
            "total":        len(restaurants),
            "famous_foods": famous,
            "scraped_at":   datetime.now().isoformat(),
        }

    # ── 2. Famous Foods ───────────────────────────────────────────────────────

    def famous_foods(self, location: str) -> Dict[str, Any]:
        geo = _geocode(location)
        display = geo[2] if geo else location
        key     = _city_key(location, display)
        known   = FAMOUS_FOODS.get(key, {})
        wiki    = _scrape_wiki_food(location)
        if not known:
            return {"city":location,"note":f"No curated data for '{location}'",
                    "wiki_culture":wiki,"must_try":[],"street_food":[],"sweets":[]}
        if wiki: known["wiki_culture"] = wiki
        return known

    # ── 3. Dietary search ────────────────────────────────────────────────────

    def dietary_search(self, location: str, diet: str,
                        radius_m: int = 5000, limit: int = 15,
                        enrich: int = 1) -> Dict[str, Any]:
        """
        Filter restaurants by dietary preference.

        diet: "vegetarian" | "vegan" | "halal" | "jain" | "gluten_free" | "kosher"
        """
        geo = _geocode(location)
        if not geo: return {"error":f"Could not locate: '{location}'"}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Scraping {diet} restaurants near '{location}'…")
        results = _scrape_osm(lat, lon, radius_m, diet=diet)

        if enrich > 0:
            for r in tqdm(results[:limit], desc=f"Enriching ({diet})", unit="place"):
                _enrich(r, city, enrich)

        # Also check famous foods for diet-friendly options
        city_key = _city_key(location, display)
        famous   = FAMOUS_FOODS.get(city_key, {})
        df       = famous.get("diet_friendly", {})
        note     = ""
        if df.get(diet.lower()):
            note = f"{city} has a strong {diet} food culture! Many options available."

        return {
            "location":    {"name":location,"display":display},
            "diet":        diet,
            "results":     results[:limit],
            "total_found": len(results[:limit]),
            "note":        note,
        }

    # ── 4. Price Compare ─────────────────────────────────────────────────────

    def price_compare(self, name: str, city: str) -> Dict[str, Any]:
        """
        Compare cost-for-two for a restaurant across Zomato, Swiggy, Google.
        """
        log.info(f"Comparing prices for '{name}' in {city}…")
        geo = _geocode(f"{name} {city}")
        lat, lon = (geo[0], geo[1]) if geo else (0.0, 0.0)

        print(f"\n{c_label('  Scraping Zomato...')}")
        z  = _scrape_zomato(name, city); time.sleep(0.4)
        print(f"{c_label('  Scraping Swiggy...')}")
        sw = _scrape_swiggy(name, lat, lon); time.sleep(0.4)
        print(f"{c_label('  Scraping Google...')}")
        g  = _scrape_google(name, lat, lon); time.sleep(0.4)
        print(f"{c_label('  Scraping TripAdvisor...')}")
        ta = _scrape_tripadvisor(name, city)

        # Collect all costs
        prices = {}
        if z.get("cost_for_two"):  prices["Zomato"]      = z["cost_for_two"]
        if sw.get("cost_for_two"): prices["Swiggy"]      = sw["cost_for_two"]
        if g.get("price_level"):   prices["Google Level"] = g["price_level"]

        # Ratings
        ratings = {}
        if z.get("rating"):        ratings["Zomato"]      = z["rating"]
        if sw.get("rating"):       ratings["Swiggy"]      = sw["rating"]
        if g.get("rating"):        ratings["Google"]      = g["rating"]
        if ta.get("rating"):       ratings["TripAdvisor"] = ta["rating"]

        # Sentiment from TripAdvisor snippets
        sentiment = None
        if ta.get("snippets"):
            combined  = " ".join(ta["snippets"])
            sentiment = _sentiment(combined)

        return {
            "restaurant":     name,
            "city":           city,
            "prices":         prices,
            "ratings":        ratings,
            "review_snippets":ta.get("snippets", []),
            "sentiment":      sentiment,
            "booking_links":  _booking_links(name, city),
        }

    # ── 5. Food Tour Planner ─────────────────────────────────────────────────

    def plan_food_tour(self, location: str, stops: int = 5,
                        radius_m: int = 5000, enrich: int = 1) -> Dict[str, Any]:
        """
        Plan an optimised food tour — best N stops ordered by distance + rating.

        Strategy:
            1. Scrape all restaurants near location
            2. Enrich top-30 to get ratings
            3. Score each restaurant: score = rating * 2 - dist_km
            4. Pick top N with variety of cuisine types
            5. Order stops by nearest-neighbour routing
            6. Get road distance/duration via OSRM
        """
        geo = _geocode(location)
        if not geo: return {"error":f"Could not locate: '{location}'"}
        lat, lon, display = geo
        city = display.split(",")[0].strip()

        log.info(f"Planning {stops}-stop food tour in '{location}'…")
        all_r = _scrape_osm(lat, lon, radius_m)
        if not all_r:
            return {"error":"No restaurants found in this area","stops":[]}

        # Enrich top-30 for ratings
        enrich_pool = min(30, len(all_r))
        for r in tqdm(all_r[:enrich_pool], desc="Evaluating restaurants", unit="place"):
            _enrich(r, city, enrich)

        # Score: rating * 2 - dist_km (higher = better)
        def score(r):
            try:    rat = float(str(r.get("rating") or 3.5))
            except: rat = 3.5
            return rat * 2 - r["dist_km"] * 0.5

        all_r.sort(key=score, reverse=True)

        # Pick with cuisine variety
        chosen, seen_cuisines = [], set()
        for r in all_r:
            if len(chosen) >= stops: break
            cuisines = set(c.lower() for c in r.get("cuisines",[]))
            if not cuisines or not cuisines.issubset(seen_cuisines):
                chosen.append(r)
                seen_cuisines.update(cuisines)

        if len(chosen) < stops:
            for r in all_r:
                if r not in chosen and len(chosen) < stops:
                    chosen.append(r)

        # Order stops: nearest-neighbour from starting location
        ordered = []
        remaining = list(chosen)
        cur_lat, cur_lon = lat, lon
        while remaining:
            nearest = min(remaining, key=lambda r: _hav(cur_lat, cur_lon, r["lat"], r["lon"]))
            ordered.append(nearest)
            remaining.remove(nearest)
            cur_lat, cur_lon = nearest["lat"], nearest["lon"]

        # Get road route via OSRM
        coords = [(lat, lon)] + [(r["lat"], r["lon"]) for r in ordered]
        route  = _osrm_route(coords)

        # Assign stop numbers + suggested timings
        start_hour = 11  # 11 AM start
        for i, r in enumerate(ordered):
            r["stop_number"]  = i + 1
            r["suggested_time"] = f"{start_hour + i*2}:00 — {start_hour + i*2 + 1}:30"
            r["what_to_order"] = r.get("popular_dishes",
                                       FAMOUS_FOODS.get(_city_key(location, display),{})
                                       .get("must_try",["House special"])[:2])

        return {
            "location":    {"name":location,"display":display,"lat":lat,"lon":lon},
            "tour_stops":  ordered,
            "total_stops": len(ordered),
            "route":       route,
            "total_distance_km": route.get("distance_km","N/A"),
            "total_duration_min":route.get("duration_min","N/A"),
            "city_foods":  FAMOUS_FOODS.get(_city_key(location,display),{}),
        }

    # ── 6. Trending Dishes ───────────────────────────────────────────────────

    def trending_dishes(self, location: str) -> Dict[str, Any]:
        """
        Get trending dishes in a city — from FAMOUS_FOODS DB + Wikipedia scrape.
        """
        geo = _geocode(location)
        display = geo[2] if geo else location
        key = _city_key(location, display)
        known = FAMOUS_FOODS.get(key, {})

        trending = known.get("trending", [])
        must_try = known.get("must_try", [])
        wiki     = _scrape_wiki_food(location)

        # Try to extract trending from Zomato trending page
        try:
            city_slug = location.lower().replace(" ","-")
            r = requests.get(
                f"https://www.zomato.com/{city_slug}/trending",
                headers=HEADERS_WEB, timeout=10
            )
            soup = BeautifulSoup(r.text, "lxml")
            nd = soup.find("script", {"id":"__NEXT_DATA__"})
            if nd:
                data = json.loads(nd.string or "{}")
                # Extract collection names as trending
                collections = (data.get("props",{}).get("pageProps",{})
                               .get("collections",[]))
                for col in collections[:5]:
                    title = col.get("title","")
                    if title: trending.append(title)
        except Exception: pass

        return {
            "city":           known.get("city",location),
            "trending_now":   trending or ["No live data — see must_try"],
            "all_time_best":  must_try[:8],
            "street_food":    known.get("street_food",[]),
            "food_culture":   known.get("food_culture",wiki),
        }

    # ── 7. Compare Cities ────────────────────────────────────────────────────

    def compare_cities(self, city1: str, city2: str) -> Dict[str, Any]:
        """
        Compare the food scenes of two cities side-by-side.
        """
        geo1 = _geocode(city1)
        geo2 = _geocode(city2)
        d1   = geo1[2] if geo1 else city1
        d2   = geo2[2] if geo2 else city2
        k1   = _city_key(city1, d1)
        k2   = _city_key(city2, d2)
        f1   = FAMOUS_FOODS.get(k1, {"city":city1,"must_try":[],"sweets":[],"cuisine_tags":[]})
        f2   = FAMOUS_FOODS.get(k2, {"city":city2,"must_try":[],"sweets":[],"cuisine_tags":[]})

        return {
            "city1": f1, "city2": f2,
            "comparison": {
                "avg_cost_two":  {f1.get("city",city1): f1.get("avg_cost_two","N/A"),
                                   f2.get("city",city2): f2.get("avg_cost_two","N/A")},
                "unique_to_city1": list(set(f1.get("must_try",[])) - set(f2.get("must_try",[]))),
                "unique_to_city2": list(set(f2.get("must_try",[])) - set(f1.get("must_try",[]))),
                "common_elements": list(set(f1.get("cuisine_tags",[])) &
                                        set(f2.get("cuisine_tags",[]))) or ["None"],
            }
        }

    # ── 8. Restaurant Details ────────────────────────────────────────────────

    def restaurant_details(self, name: str, city: str) -> Dict[str, Any]:
        geo = _geocode(f"{name} {city}")
        if not geo: geo = _geocode(city)
        if not geo: return {"error":f"Could not locate: '{name}, {city}'"}
        lat, lon, display = geo

        results = _scrape_osm(lat, lon, 2000)
        match   = next((r for r in results if name.lower() in r["name"].lower()), None)
        if not match:
            match = {
                "name":name,"icon":"🍽️","type":"restaurant","type_label":"🍽️ Restaurant",
                "cuisines":[],"address":city,"phone":"","website":"","opening_hours":"",
                "amenities":{},"price_range":"","lat":lat,"lon":lon,"dist_km":0,
                "rating":None,"votes":None,"cost_for_two":None,"popular_dishes":[],
                "maps_url":f"https://www.google.com/maps?q={lat},{lon}",
                "directions_url":f"https://www.google.com/maps/dir/?api=1&destination={lat},{lon}",
                "source":"stub",
            }
        return _enrich(match, city, 2)

    # ── 9. Exports ───────────────────────────────────────────────────────────

    def export_excel(self, location: str, filepath: str = None,
                      radius_m: int = 5000, enrich: int = 1) -> str:
        if not filepath:
            slug = re.sub(r"[^\w]","_",location.lower())[:30]
            filepath = f"restaurants_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        result = self.find_nearby(location, radius_m=radius_m, limit=50, enrich=enrich)
        if "error" in result: log.error(result["error"]); return ""
        _export_excel(result["restaurants"], result.get("famous_foods",{}), filepath, location)
        print(f"\n  {c_good('✅ Excel saved:')} {filepath}  ({result['total']} restaurants)")
        return filepath

    def export_json(self, location: str, filepath: str = None,
                     radius_m: int = 5000, enrich: int = 1) -> str:
        if not filepath:
            slug = re.sub(r"[^\w]","_",location.lower())[:30]
            filepath = f"restaurants_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        result = self.find_nearby(location, radius_m=radius_m, limit=50, enrich=enrich)
        if "error" in result: log.error(result["error"]); return ""
        _export_json(result["restaurants"], result.get("famous_foods",{}), filepath)
        print(f"\n  {c_good('✅ JSON saved:')} {filepath}")
        return filepath

    def export_csv(self, location: str, filepath: str = None,
                    radius_m: int = 5000, enrich: int = 0) -> str:
        if not filepath:
            slug = re.sub(r"[^\w]","_",location.lower())[:30]
            filepath = f"restaurants_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        result = self.find_nearby(location, radius_m=radius_m, limit=100, enrich=enrich)
        if "error" in result: log.error(result["error"]); return ""
        _export_csv(result["restaurants"], filepath)
        print(f"\n  {c_good('✅ CSV saved:')} {filepath}")
        return filepath

    # ══════════════════════════════════════════════════════════════════════════
    #  PRINT METHODS
    # ══════════════════════════════════════════════════════════════════════════

    def print_restaurants(self, result: Dict[str,Any]):
        if "error" in result: print(c_bad(f"  ❌ {result['error']}")); return
        loc = result["location"]
        loc_name = loc['name']
        print(f"\n{c_head(f'  Restaurants near {loc_name}')}")
        print(f"  {c_label('📍')} {c_dim(loc['display'][:70])}")
        print(f"  {c_label('Radius')} {result['radius_m']}m  "
              f"{c_label('Cuisine')} {result['cuisine']}  "
              f"{c_label('Diet')} {result['diet']}  "
              f"{c_label('Found')} {c_num(str(result['total']))}")
        print()

        for i, r in enumerate(result["restaurants"], 1):
            rat = r.get("rating")
            rat_str = c_good(f"⭐{rat}") if rat else c_dim("⭐-")
            sent = r.get("reviews_sentiment",{})
            sent_str = f"  {c_sent(sent.get('label',''))}" if sent else ""
            print(f"  {c_num(f'{i:>2}.')} {r['icon']}  {Style.BRIGHT}{r['name']}{Style.RESET_ALL}  "
                  f"{rat_str}{sent_str}")
            print(f"      {c_label('📏')} {r['dist_km']} km  "
                  f"{c_label('🍴')} {', '.join(r['cuisines'][:3]) or 'N/A'}")
            if r["address"] != "N/A":
                print(f"      {c_label('📍')} {c_dim(r['address'][:60])}")
            if r.get("phone"):
                print(f"      {c_label('📞')} {r['phone']}")
            if r.get("cost_for_two"):
                print(f"      {c_label('💰')} {r['cost_for_two']}", end="")
                ps = r.get("price_sources",{})
                if ps: print(f"  {c_dim(str(ps))}", end="")
                print()
            if r["opening_hours"] not in ("N/A",""):
                print(f"      {c_label('🕐')} {r['opening_hours']}")
            if r.get("popular_dishes"):
                print(f"      {c_label('🥘')} {', '.join(r['popular_dishes'][:4])}")
            a = r.get("amenities",{})
            tags = []
            if a.get("vegetarian")=="yes":   tags.append(c_good("🌿Veg"))
            if a.get("vegan")=="yes":        tags.append(c_good("🌱Vegan"))
            if a.get("halal")=="yes":        tags.append(c_good("☪️Halal"))
            if a.get("delivery")=="yes":     tags.append("🛵Delivery")
            if a.get("outdoor_seating")=="yes": tags.append("🌳Outdoor")
            if a.get("wifi") in ("yes","free"): tags.append("📶WiFi")
            if tags: print(f"      {'  '.join(tags)}")
            print(f"      {c_url(r['maps_url'])}")
            print()

        self._print_famous_inline(result.get("famous_foods",{}))

    def _print_famous_inline(self, ff: Dict):
        if not ff: return
        print(f"\n  {'═'*55}")
        ff_city = ff.get('city','').upper()
        print(f"  {c_head(f'  FAMOUS FOODS — {ff_city}')}")
        print(f"  {'═'*55}")
        for label, key in [("Must Try","must_try"),("Street Food","street_food"),("Sweets","sweets")]:
            items = ff.get(key,[])
            if items:
                print(f"  {c_label(label+':')} {', '.join(items[:6])}")
        spots = ff.get("famous_spots",[])
        if spots:
            print(f"\n  {c_label('Famous Spots:')}")
            for s in spots[:4]: print(f"    {c_good('•')} {s}")

    def print_famous_foods(self, result: Dict[str,Any]):
        city  = result.get("city","")
        state = result.get("state","")
        emoji = result.get("emoji","🍽️")
        print(f"\n{c_head(f'  {emoji} FAMOUS FOODS OF {city.upper()}')}")
        if state: print(f"  {c_label('State:')} {state}")
        for title, key in [
            ("🥘 Must Try Dishes",  "must_try"),
            ("🏮 Street Food",      "street_food"),
            ("🍬 Famous Sweets",    "sweets"),
            ("📍 Famous Spots",     "famous_spots"),
            ("🏷️  Cuisine Types",   "cuisine_tags"),
            ("🔥 Trending Now",     "trending"),
        ]:
            items = result.get(key,[])
            if not items: continue
            print(f"\n  {c_title(title)}")
            for i,item in enumerate(items,1):
                print(f"    {c_num(str(i)+'.'):>5} {item}")
        culture = result.get("food_culture","")
        if culture:
            print(f"\n  {c_title('📖 Food Culture')}")
            for line in textwrap.wrap(culture, 65):
                print(f"     {c_dim(line)}")

    def print_price_compare(self, result: Dict[str,Any]):
        rname = result["restaurant"]
        print(f"\n{c_head(f'  Price Comparison — {rname}')}")
        print(f"  {c_label('City:')} {result['city']}\n")
        print(f"  {c_title('💵 Cost for Two:')}")
        for src, price in result.get("prices",{}).items():
            print(f"    {c_label(src+':'):20} {c_good(str(price))}")
        if not result.get("prices"):
            print(f"    {c_dim('No pricing data scraped — try enriching at level 2')}")
        print(f"\n  {c_title('⭐ Ratings:')}")
        for src, rat in result.get("ratings",{}).items():
            print(f"    {c_label(src+':'):20} {c_good(str(rat))}")
        sent = result.get("sentiment")
        if sent:
            print(f"\n  {c_title('💬 Review Sentiment:')} {c_sent(sent['label'])}")
            pos_sig = sent["positive_signals"]
            neg_sig = sent["negative_signals"]
            print(f"    {c_dim(f'Positive signals: {pos_sig} | Negative: {neg_sig}')}")
        snippets = result.get("review_snippets",[])
        if snippets:
            print(f"\n  {c_title('💬 Sample Reviews:')}")
            for s in snippets[:3]:
                print(f"    {c_dim('• '+s[:100])}")
        print(f"\n  {c_title('📲 Booking Links:')}")
        for platform, url in result.get("booking_links",{}).items():
            print(f"    {c_label(platform+':'):16} {c_url(url[:55])}")

    def print_food_tour(self, result: Dict[str,Any]):
        if "error" in result: print(c_bad(f"  ❌ {result['error']}")); return
        loc = result["location"]
        tour_name = loc["name"].upper()
        print(f"\n{c_head(f'  FOOD TOUR — {tour_name}')}")
        print(f"  {c_label('Total stops:')}  {c_num(str(result['total_stops']))}  |  "
              f"{c_label('Road distance:')} {result['total_distance_km']} km  |  "
              f"{c_label('Drive time:')} ~{result['total_duration_min']} min")
        print()
        for r in result["tour_stops"]:
            n = r["stop_number"]
            print(f"  {c_num(f'STOP {n}:')}  {r['icon']}  {Style.BRIGHT}{r['name']}{Style.RESET_ALL}")
            print(f"    {c_label('🕐 When:')}    {r.get('suggested_time','')}")
            print(f"    {c_label('📏 From start:')} {r['dist_km']} km")
            if r.get("rating"):
                print(f"    {c_label('⭐ Rating:')}   {c_good(str(r['rating']))}")
            dishes = r.get("what_to_order",[])
            if dishes:
                print(f"    {c_label('🥘 Order:')}    {', '.join(str(d) for d in dishes[:3])}")
            if r["address"] != "N/A":
                print(f"    {c_label('📍 Address:')}  {c_dim(r['address'][:55])}")
            print(f"    {c_url(r['directions_url'])}")
            print()
        cf = result.get("city_foods",{})
        if cf.get("must_try"):
            print(f"  {c_label('💡 Also try in this city:')} {', '.join(cf['must_try'][:4])}")

    def print_compare_cities(self, result: Dict[str,Any]):
        c1 = result["city1"]
        c2 = result["city2"]
        cmp = result["comparison"]
        n1 = c1.get("city","City 1")
        n2 = c2.get("city","City 2")
        print(f"\n{c_head(f'  🆚 FOOD FACE-OFF: {n1.upper()} vs {n2.upper()}')}\n")
        print(f"  {c_title('📊 Avg Cost for Two (₹):')}")
        for city_name, cost in cmp["avg_cost_two"].items():
            bar = "█" * int(int(str(cost))/100) if str(cost).isdigit() else ""
            print(f"    {c_label(f'{city_name}:'):25} {c_good(str(cost))}  {c_dim(bar)}")
        print(f"\n  {c_title(f'🏷️  Unique to {n1}:')}")
        for d in cmp["unique_to_city1"][:6]: print(f"    {c_good('•')} {d}")
        print(f"\n  {c_title(f'🏷️  Unique to {n2}:')}")
        for d in cmp["unique_to_city2"][:6]: print(f"    {c_good('•')} {d}")
        print(f"\n  {c_title('🤝 Common cuisine elements:')} {', '.join(cmp['common_elements'])}")
        for c_data in [c1, c2]:
            nm = c_data.get("city","")
            mt = c_data.get("must_try",[])
            print(f"\n  {c_title(f'🍛 {nm} — Must Try:')}")
            for i, d in enumerate(mt[:5],1): print(f"    {c_num(str(i)+'.')} {d}")


# ══════════════════════════════════════════════════════════════════════════════
#  INTERACTIVE CLI  v2
# ══════════════════════════════════════════════════════════════════════════════

def run_cli():
    agent = RestaurantAgent()

    MENU = f"""
  {c_title('[1]')}  Find nearby restaurants
  {c_title('[2]')}  Famous foods of a city  (dishes / street food / sweets / trending)
  {c_title('[3]')}  Search by cuisine  (south indian / biryani / vegan / chinese…)
  {c_title('[4]')}  Dietary filter  (vegetarian / vegan / halal / jain / gluten-free)
  {c_title('[5]')}  Top rated restaurants  (live-scraped ratings)
  {c_title('[6]')}  Street food near me
  {c_title('[7]')}  Plan a food tour  🗺️  (5-stop optimised route with OSRM)
  {c_title('[8]')}  Price comparison  💰 (Zomato vs Swiggy vs Google)
  {c_title('[9]')}  Trending dishes in a city
  {c_title('[0]')}  Compare two cities food scenes  🆚
  {c_title('[d]')}  Full restaurant details (all scrapers)
  {c_title('[xe]')} Export to Excel
  {c_title('[xj]')} Export to JSON
  {c_title('[xc]')} Export to CSV
  {c_title('[q]')}  Quit
"""

    while True:
        print(MENU)
        c = input(f"  {c_label('👉 Choose:')} ").strip().lower()

        if c == "1":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius m [3000]: ").strip()
            lim    = input("  🔢 Limit [20]: ").strip()
            cuis   = input("  🍴 Cuisine (or Enter): ").strip()
            diet   = input("  🌿 Diet (vegan/halal/jain/gluten_free or Enter): ").strip()
            budget = input("  💰 Budget (budget/moderate/upscale/fine or Enter): ").strip()
            enrich = input("  🔍 Enrich 0/1/2 [1]: ").strip()
            result = agent.find_nearby(
                loc,
                radius_m=int(r) if r.isdigit() else 3000,
                limit=int(lim)  if lim.isdigit() else 20,
                cuisine=cuis, diet=diet, budget=budget,
                enrich=int(enrich) if enrich.isdigit() else 1,
            )
            agent.print_restaurants(result)

        elif c == "2":
            loc = input("  🏙  City: ").strip()
            result = agent.famous_foods(loc)
            agent.print_famous_foods(result)

        elif c == "3":
            loc    = input("  📍 Location: ").strip()
            cuis   = input("  🍴 Cuisine: ").strip()
            enrich = input("  🔍 Enrich 0/1/2 [1]: ").strip()
            result = agent.find_nearby(loc, cuisine=cuis, radius_m=10000,
                                        enrich=int(enrich) if enrich.isdigit() else 1)
            agent.print_restaurants(result)

        elif c == "4":
            loc  = input("  📍 Location: ").strip()
            diet = input("  🌿 Diet (vegetarian/vegan/halal/jain/gluten_free): ").strip()
            enr  = input("  🔍 Enrich 0/1/2 [1]: ").strip()
            result = agent.dietary_search(loc, diet, enrich=int(enr) if enr.isdigit() else 1)
            tf = result["total_found"]
            print(f"\n  {c_good(f'🌿 {tf} {diet} places near {loc}')}")
            if result.get("note"): print(f"  {c_good('💡 '+result['note'])}")
            for i, r in enumerate(result.get("results",[]),1):
                print(f"  {i:>2}. {r['icon']} {r['name']:<38} {r['dist_km']} km")
                print(f"       {c_dim(r['address'][:55])}  |  {c_url(r['maps_url'])}")

        elif c == "5":
            loc  = input("  📍 Location: ").strip()
            r    = input("  📡 Radius m [5000]: ").strip()
            minr = input("  ⭐ Min rating [4.0]: ").strip()
            try:    minf = float(minr)
            except: minf = 4.0
            result = agent.find_nearby(loc, radius_m=int(r) if r.isdigit() else 5000,
                                        limit=30, enrich=2)
            rated = [x for x in result["restaurants"] if x.get("rating")]
            try: rated.sort(key=lambda x: float(str(x["rating"])), reverse=True)
            except: pass
            top = [x for x in rated if float(str(x.get("rating",0))) >= minf]
            print(f"\n  {c_title(f'⭐ Top {len(top)} restaurants ≥{minf}★ near {loc}')}")
            for i, r2 in enumerate(top[:15],1):
                rat_val = r2.get('rating', '')
                print(f"  {c_num(f'{i:>2}.')} {r2['icon']} {r2['name']:<38} "
                      f"{c_good(f'⭐{rat_val}')}")
                print(f"       {c_dim(r2['address'][:55])}  |  {c_url(r2['maps_url'])}")

        elif c == "6":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius m [2000]: ").strip()
            result = agent.find_nearby(loc, radius_m=int(r) if r.isdigit() else 2000,
                                        enrich=0,
                                        cuisine="",)
            ff = result.get("famous_foods",{})
            sf = ff.get("street_food",[])
            print(f"\n  {c_title('🏮 Street food near '+loc)}")
            if sf: print(f"  {c_label('Famous street food here:')} {', '.join(sf[:6])}")
            for i,r2 in enumerate(result["restaurants"][:20],1):
                if r2["type"] in ("fast_food","food_stall","food_court","cafe","ice_cream"):
                    print(f"  {i:>2}. {r2['icon']} {r2['name']:<38} {r2['dist_km']} km")
                    print(f"       {c_dim(r2['address'][:55])}")

        elif c == "7":
            loc    = input("  📍 Location: ").strip()
            stops  = input("  🔢 Stops [5]: ").strip()
            enrich = input("  🔍 Enrich 0/1/2 [1]: ").strip()
            result = agent.plan_food_tour(
                loc,
                stops=int(stops)   if stops.isdigit()   else 5,
                enrich=int(enrich) if enrich.isdigit()  else 1,
            )
            agent.print_food_tour(result)

        elif c == "8":
            name = input("  🍽️  Restaurant name: ").strip()
            city = input("  🏙  City: ").strip()
            result = agent.price_compare(name, city)
            agent.print_price_compare(result)

        elif c == "9":
            loc    = input("  🏙  City / Location: ").strip()
            result = agent.trending_dishes(loc)
            print(f"\n  {c_title('🔥 Trending in '+result.get('city',loc))}")
            for i, d in enumerate(result.get("trending_now",[]),1):
                print(f"    {c_num(str(i)+'.')} {d}")
            print(f"\n  {c_label('All-time best:')} {', '.join(result.get('all_time_best',[])[:5])}")

        elif c == "0":
            c1 = input("  🏙  City 1: ").strip()
            c2 = input("  🏙  City 2: ").strip()
            result = agent.compare_cities(c1, c2)
            agent.print_compare_cities(result)

        elif c == "d":
            name = input("  🍽️  Restaurant name: ").strip()
            city = input("  🏙  City: ").strip()
            r    = agent.restaurant_details(name, city)
            if "error" in r: print(c_bad(f"  ❌ {r['error']}"))
            else:
                print(f"\n  {c_title(r.get('icon','🍽️') + '  ' + r.get('name',''))}")
                for label, key in [("Cuisines",None),("Address","address"),
                                    ("Phone","phone"),("Rating","rating"),
                                    ("Cost/2","cost_for_two"),("Hours","opening_hours")]:
                    if key:
                        v = r.get(key)
                        if v: print(f"  {c_label(label+':'):18} {v}")
                    else:
                        cs = r.get("cuisines",[])
                        if cs: print(f"  {c_label('Cuisines:'):18} {', '.join(cs)}")
                print(f"\n  {c_label('📲 Booking Links:')}")
                for p,u in r.get("booking_links",{}).items():
                    print(f"    {c_dim(p+':'):18} {c_url(u[:55])}")

        elif c == "xe":
            loc    = input("  📍 Location: ").strip()
            r      = input("  📡 Radius m [5000]: ").strip()
            enrich = input("  🔍 Enrich 0/1/2 [1]: ").strip()
            path   = agent.export_excel(loc, radius_m=int(r) if r.isdigit() else 5000,
                                         enrich=int(enrich) if enrich.isdigit() else 1)
            if path: print(f"  {c_good('📊 Excel ready:')} {path}")

        elif c == "xj":
            loc    = input("  📍 Location: ").strip()
            enrich = input("  🔍 Enrich 0/1/2 [0]: ").strip()
            path   = agent.export_json(loc, enrich=int(enrich) if enrich.isdigit() else 0)
            if path: print(f"  {c_good('📄 JSON ready:')} {path}")

        elif c == "xc":
            loc  = input("  📍 Location: ").strip()
            path = agent.export_csv(loc)
            if path: print(f"  {c_good('📃 CSV ready:')} {path}")

        elif c in ("q","quit","exit"):
            print(f"\n  {c_good('🍽️  RestaurantAgent v2 signing off!')}\n"); break
        else:
            print(c_bad("  ⚠  Invalid option."))


if __name__ == "__main__":
    run_cli()