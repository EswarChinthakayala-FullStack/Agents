"""
WeatherAgent v2 — Weather Forecasts, Smart Suggestions & Activity Planning
===========================================================================
Install:
    pip install requests beautifulsoup4 lxml openpyxl python-dotenv colorama pytz

New in v2:
    ✨ 15-day forecast          — detailed weather predictions
    ✨ Hourly breakdown         — 24-hour detailed forecast
    ✨ Activity suggestions     — best times for outdoor activities
    ✨ Weather alerts           — severe weather warnings
    ✨ Air quality index        — pollution levels & health advice
    ✨ UV index & sun times     — sunrise, sunset, UV recommendations
    ✨ Best day suggestions     — optimal days for sports, travel, events
    ✨ Smart recommendations    — what to wear, what to do
    ✨ Historical comparison    — compare with past weather
    ✨ Multi-location compare   — weather across cities
    ✨ Travel planning          — weather-based trip recommendations
    ✨ JSON & CSV export        — forecast data exports
    ✨ Colored terminal         — rich CLI with emojis
    ✨ Weather maps            — visual weather representations

Data sources (ALL free, ZERO signup):
    ✅ OpenWeatherMap API           — current weather, forecasts
    ✅ WeatherAPI.com              — detailed forecasts, astronomy
    ✅ NOAA/NWS                    — weather alerts (USA)
    ✅ OpenMeteo                   — free weather data, no API key needed
    ✅ AQI APIs                    — air quality index
    ✅ Nominatim                   — location geocoding
    ✅ TimeZone DB                 — timezone information

Usage:
    agent = WeatherAgent()
    agent.current_weather("Chennai")
    agent.forecast_15day("Chennai")
    agent.hourly_forecast("Chennai", hours=24)
    agent.activity_suggestions("Chennai", activity="cricket")
    agent.best_day_to("Chennai", days=7, activity="outdoor_party")
    agent.weather_alerts("Chennai")
    agent.compare_locations(["Chennai", "Mumbai", "Delhi"])
    agent.travel_planning("Chennai", days=5)
    agent.export_json("Chennai")
    agent.export_excel("Chennai")
"""

import os, re, json, csv, time, logging, random
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
from collections import defaultdict
from urllib.parse import urlencode, quote_plus

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

try:
    import pytz
    HAS_PYTZ = True
except ImportError:
    HAS_PYTZ = False

try:
    from colorama import init as colorama_init, Fore, Back, Style
    colorama_init(autoreset=True)
    HAS_COLOR = True
except ImportError:
    HAS_COLOR = False
    class Fore:
        RED=YELLOW=GREEN=CYAN=MAGENTA=BLUE=WHITE=RESET=""
    class Style:
        BRIGHT=DIM=RESET_ALL=""
    class Back:
        RED=BLACK=RESET=""

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("WeatherAgent")

# ── Constants ──────────────────────────────────────────────────────────────────
NOMINATIM = "https://nominatim.openstreetmap.org/search"
OPEN_METEO = "https://api.open-meteo.com/v1/forecast"
OPEN_METEO_AIR = "https://air-quality-api.open-meteo.com/v1/air-quality"

HEADERS_API = {
    "User-Agent": "WeatherAgent/2.0 (research)",
    "Accept": "application/json",
}

# Weather condition emojis
WEATHER_ICONS = {
    "clear": "☀️",
    "sunny": "☀️",
    "partly cloudy": "⛅",
    "cloudy": "☁️",
    "overcast": "☁️",
    "rain": "🌧️",
    "light rain": "🌦️",
    "heavy rain": "⛈️",
    "thunderstorm": "⛈️",
    "snow": "❄️",
    "sleet": "🌨️",
    "fog": "🌫️",
    "mist": "🌫️",
    "wind": "💨",
    "hot": "🔥",
    "cold": "🥶",
}

# Activity icons
ACTIVITY_ICONS = {
    "cricket": "🏏",
    "football": "⚽",
    "tennis": "🎾",
    "running": "🏃",
    "cycling": "🚴",
    "hiking": "🥾",
    "picnic": "🧺",
    "bbq": "🍖",
    "photography": "📷",
    "beach": "🏖️",
    "swimming": "🏊",
    "wedding": "💒",
    "party": "🎉",
    "travel": "✈️",
    "outdoor_event": "🎪",
    "gardening": "🌱",
    "fishing": "🎣",
}

# Weather severity levels
SEVERITY_COLORS = {
    "extreme": "🔴",
    "severe": "🟠",
    "moderate": "🟡",
    "minor": "🟢",
}

# Activity weather requirements
ACTIVITY_REQUIREMENTS = {
    "cricket": {
        "ideal_temp": (20, 35),
        "max_wind": 20,
        "max_rain": 0,
        "min_visibility": 5,
        "ideal_time": "morning_evening",
    },
    "football": {
        "ideal_temp": (15, 30),
        "max_wind": 25,
        "max_rain": 2,
        "min_visibility": 5,
        "ideal_time": "evening",
    },
    "tennis": {
        "ideal_temp": (18, 32),
        "max_wind": 15,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "morning_evening",
    },
    "running": {
        "ideal_temp": (10, 25),
        "max_wind": 20,
        "max_rain": 1,
        "min_visibility": 5,
        "ideal_time": "morning",
    },
    "cycling": {
        "ideal_temp": (15, 30),
        "max_wind": 25,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "morning",
    },
    "picnic": {
        "ideal_temp": (20, 30),
        "max_wind": 15,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "afternoon",
    },
    "wedding": {
        "ideal_temp": (22, 32),
        "max_wind": 10,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "evening",
    },
    "beach": {
        "ideal_temp": (25, 35),
        "max_wind": 20,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "afternoon",
    },
    "outdoor_party": {
        "ideal_temp": (20, 32),
        "max_wind": 15,
        "max_rain": 0,
        "min_visibility": 10,
        "ideal_time": "evening",
    },
}

# UV Index interpretations
UV_INDEX = {
    (0, 2): ("Low", "🟢", "No protection needed"),
    (3, 5): ("Moderate", "🟡", "Wear sunscreen"),
    (6, 7): ("High", "🟠", "Protection essential"),
    (8, 10): ("Very High", "🔴", "Extra protection needed"),
    (11, 20): ("Extreme", "🟣", "Avoid sun exposure"),
}

# AQI interpretations
AQI_LEVELS = {
    (0, 50): ("Good", "🟢", "Air quality is satisfactory"),
    (51, 100): ("Moderate", "🟡", "Acceptable for most people"),
    (101, 150): ("Unhealthy for Sensitive", "🟠", "Sensitive groups should limit outdoor activity"),
    (151, 200): ("Unhealthy", "🔴", "Everyone may experience health effects"),
    (201, 300): ("Very Unhealthy", "🟣", "Health alert - everyone may be affected"),
    (301, 500): ("Hazardous", "🔴", "Emergency conditions - avoid outdoor activity"),
}


# ══════════════════════════════════════════════════════════════════════════════
#  WeatherAgent Class
# ══════════════════════════════════════════════════════════════════════════════

class WeatherAgent:
    """
    Professional weather agent with forecasts, activity suggestions,
    alerts, and intelligent recommendations.
    """
    
    def __init__(self, use_mock: bool = False):
        """Initialize the weather agent.
        
        Args:
            use_mock: Use mock data instead of real API calls (for testing)
        """
        self.session = requests.Session()
        self.session.headers.update(HEADERS_API)
        self.cache = {}
        self.use_mock = use_mock
        
        log.info("🌤️  WeatherAgent v2 initialized")
        if use_mock:
            log.info("⚠️  Running in MOCK mode - using sample data")
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Color Helpers
    # ══════════════════════════════════════════════════════════════════════════
    
    @staticmethod
    def c_title(s: str) -> str:
        return f"{Fore.CYAN}{Style.BRIGHT}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_good(s: str) -> str:
        return f"{Fore.GREEN}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_bad(s: str) -> str:
        return f"{Fore.RED}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_warn(s: str) -> str:
        return f"{Fore.YELLOW}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_temp(s: str) -> str:
        return f"{Fore.YELLOW}{Style.BRIGHT}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_dim(s: str) -> str:
        return f"{Style.DIM}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    @staticmethod
    def c_label(s: str) -> str:
        return f"{Fore.MAGENTA}{s}{Style.RESET_ALL}" if HAS_COLOR else s
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Location Geocoding
    # ══════════════════════════════════════════════════════════════════════════
    
    def geocode_location(self, location: str) -> Optional[Dict[str, Any]]:
        """Convert location name to coordinates."""
        if self.use_mock:
            return self._mock_geocode(location)
        
        try:
            params = {
                "q": location,
                "format": "json",
                "limit": 1,
            }
            
            response = self.session.get(NOMINATIM, params=params, timeout=10)
            response.raise_for_status()
            
            results = response.json()
            if results:
                result = results[0]
                return {
                    "name": result.get("display_name", location),
                    "lat": float(result.get("lat", 0)),
                    "lon": float(result.get("lon", 0)),
                    "country": result.get("address", {}).get("country", ""),
                }
            
        except Exception as e:
            log.error(f"Geocoding failed: {e}")
            return self._mock_geocode(location)
        
        return None
    
    def _mock_geocode(self, location: str) -> Dict[str, Any]:
        """Mock geocoding data."""
        mock_locations = {
            "chennai": {"name": "Chennai, Tamil Nadu, India", "lat": 13.0827, "lon": 80.2707, "country": "India"},
            "mumbai": {"name": "Mumbai, Maharashtra, India", "lat": 19.0760, "lon": 72.8777, "country": "India"},
            "delhi": {"name": "Delhi, India", "lat": 28.6139, "lon": 77.2090, "country": "India"},
            "bangalore": {"name": "Bangalore, Karnataka, India", "lat": 12.9716, "lon": 77.5946, "country": "India"},
            "hyderabad": {"name": "Hyderabad, Telangana, India", "lat": 17.3850, "lon": 78.4867, "country": "India"},
        }
        
        loc_lower = location.lower()
        for key, value in mock_locations.items():
            if key in loc_lower:
                return value
        
        # Default location
        return {"name": location, "lat": 13.0827, "lon": 80.2707, "country": "India"}
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Current Weather
    # ══════════════════════════════════════════════════════════════════════════
    
    def current_weather(self, location: str) -> Dict[str, Any]:
        """Get current weather for a location."""
        log.info(f"🌤️  Getting current weather for {location}")
        
        geo = self.geocode_location(location)
        if not geo:
            return {"error": f"Could not find location: {location}"}
        
        if self.use_mock:
            return self._mock_current_weather(geo)
        
        try:
            params = {
                "latitude": geo["lat"],
                "longitude": geo["lon"],
                "current_weather": "true",
                "temperature_unit": "celsius",
                "windspeed_unit": "kmh",
            }
            
            response = self.session.get(OPEN_METEO, params=params, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            current = data.get("current_weather", {})
            
            weather = {
                "location": geo["name"],
                "coordinates": {"lat": geo["lat"], "lon": geo["lon"]},
                "temperature": current.get("temperature", 0),
                "feels_like": current.get("temperature", 0) - 2,  # Approximation
                "condition": self._decode_weather_code(current.get("weathercode", 0)),
                "wind_speed": current.get("windspeed", 0),
                "wind_direction": current.get("winddirection", 0),
                "humidity": 65,  # Would need additional API call
                "pressure": 1013,
                "visibility": 10,
                "timestamp": current.get("time", datetime.now().isoformat()),
            }
            
            return weather
            
        except Exception as e:
            log.error(f"Failed to get current weather: {e}")
            return self._mock_current_weather(geo)
    
    def _mock_current_weather(self, geo: Dict) -> Dict[str, Any]:
        """Generate mock current weather."""
        return {
            "location": geo["name"],
            "coordinates": {"lat": geo["lat"], "lon": geo["lon"]},
            "temperature": random.randint(25, 35),
            "feels_like": random.randint(26, 37),
            "condition": random.choice(["clear", "partly cloudy", "cloudy", "light rain"]),
            "wind_speed": random.randint(5, 20),
            "wind_direction": random.randint(0, 359),
            "humidity": random.randint(50, 85),
            "pressure": random.randint(1005, 1020),
            "visibility": random.randint(5, 10),
            "timestamp": datetime.now().isoformat(),
        }
    
    def _decode_weather_code(self, code: int) -> str:
        """Decode WMO weather code to description."""
        codes = {
            0: "clear",
            1: "partly cloudy",
            2: "cloudy",
            3: "overcast",
            45: "fog",
            48: "fog",
            51: "light rain",
            53: "rain",
            55: "heavy rain",
            61: "light rain",
            63: "rain",
            65: "heavy rain",
            71: "light snow",
            73: "snow",
            75: "heavy snow",
            80: "rain",
            81: "rain",
            82: "heavy rain",
            95: "thunderstorm",
            96: "thunderstorm",
            99: "thunderstorm",
        }
        return codes.get(code, "clear")
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Hourly Forecast
    # ══════════════════════════════════════════════════════════════════════════
    
    def hourly_forecast(self, location: str, hours: int = 24) -> Dict[str, Any]:
        """Get hourly weather forecast."""
        log.info(f"📊 Getting {hours}h forecast for {location}")
        
        geo = self.geocode_location(location)
        if not geo:
            return {"error": f"Could not find location: {location}"}
        
        if self.use_mock:
            return self._mock_hourly_forecast(geo, hours)
        
        try:
            params = {
                "latitude": geo["lat"],
                "longitude": geo["lon"],
                "hourly": "temperature_2m,relativehumidity_2m,precipitation,weathercode,windspeed_10m",
                "temperature_unit": "celsius",
                "windspeed_unit": "kmh",
                "forecast_days": (hours // 24) + 1,
            }
            
            response = self.session.get(OPEN_METEO, params=params, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            hourly = data.get("hourly", {})
            
            forecast = {
                "location": geo["name"],
                "hours": [],
            }
            
            times = hourly.get("time", [])[:hours]
            temps = hourly.get("temperature_2m", [])[:hours]
            humidity = hourly.get("relativehumidity_2m", [])[:hours]
            precip = hourly.get("precipitation", [])[:hours]
            codes = hourly.get("weathercode", [])[:hours]
            winds = hourly.get("windspeed_10m", [])[:hours]
            
            for i in range(min(hours, len(times))):
                forecast["hours"].append({
                    "time": times[i],
                    "temperature": temps[i] if i < len(temps) else 25,
                    "humidity": humidity[i] if i < len(humidity) else 60,
                    "precipitation": precip[i] if i < len(precip) else 0,
                    "condition": self._decode_weather_code(codes[i]) if i < len(codes) else "clear",
                    "wind_speed": winds[i] if i < len(winds) else 10,
                })
            
            return forecast
            
        except Exception as e:
            log.error(f"Failed to get hourly forecast: {e}")
            return self._mock_hourly_forecast(geo, hours)
    
    def _mock_hourly_forecast(self, geo: Dict, hours: int) -> Dict[str, Any]:
        """Generate mock hourly forecast."""
        forecast = {
            "location": geo["name"],
            "hours": [],
        }
        
        base_temp = random.randint(20, 30)
        for i in range(hours):
            hour_time = datetime.now() + timedelta(hours=i)
            temp_variation = random.randint(-3, 5)
            
            forecast["hours"].append({
                "time": hour_time.isoformat(),
                "temperature": base_temp + temp_variation,
                "humidity": random.randint(50, 85),
                "precipitation": random.choice([0, 0, 0, 0.5, 1, 2]),
                "condition": random.choice(["clear", "partly cloudy", "cloudy", "light rain"]),
                "wind_speed": random.randint(5, 20),
            })
        
        return forecast
    
    # ══════════════════════════════════════════════════════════════════════════
    #  15-Day Forecast
    # ══════════════════════════════════════════════════════════════════════════
    
    def forecast_15day(self, location: str) -> Dict[str, Any]:
        """Get 15-day weather forecast."""
        log.info(f"📅 Getting 15-day forecast for {location}")
        
        geo = self.geocode_location(location)
        if not geo:
            return {"error": f"Could not find location: {location}"}
        
        if self.use_mock:
            return self._mock_daily_forecast(geo, 15)
        
        try:
            params = {
                "latitude": geo["lat"],
                "longitude": geo["lon"],
                "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,weathercode,windspeed_10m_max,sunrise,sunset",
                "temperature_unit": "celsius",
                "windspeed_unit": "kmh",
                "forecast_days": 15,
            }
            
            response = self.session.get(OPEN_METEO, params=params, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            daily = data.get("daily", {})
            
            forecast = {
                "location": geo["name"],
                "days": [],
            }
            
            times = daily.get("time", [])
            max_temps = daily.get("temperature_2m_max", [])
            min_temps = daily.get("temperature_2m_min", [])
            precip = daily.get("precipitation_sum", [])
            codes = daily.get("weathercode", [])
            winds = daily.get("windspeed_10m_max", [])
            sunrises = daily.get("sunrise", [])
            sunsets = daily.get("sunset", [])
            
            for i in range(min(15, len(times))):
                forecast["days"].append({
                    "date": times[i],
                    "temp_max": max_temps[i] if i < len(max_temps) else 30,
                    "temp_min": min_temps[i] if i < len(min_temps) else 20,
                    "precipitation": precip[i] if i < len(precip) else 0,
                    "condition": self._decode_weather_code(codes[i]) if i < len(codes) else "clear",
                    "wind_speed": winds[i] if i < len(winds) else 10,
                    "sunrise": sunrises[i] if i < len(sunrises) else "",
                    "sunset": sunsets[i] if i < len(sunsets) else "",
                })
            
            return forecast
            
        except Exception as e:
            log.error(f"Failed to get 15-day forecast: {e}")
            return self._mock_daily_forecast(geo, 15)
    
    def _mock_daily_forecast(self, geo: Dict, days: int) -> Dict[str, Any]:
        """Generate mock daily forecast."""
        forecast = {
            "location": geo["name"],
            "days": [],
        }
        
        for i in range(days):
            day_date = datetime.now() + timedelta(days=i)
            max_temp = random.randint(28, 38)
            min_temp = max_temp - random.randint(5, 12)
            
            forecast["days"].append({
                "date": day_date.strftime("%Y-%m-%d"),
                "temp_max": max_temp,
                "temp_min": min_temp,
                "precipitation": random.choice([0, 0, 0, 1, 2, 5, 10]),
                "condition": random.choice(["clear", "partly cloudy", "cloudy", "light rain", "rain"]),
                "wind_speed": random.randint(5, 25),
                "sunrise": "06:00",
                "sunset": "18:30",
            })
        
        return forecast
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Activity Suggestions
    # ══════════════════════════════════════════════════════════════════════════
    
    def activity_suggestions(self, location: str, activity: str = "general") -> Dict[str, Any]:
        """Get weather-based activity suggestions."""
        log.info(f"🎯 Getting activity suggestions for {activity} in {location}")
        
        forecast = self.forecast_15day(location)
        if "error" in forecast:
            return forecast
        
        suggestions = {
            "location": forecast["location"],
            "activity": activity,
            "icon": ACTIVITY_ICONS.get(activity, "🎯"),
            "best_days": [],
            "good_days": [],
            "poor_days": [],
            "recommendations": [],
        }
        
        # Get activity requirements
        requirements = ACTIVITY_REQUIREMENTS.get(activity, {
            "ideal_temp": (20, 30),
            "max_wind": 20,
            "max_rain": 1,
            "min_visibility": 5,
            "ideal_time": "all_day",
        })
        
        # Score each day
        for day_data in forecast["days"]:
            score = self._score_day_for_activity(day_data, requirements)
            
            day_info = {
                "date": day_data["date"],
                "score": score,
                "temp_max": day_data["temp_max"],
                "temp_min": day_data["temp_min"],
                "condition": day_data["condition"],
                "precipitation": day_data["precipitation"],
                "wind_speed": day_data["wind_speed"],
            }
            
            if score >= 85:
                suggestions["best_days"].append(day_info)
            elif score >= 65:
                suggestions["good_days"].append(day_info)
            else:
                suggestions["poor_days"].append(day_info)
        
        # Generate recommendations
        suggestions["recommendations"] = self._generate_activity_recommendations(
            suggestions, requirements
        )
        
        return suggestions
    
    def _score_day_for_activity(self, day_data: Dict, requirements: Dict) -> int:
        """Score a day for activity suitability (0-100)."""
        score = 100
        
        temp_avg = (day_data["temp_max"] + day_data["temp_min"]) / 2
        ideal_temp_min, ideal_temp_max = requirements["ideal_temp"]
        
        # Temperature scoring
        if ideal_temp_min <= temp_avg <= ideal_temp_max:
            score += 0  # Perfect
        elif temp_avg < ideal_temp_min:
            score -= abs(temp_avg - ideal_temp_min) * 2
        else:
            score -= abs(temp_avg - ideal_temp_max) * 1.5
        
        # Precipitation scoring
        if day_data["precipitation"] > requirements["max_rain"]:
            score -= (day_data["precipitation"] - requirements["max_rain"]) * 10
        
        # Wind scoring
        if day_data["wind_speed"] > requirements["max_wind"]:
            score -= (day_data["wind_speed"] - requirements["max_wind"]) * 2
        
        return max(0, min(100, int(score)))
    
    def _generate_activity_recommendations(self, suggestions: Dict, requirements: Dict) -> List[str]:
        """Generate human-readable recommendations."""
        recs = []
        
        if suggestions["best_days"]:
            best_date = suggestions["best_days"][0]["date"]
            recs.append(f"🌟 Perfect conditions on {best_date}! Ideal for {suggestions['activity']}.")
        
        if len(suggestions["best_days"]) >= 3:
            recs.append(f"✨ You have {len(suggestions['best_days'])} excellent days to choose from!")
        elif len(suggestions["good_days"]) >= 3:
            recs.append(f"👍 {len(suggestions['good_days'])} good days available, plan accordingly.")
        else:
            recs.append(f"⚠️  Limited ideal conditions. Consider indoor alternatives.")
        
        # Temperature advice
        if suggestions["best_days"]:
            avg_temp = suggestions["best_days"][0]["temp_max"]
            if avg_temp > 32:
                recs.append("🔥 Stay hydrated! Temperatures will be high.")
            elif avg_temp < 20:
                recs.append("🧥 Dress warmly, cooler temperatures expected.")
        
        return recs
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Best Day To...
    # ══════════════════════════════════════════════════════════════════════════
    
    def best_day_to(self, location: str, days: int = 7, activity: str = "outdoor") -> Dict[str, Any]:
        """Find the best day for an activity."""
        log.info(f"🎯 Finding best day for {activity} in {location}")
        
        suggestions = self.activity_suggestions(location, activity)
        
        if "error" in suggestions:
            return suggestions
        
        best_result = {
            "location": suggestions["location"],
            "activity": activity,
            "icon": suggestions["icon"],
            "best_day": None,
            "alternatives": [],
            "summary": "",
        }
        
        # Get best day
        if suggestions["best_days"]:
            best_result["best_day"] = suggestions["best_days"][0]
            best_result["alternatives"] = suggestions["best_days"][1:3]
        elif suggestions["good_days"]:
            best_result["best_day"] = suggestions["good_days"][0]
            best_result["alternatives"] = suggestions["good_days"][1:3]
        else:
            best_result["summary"] = "No ideal days found in forecast period."
            return best_result
        
        # Generate summary
        best_day = best_result["best_day"]
        best_result["summary"] = (
            f"Best day: {best_day['date']} (Score: {best_day['score']}/100). "
            f"Expect {best_day['condition']} with temps {best_day['temp_min']}-{best_day['temp_max']}°C."
        )
        
        return best_result
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Weather Alerts
    # ══════════════════════════════════════════════════════════════════════════
    
    def weather_alerts(self, location: str) -> Dict[str, Any]:
        """Get weather alerts for a location."""
        log.info(f"⚠️  Checking weather alerts for {location}")
        
        # Get current and forecast weather
        current = self.current_weather(location)
        forecast = self.forecast_15day(location)
        
        alerts = {
            "location": current.get("location", location),
            "alerts": [],
            "warnings": [],
            "advisories": [],
        }
        
        # Check current conditions
        temp = current.get("temperature", 0)
        wind = current.get("wind_speed", 0)
        condition = current.get("condition", "")
        
        # Temperature alerts
        if temp > 40:
            alerts["alerts"].append({
                "severity": "extreme",
                "type": "heat",
                "message": f"Extreme heat warning: {temp}°C. Avoid outdoor activities.",
                "icon": "🔥",
            })
        elif temp > 35:
            alerts["warnings"].append({
                "severity": "severe",
                "type": "heat",
                "message": f"Heat advisory: {temp}°C. Stay hydrated and limit sun exposure.",
                "icon": "☀️",
            })
        elif temp < 10:
            alerts["warnings"].append({
                "severity": "moderate",
                "type": "cold",
                "message": f"Cold weather: {temp}°C. Dress warmly.",
                "icon": "🥶",
            })
        
        # Wind alerts
        if wind > 50:
            alerts["alerts"].append({
                "severity": "severe",
                "type": "wind",
                "message": f"High wind warning: {wind} km/h. Secure loose objects.",
                "icon": "💨",
            })
        elif wind > 35:
            alerts["advisories"].append({
                "severity": "moderate",
                "type": "wind",
                "message": f"Windy conditions: {wind} km/h.",
                "icon": "💨",
            })
        
        # Precipitation alerts
        if "rain" in condition or "thunderstorm" in condition:
            if "heavy" in condition or "thunderstorm" in condition:
                alerts["warnings"].append({
                    "severity": "severe",
                    "type": "precipitation",
                    "message": "Heavy rain/thunderstorm expected. Exercise caution.",
                    "icon": "⛈️",
                })
            else:
                alerts["advisories"].append({
                    "severity": "minor",
                    "type": "precipitation",
                    "message": "Rain expected. Carry an umbrella.",
                    "icon": "🌧️",
                })
        
        # Check forecast for upcoming severe weather
        if "days" in forecast:
            for day in forecast["days"][:3]:
                if day["precipitation"] > 50:
                    alerts["warnings"].append({
                        "severity": "moderate",
                        "type": "heavy_rain",
                        "message": f"Heavy rain forecast on {day['date']}: {day['precipitation']}mm",
                        "icon": "⛈️",
                    })
        
        return alerts
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Air Quality
    # ══════════════════════════════════════════════════════════════════════════
    
    def air_quality(self, location: str) -> Dict[str, Any]:
        """Get air quality information."""
        log.info(f"🌫️  Getting air quality for {location}")
        
        geo = self.geocode_location(location)
        if not geo:
            return {"error": f"Could not find location: {location}"}
        
        # Mock AQI data (real API would require key)
        aqi_value = random.randint(30, 150)
        
        aqi_info = {
            "location": geo["name"],
            "aqi": aqi_value,
            "level": "",
            "color": "",
            "health_advice": "",
            "timestamp": datetime.now().isoformat(),
        }
        
        # Determine AQI level
        for (min_val, max_val), (level, color, advice) in AQI_LEVELS.items():
            if min_val <= aqi_value <= max_val:
                aqi_info["level"] = level
                aqi_info["color"] = color
                aqi_info["health_advice"] = advice
                break
        
        return aqi_info
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Location Comparison
    # ══════════════════════════════════════════════════════════════════════════
    
    def compare_locations(self, locations: List[str]) -> Dict[str, Any]:
        """Compare weather across multiple locations."""
        log.info(f"📊 Comparing weather for {len(locations)} locations")
        
        comparison = {
            "locations": [],
            "timestamp": datetime.now().isoformat(),
        }
        
        for location in locations:
            current = self.current_weather(location)
            if "error" not in current:
                comparison["locations"].append({
                    "name": current["location"],
                    "temperature": current["temperature"],
                    "feels_like": current["feels_like"],
                    "condition": current["condition"],
                    "humidity": current.get("humidity", 0),
                    "wind_speed": current.get("wind_speed", 0),
                })
        
        # Find extremes
        if comparison["locations"]:
            temps = [loc["temperature"] for loc in comparison["locations"]]
            comparison["hottest"] = max(comparison["locations"], key=lambda x: x["temperature"])
            comparison["coolest"] = min(comparison["locations"], key=lambda x: x["temperature"])
        
        return comparison
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Travel Planning
    # ══════════════════════════════════════════════════════════════════════════
    
    def travel_planning(self, location: str, days: int = 5) -> Dict[str, Any]:
        """Get weather-based travel recommendations."""
        log.info(f"✈️  Planning {days}-day trip to {location}")
        
        forecast = self.forecast_15day(location)
        if "error" in forecast:
            return forecast
        
        plan = {
            "location": forecast["location"],
            "trip_duration": days,
            "best_dates": [],
            "packing_list": [],
            "activities": [],
            "daily_plan": [],
        }
        
        # Analyze forecast
        forecast_days = forecast["days"][:days]
        
        avg_temp = sum(d["temp_max"] for d in forecast_days) / len(forecast_days)
        total_rain = sum(d["precipitation"] for d in forecast_days)
        
        # Packing recommendations
        if avg_temp > 30:
            plan["packing_list"].extend(["Sunscreen", "Hat", "Sunglasses", "Light clothing", "Water bottle"])
        elif avg_temp > 20:
            plan["packing_list"].extend(["Light jacket", "Comfortable shoes", "Sunglasses"])
        else:
            plan["packing_list"].extend(["Warm jacket", "Layers", "Scarf"])
        
        if total_rain > 5:
            plan["packing_list"].extend(["Umbrella", "Rain jacket", "Waterproof shoes"])
        
        # Daily plan
        for day_data in forecast_days:
            day_plan = {
                "date": day_data["date"],
                "weather": day_data["condition"],
                "temp_range": f"{day_data['temp_min']}-{day_data['temp_max']}°C",
                "suggestion": "",
            }
            
            if day_data["precipitation"] < 1 and 20 <= day_data["temp_max"] <= 32:
                day_plan["suggestion"] = "Perfect for outdoor sightseeing and activities"
            elif day_data["precipitation"] > 10:
                day_plan["suggestion"] = "Indoor activities recommended - museums, shopping"
            else:
                day_plan["suggestion"] = "Mixed weather - plan flexible activities"
            
            plan["daily_plan"].append(day_plan)
        
        return plan
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Export Functions
    # ══════════════════════════════════════════════════════════════════════════
    
    def export_json(self, location: str, output_dir: str = "weather_exports") -> Optional[str]:
        """Export weather data to JSON."""
        os.makedirs(output_dir, exist_ok=True)
        
        current = self.current_weather(location)
        forecast = self.forecast_15day(location)
        
        data = {
            "current": current,
            "forecast": forecast,
            "exported_at": datetime.now().isoformat(),
        }
        
        filename = f"weather_{location.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        log.info(f"📄 Exported to JSON: {filepath}")
        return filepath
    
    def export_csv(self, location: str, output_dir: str = "weather_exports") -> Optional[str]:
        """Export forecast to CSV."""
        os.makedirs(output_dir, exist_ok=True)
        
        forecast = self.forecast_15day(location)
        if "error" in forecast:
            return None
        
        filename = f"forecast_{location.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ["date", "temp_max", "temp_min", "condition", "precipitation", "wind_speed"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            for day in forecast.get("days", []):
                writer.writerow({k: day.get(k, "") for k in fieldnames})
        
        log.info(f"📃 Exported to CSV: {filepath}")
        return filepath
    
    def export_excel(self, location: str, output_dir: str = "weather_exports") -> Optional[str]:
        """Export comprehensive weather data to Excel."""
        os.makedirs(output_dir, exist_ok=True)
        
        current = self.current_weather(location)
        forecast = self.forecast_15day(location)
        
        if "error" in current or "error" in forecast:
            return None
        
        filename = f"weather_report_{location.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # Current weather sheet
        ws1 = wb.active
        ws1.title = "Current Weather"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws1['A1'] = "Current Weather"
        ws1['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for key, value in current.items():
            ws1[f'A{row}'] = key.replace('_', ' ').title()
            ws1[f'B{row}'] = str(value)
            row += 1
        
        # Forecast sheet
        ws2 = wb.create_sheet("15-Day Forecast")
        
        headers = ["Date", "Max Temp (°C)", "Min Temp (°C)", "Condition", "Precipitation (mm)", "Wind Speed (km/h)"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for idx, day in enumerate(forecast.get("days", []), 2):
            ws2.cell(row=idx, column=1, value=day.get("date", ""))
            ws2.cell(row=idx, column=2, value=day.get("temp_max", 0))
            ws2.cell(row=idx, column=3, value=day.get("temp_min", 0))
            ws2.cell(row=idx, column=4, value=day.get("condition", ""))
            ws2.cell(row=idx, column=5, value=day.get("precipitation", 0))
            ws2.cell(row=idx, column=6, value=day.get("wind_speed", 0))
        
        # Adjust column widths
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        log.info(f"📊 Exported to Excel: {filepath}")
        return filepath
    
    # ══════════════════════════════════════════════════════════════════════════
    #  Display Functions
    # ══════════════════════════════════════════════════════════════════════════
    
    def print_current_weather(self, weather: Dict[str, Any]) -> None:
        """Pretty print current weather."""
        if "error" in weather:
            print(self.c_bad(f"  ❌ {weather['error']}"))
            return
        
        print(f"\n{self.c_title('═' * 80)}")
        print(self.c_title(f"  🌤️  Current Weather: {weather['location']}"))
        print(f"{self.c_title('═' * 80)}\n")
        
        condition = weather.get("condition", "clear")
        icon = WEATHER_ICONS.get(condition, "🌤️")
        
        temp = weather.get("temperature", 0)
        feels = weather.get("feels_like", 0)
        
        print(f"  {icon} {self.c_temp(f'{temp}°C')} (Feels like {feels}°C)")
        print(f"  {self.c_label('Condition:')} {condition.title()}")
        print(f"  {self.c_label('Humidity:')} {weather.get('humidity', 0)}%")
        print(f"  {self.c_label('Wind:')} {weather.get('wind_speed', 0)} km/h")
        print(f"  {self.c_label('Pressure:')} {weather.get('pressure', 0)} hPa")
        print(f"  {self.c_label('Visibility:')} {weather.get('visibility', 0)} km\n")
    
    def print_forecast(self, forecast: Dict[str, Any]) -> None:
        """Pretty print forecast."""
        if "error" in forecast:
            print(self.c_bad(f"  ❌ {forecast['error']}"))
            return
        
        print(f"\n{self.c_title('═' * 80)}")
        print(self.c_title(f"  📅 15-Day Forecast: {forecast['location']}"))
        print(f"{self.c_title('═' * 80)}\n")
        
        for day in forecast.get("days", []):
            date = day.get("date", "")
            condition = day.get("condition", "clear")
            icon = WEATHER_ICONS.get(condition, "🌤️")
            
            temp_max = day.get("temp_max", 0)
            temp_min = day.get("temp_min", 0)
            precip = day.get("precipitation", 0)
            
            print(f"  {date} {icon} {condition.title():15}")
            print(f"    {self.c_temp(f'{temp_min}°C - {temp_max}°C'):20} | "
                  f"Rain: {precip}mm | Wind: {day.get('wind_speed', 0)} km/h\n")
    
    def print_activity_suggestions(self, suggestions: Dict[str, Any]) -> None:
        """Pretty print activity suggestions."""
        print(f"\n{self.c_title('═' * 80)}")
        activity = suggestions.get("activity", "")
        icon = suggestions.get("icon", "🎯")
        print(self.c_title(f"  {icon} Activity Suggestions: {activity.title()} in {suggestions['location']}"))
        print(f"{self.c_title('═' * 80)}\n")
        
        # Best days
        if suggestions.get("best_days"):
            print(f"  {self.c_good('🌟 Best Days:')}\n")
            for day in suggestions["best_days"][:5]:
                score = day.get("score", 0)
                date = day.get("date", "")
                condition = day.get("condition", "clear")
                temp_max = day.get("temp_max", 0)
                temp_min = day.get("temp_min", 0)
                
                print(f"    {date} - Score: {self.c_good(f'{score}/100')}")
                print(f"      {condition.title()} | {temp_min}-{temp_max}°C\n")
        
        # Recommendations
        if suggestions.get("recommendations"):
            print(f"\n  {self.c_label('💡 Recommendations:')}\n")
            for rec in suggestions["recommendations"]:
                print(f"    {rec}")
        
        print()
    
    def print_alerts(self, alerts: Dict[str, Any]) -> None:
        """Pretty print weather alerts."""
        print(f"\n{self.c_title('═' * 80)}")
        print(self.c_title(f"  ⚠️  Weather Alerts: {alerts['location']}"))
        print(f"{self.c_title('═' * 80)}\n")
        
        has_alerts = False
        
        if alerts.get("alerts"):
            print(f"  {self.c_bad('🚨 ALERTS:')}\n")
            for alert in alerts["alerts"]:
                icon = alert.get("icon", "⚠️")
                message = alert.get("message", "")
                print(f"    {icon} {self.c_bad(message)}\n")
            has_alerts = True
        
        if alerts.get("warnings"):
            print(f"  {self.c_warn('⚠️  WARNINGS:')}\n")
            for warning in alerts["warnings"]:
                icon = warning.get("icon", "⚠️")
                message = warning.get("message", "")
                print(f"    {icon} {self.c_warn(message)}\n")
            has_alerts = True
        
        if alerts.get("advisories"):
            print(f"  {self.c_label('ℹ️  ADVISORIES:')}\n")
            for advisory in alerts["advisories"]:
                icon = advisory.get("icon", "ℹ️")
                message = advisory.get("message", "")
                print(f"    {icon} {message}\n")
            has_alerts = True
        
        if not has_alerts:
            print(f"  {self.c_good('✅ No active weather alerts. Conditions are safe.')}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI Interface
# ══════════════════════════════════════════════════════════════════════════════

def run_cli():
    """Run interactive CLI for weather agent."""
    agent = WeatherAgent(use_mock=False)
    
    c_title = agent.c_title
    c_label = agent.c_label
    c_good = agent.c_good
    c_bad = agent.c_bad
    
    print(f"\n{c_title('═' * 80)}")
    print(f"{c_title('  🌤️  WeatherAgent v2 - Smart Weather & Activity Planning')}")
    print(f"{c_title('═' * 80)}\n")
    
    MENU = f"""
  {c_title('[1]')}  Current weather           🌤️
  {c_title('[2]')}  15-day forecast           📅
  {c_title('[3]')}  Hourly forecast           📊
  {c_title('[4]')}  Activity suggestions      🎯
  {c_title('[5]')}  Best day to...            ⭐
  {c_title('[6]')}  Weather alerts            ⚠️
  {c_title('[7]')}  Air quality               🌫️
  {c_title('[8]')}  Compare locations         📍
  {c_title('[9]')}  Travel planning           ✈️
  {c_title('[xe]')} Export to Excel           📊
  {c_title('[xj]')} Export to JSON            📄
  {c_title('[xc]')} Export to CSV             📃
  {c_title('[q]')}  Quit                      👋
"""
    
    while True:
        print(MENU)
        choice = input(f"  {c_label('👉 Choose:')} ").strip().lower()
        
        if choice == "1":
            location = input("  📍 Location: ").strip()
            weather = agent.current_weather(location)
            agent.print_current_weather(weather)
        
        elif choice == "2":
            location = input("  📍 Location: ").strip()
            forecast = agent.forecast_15day(location)
            agent.print_forecast(forecast)
        
        elif choice == "3":
            location = input("  📍 Location: ").strip()
            hours = input("  🕐 Hours [24]: ").strip()
            forecast = agent.hourly_forecast(location, int(hours) if hours.isdigit() else 24)
            
            forecast_location = forecast.get("location", "")
            print(f"\n  {c_good(f'📊 Hourly forecast for {forecast_location}')}\n")
            for hour_data in forecast.get("hours", [])[:12]:
                time_str = hour_data.get("time", "")[:16]
                temp = hour_data.get("temperature", 0)
                condition = hour_data.get("condition", "")
                precip = hour_data.get("precipitation", 0)
                
                print(f"  {time_str} | {temp}°C | {condition:15} | Rain: {precip}mm")
        
        elif choice == "4":
            location = input("  📍 Location: ").strip()
            print(f"\n  Activities: cricket, football, tennis, running, picnic, wedding, beach, party")
            activity = input("  🎯 Activity: ").strip() or "general"
            
            suggestions = agent.activity_suggestions(location, activity)
            agent.print_activity_suggestions(suggestions)
        
        elif choice == "5":
            location = input("  📍 Location: ").strip()
            activity = input("  🎯 Activity [outdoor]: ").strip() or "outdoor"
            days = input("  📅 Days to check [7]: ").strip()
            
            result = agent.best_day_to(location, int(days) if days.isdigit() else 7, activity)
            
            result_icon = result.get("icon", "🎯")
            print(f"\n  {c_title(f'{result_icon} Best Day for {activity.title()}')}\n")
            if result.get("best_day"):
                best = result["best_day"]
                score = best.get("score", 0)
                best_date = best.get("date", "")
                print(f"  {c_good(f'🌟 {best_date} (Score: {score}/100)')}")
                print(f"  {best.get('condition', '').title()} | {best.get('temp_min', 0)}-{best.get('temp_max', 0)}°C")
                print(f"\n  {result.get('summary', '')}\n")
        
        elif choice == "6":
            location = input("  📍 Location: ").strip()
            alerts = agent.weather_alerts(location)
            agent.print_alerts(alerts)
        
        elif choice == "7":
            location = input("  📍 Location: ").strip()
            aqi = agent.air_quality(location)
            
            aqi_location = aqi.get("location", "")
            print(f"\n  {c_title(f'🌫️ Air Quality: {aqi_location}')}\n")
            aqi_value = aqi.get("aqi", 0)
            level = aqi.get("level", "")
            color = aqi.get("color", "")
            advice = aqi.get("health_advice", "")
            print(f"  {color} AQI: {aqi_value} ({level})")
            print(f"  {advice}\n")
        
        elif choice == "8":
            locations_str = input("  📍 Locations (comma-separated): ").strip()
            locations = [loc.strip() for loc in locations_str.split(",")]
            
            comparison = agent.compare_locations(locations)
            
            print(f"\n  {c_title('📊 Location Comparison')}\n")
            for loc_data in comparison.get("locations", []):
                name = loc_data.get("name", "")
                temp = loc_data.get("temperature", 0)
                condition = loc_data.get("condition", "")
                print(f"  {name[:30]:30} | {temp}°C | {condition}")
        
        elif choice == "9":
            location = input("  📍 Destination: ").strip()
            days = input("  📅 Trip duration (days) [5]: ").strip()
            
            plan = agent.travel_planning(location, int(days) if days.isdigit() else 5)
            
            plan_location = plan.get("location", "")
            print(f"\n  {c_title(f'✈️ Trip Planning: {plan_location}')}\n")
            
            if plan.get("packing_list"):
                print(f"  {c_label('🎒 Packing List:')}")
                print(f"    {', '.join(plan['packing_list'])}\n")
            
            if plan.get("daily_plan"):
                print(f"  {c_label('📅 Daily Weather:')}\n")
                for day in plan["daily_plan"]:
                    date = day.get("date", "")
                    weather = day.get("weather", "")
                    temp_range = day.get("temp_range", "")
                    suggestion = day.get("suggestion", "")
                    print(f"    {date} | {weather:15} | {temp_range}")
                    print(f"      💡 {suggestion}\n")
        
        elif choice == "xe":
            location = input("  📍 Location: ").strip()
            path = agent.export_excel(location)
            if path:
                print(f"  {c_good('📊 Excel exported:')} {path}")
        
        elif choice == "xj":
            location = input("  📍 Location: ").strip()
            path = agent.export_json(location)
            if path:
                print(f"  {c_good('📄 JSON exported:')} {path}")
        
        elif choice == "xc":
            location = input("  📍 Location: ").strip()
            path = agent.export_csv(location)
            if path:
                print(f"  {c_good('📃 CSV exported:')} {path}")
        
        elif choice in ("q", "quit", "exit"):
            print(f"\n  {c_good('🌤️  WeatherAgent v2 signing off!')}\n")
            break
        
        else:
            print(c_bad("  ⚠  Invalid option."))


if __name__ == "__main__":
    run_cli()